"""API: POST /api/invoice/upload-packing-list."""
from __future__ import annotations

import io
import unittest
from unittest.mock import AsyncMock, patch

import importlib.util

from openpyxl import Workbook

from tests.support_auth import login_declarant


def _sample_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "票号"
    ws["B1"] = "品名"
    ws["C1"] = "材质"
    ws["D1"] = "总数量"
    ws["A2"] = "X1"
    ws["B2"] = "塑料盒"
    ws["C2"] = "PP"
    ws["D2"] = 50
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@unittest.skipUnless(importlib.util.find_spec("fastapi"), "fastapi")
class UploadPackingListApiTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        from fastapi.testclient import TestClient
        from app.main import app

        cls.client = TestClient(app)
        login_declarant(cls.client)

    def test_upload_packing_list_meta(self) -> None:
        r = self.client.post(
            "/api/invoice/upload-packing-list",
            files={"file": ("pack.xlsx", _sample_xlsx(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"classify": "false", "max_rows": "10"},
        )
        self.assertEqual(r.status_code, 200, r.text)
        body = r.json()
        self.assertEqual(body["status"], "OK")
        self.assertIn("name", body["meta"]["columns_found"])
        self.assertEqual(body["meta"]["total_rows"], 1)
        self.assertEqual(body["results"][0]["name_cn"], "塑料盒")

    @patch("app.api.invoice.get_smart_classifier")
    def test_upload_packing_list_classify(self, mock_get_clf) -> None:
        from app.services.smart_classifier import ClassifyResult

        mock_clf = AsyncMock()
        mock_clf.classify = AsyncMock(
            return_value=ClassifyResult(
                results=[{
                    "hs_code": "3924100000",
                    "confidence": 0.85,
                    "description": "Пластиковые изделия",
                    "rationale": "PP plastic box",
                }],
                translation_used="пластиковая коробка",
                visual_analysis="На фото прозрачный контейнер",
                status="OK",
            )
        )
        mock_get_clf.return_value = mock_clf

        r = self.client.post(
            "/api/invoice/upload-packing-list",
            files={"file": ("pack.xlsx", _sample_xlsx(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"classify": "true", "max_rows": "1"},
        )
        self.assertEqual(r.status_code, 200, r.text)
        row = r.json()["results"][0]
        self.assertEqual(row["translation_used"], "пластиковая коробка")
        self.assertEqual(row["visual_analysis"], "На фото прозрачный контейнер")
        self.assertEqual(row["hs_code"], "3924100000")
