"""Тесты универсального парсера пакинг-листов."""
from __future__ import annotations

import io
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

from app.services.packing_list_parser import detect_columns, parse_packing_list


def _build_chinese_packing_xlsx(*, with_image: bool = True) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "PackingList"
    headers = ["票号", "品名", "材质", "件数", "装箱数", "总数量", "总毛重", "产品图片"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)

    rows_data = [
        ("A-001", "塑料玩具车", "ABS塑料", 5, 12, 60, 18.5),
        ("A-002", "不锈钢水杯", "304不锈钢", 3, 24, 72, 22.0),
        ("A-003", "棉质T恤", "100%棉", 10, 6, 60, 15.2),
    ]
    for i, (art, name, mat, boxes, per_box, total, gw) in enumerate(rows_data, start=2):
        ws.cell(row=i, column=1, value=art)
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=mat)
        ws.cell(row=i, column=4, value=boxes)
        ws.cell(row=i, column=5, value=per_box)
        ws.cell(row=i, column=6, value=total)
        ws.cell(row=i, column=7, value=gw)
        if with_image:
            png_buf = io.BytesIO()
            Image.new("RGB", (20, 20), color=(200, 100, 50)).save(png_buf, format="PNG")
            png_buf.seek(0)
            ws.add_image(XLImage(png_buf), f"H{i}")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_english_packing_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "SKU"
    ws["B1"] = "Product Name"
    ws["C1"] = "Material"
    ws["D1"] = "Total Qty"
    ws["A2"] = "SKU-99"
    ws["B2"] = "LED lamp"
    ws["C2"] = "Aluminum"
    ws["D2"] = 100
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


@pytest.fixture
def chinese_packing_path(tmp_path: Path) -> Path:
    p = tmp_path / "packing_cn.xlsx"
    p.write_bytes(_build_chinese_packing_xlsx())
    return p


def test_detect_columns_chinese(chinese_packing_path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(chinese_packing_path, data_only=True)
    ws = wb.active
    col_map = detect_columns(ws)
    wb.close()
    assert col_map.get("article") == 1
    assert col_map.get("name") == 2
    assert col_map.get("material") == 3
    assert col_map.get("qty_total") == 6


def test_parse_packing_list_chinese(chinese_packing_path: Path) -> None:
    rows, meta = parse_packing_list(chinese_packing_path)
    assert meta["total_rows"] == 3
    assert meta["rows_with_images"] == 3
    assert "article" in meta["columns_found"]
    assert rows[0].article == "A-001"
    assert rows[0].name_cn == "塑料玩具车"
    assert rows[0].total_qty == 60
    assert rows[0].image_base64


def test_parse_packing_list_english(tmp_path: Path) -> None:
    p = tmp_path / "packing_en.xlsx"
    p.write_bytes(_build_english_packing_xlsx())
    rows, meta = parse_packing_list(p)
    assert meta["total_rows"] == 1
    assert rows[0].article == "SKU-99"
    assert rows[0].name_cn == "LED lamp"
    assert rows[0].total_qty == 100


def test_parse_partial_columns(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Description"
    ws["B1"] = "Quantity"
    ws["A2"] = "Mystery item"
    ws["B2"] = 5
    p = tmp_path / "partial.xlsx"
    wb.save(p)
    rows, meta = parse_packing_list(p)
    assert meta["total_rows"] == 1
    assert rows[0].name_cn == "Mystery item"
    assert rows[0].total_qty == 5
