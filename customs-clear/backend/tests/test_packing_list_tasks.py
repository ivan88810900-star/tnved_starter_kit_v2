"""Тесты экспорта и async-задач пакинг-листа."""
from __future__ import annotations

import io
from pathlib import Path
from unittest.mock import AsyncMock, patch

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
from starlette.background import BackgroundTasks

from app.services.packing_list_export import export_classified_packing_list
from app.services.packing_list_tasks import create_packing_list_task, get_task


def _build_sample_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "票号"
    ws["B1"] = "品名"
    ws["C1"] = "材质"
    ws["D1"] = "总数量"
    ws["A2"] = "X1"
    ws["B2"] = "塑料盒"
    ws["C2"] = "PP"
    ws["D2"] = 50
    png = io.BytesIO()
    Image.new("RGB", (16, 16), color=(100, 150, 200)).save(png, format="PNG")
    png.seek(0)
    ws.add_image(XLImage(png), "E2")
    wb.save(path)


def test_export_inserts_columns_after_name(tmp_path: Path) -> None:
    src = tmp_path / "pack.xlsx"
    _build_sample_xlsx(src)
    out = tmp_path / "classified_pack.xlsx"
    export_classified_packing_list(
        src,
        results_by_row={
            2: {
                "translation_used": "пластиковая коробка",
                "hs_code": "3924100000",
                "hs_confidence": 0.85,
                "hs_rationale": "Изделие из PP",
            }
        },
        col_map={"name": 2, "article": 1},
        data_start_row=2,
        output_path=out,
    )
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(row=1, column=3).value == "Перевод (RU)"
    assert ws.cell(row=1, column=4).value == "Код ТН ВЭД"
    assert ws.cell(row=2, column=3).value == "пластиковая коробка"
    assert ws.cell(row=2, column=4).value == "3924100000"
    assert ws.cell(row=2, column=5).value == 85.0
    wb.close()


@pytest.mark.asyncio
async def test_create_task_sync_parse_only(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("PACKING_TASK_DIR", str(tmp_path / "tasks"))
    src = tmp_path / "pack.xlsx"
    _build_sample_xlsx(src)
    data = src.read_bytes()
    bg = BackgroundTasks()
    result = await create_packing_list_task(
        file_bytes=data,
        original_filename="pack.xlsx",
        background_tasks=bg,
        classify=False,
    )
    assert result["status"] == "done"
    assert result["total_rows"] == 1
    assert result["results"][0]["name_cn"] == "塑料盒"


@pytest.mark.asyncio
async def test_create_task_async_classify(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("PACKING_TASK_DIR", str(tmp_path / "tasks"))
    src = tmp_path / "pack.xlsx"
    _build_sample_xlsx(src)

    from app.services.smart_classifier import ClassifyResult, SmartClassifier

    mock_clf = AsyncMock()
    mock_clf.prepare_translations = AsyncMock(return_value={})
    mock_clf.translate_cached = AsyncMock(return_value="коробка")
    mock_clf.get_or_analyze_vision = AsyncMock(return_value="фото")
    mock_clf.get_or_classify_group = AsyncMock(
        return_value=ClassifyResult(
            results=[{"hs_code": "3924100000", "confidence": 0.9, "description": "x", "rationale": "y"}],
            translation_used="коробка",
            visual_analysis="фото",
            status="OK",
        )
    )

    bg = BackgroundTasks()
    with patch("app.services.packing_list_classify.get_smart_classifier", return_value=mock_clf):
        result = await create_packing_list_task(
            file_bytes=src.read_bytes(),
            original_filename="pack.xlsx",
            background_tasks=bg,
            classify=True,
        )
        assert result["status"] == "processing"
        task_id = result["task_id"]
        await bg()

    task = await get_task(task_id)
    assert task is not None
    assert task["status"] == "done"
    assert task["processed"] == 1
    assert task["results"][0]["hs_code"] == "3924100000"
    assert Path(task["export_path"]).is_file()
