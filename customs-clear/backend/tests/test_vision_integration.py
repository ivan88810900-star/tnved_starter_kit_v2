from __future__ import annotations

import io
import json
import sys
import types

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image


def _build_xlsx_with_embedded_image() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Packing"
    ws["A1"] = "name"
    ws["A2"] = "Smartphone device"

    # Create in-memory image and attach to row 2.
    png_buf = io.BytesIO()
    Image.new("RGB", (24, 24), color=(30, 144, 255)).save(png_buf, format="PNG")
    png_buf.seek(0)
    xl_img = XLImage(png_buf)
    ws.add_image(xl_img, "B2")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_xlsx_without_image() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Packing"
    ws["A1"] = "name"
    ws["A2"] = "Simple product without photos"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_conflict_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Packing"
    ws["A1"] = "name"
    ws["A2"] = "Plastic Toy Boat"

    # Synthetic "steel bearing-like" visual (metallic circle/ring).
    png_buf = io.BytesIO()
    img = Image.new("RGB", (64, 64), color=(180, 180, 180))
    for x in range(64):
        for y in range(64):
            dx = x - 32
            dy = y - 32
            r2 = dx * dx + dy * dy
            if 18 * 18 < r2 < 28 * 28:
                img.putpixel((x, y), (110, 110, 110))
            elif r2 <= 10 * 10:
                img.putpixel((x, y), (220, 220, 220))
    img.save(png_buf, format="PNG")
    png_buf.seek(0)
    ws.add_image(XLImage(png_buf), "B2")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _install_fake_genai(monkeypatch, dia, mode: str) -> None:
    class _Resp:
        def __init__(self, text: str):
            self.text = text

    class _FakeModel:
        def __init__(self, *_args, **_kwargs):
            pass

        def generate_content(self, parts, generation_config=None):  # noqa: ARG002
            if isinstance(parts, list) and parts and isinstance(parts[0], str) and "табличные данные из Excel" in parts[0]:
                text_block = parts[0]
                if mode == "conflict" and "Plastic Toy Boat" in text_block:
                    payload = {
                        "items": [
                            {
                                "name": "Steel bearing unit",
                                "suggested_hs_code": "8482109000",
                                "price": 500,
                                "net_weight_kg": 0.4,
                                "currency": "USD",
                            }
                        ]
                    }
                else:
                    payload = {
                        "items": [
                            {
                                "name": "Smartphone device" if mode == "default" else "Simple product without photos",
                                "suggested_hs_code": "8517130000" if mode == "default" else "3926909709",
                                "price": 1000,
                                "net_weight_kg": 0.2,
                                "currency": "USD",
                            }
                        ]
                    }
                return _Resp(json.dumps(payload, ensure_ascii=False))
            if mode == "conflict":
                return _Resp(
                    "Техническое описание: на фото металлический подшипник качения; текст 'Plastic Toy Boat' "
                    "не соответствует изображению, для таможенных целей приоритет у визуальных признаков."
                )
            return _Resp("Техническое описание: портативное электронное устройство связи с радиомодулями.")

    fake_genai_mod = types.ModuleType("google.generativeai")
    fake_genai_mod.GenerativeModel = _FakeModel
    fake_google_mod = types.ModuleType("google")
    fake_google_mod.generativeai = fake_genai_mod
    monkeypatch.setitem(sys.modules, "google", fake_google_mod)
    monkeypatch.setitem(sys.modules, "google.generativeai", fake_genai_mod)
    monkeypatch.setattr(dia, "configure_google_generativeai", lambda *args, **kwargs: None)
    monkeypatch.setattr(dia, "resolved_gemini_model_name", lambda: "fake-gemini-model")


def test_xlsx_vision_integration_adds_ai_visual_description(monkeypatch):
    import app.services.document_invoice_analyze as dia

    # Ensure key check passes.
    monkeypatch.setenv("GEMINI_API_KEY", "test-key")
    _install_fake_genai(monkeypatch, dia, mode="default")

    xlsx_data = _build_xlsx_with_embedded_image()

    import asyncio

    result = asyncio.run(
        dia.analyze_invoice_file(
            data=xlsx_data,
            filename="packing_with_photo.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    )

    assert result["status"] == "OK"
    assert result["items_count"] >= 1
    item0 = result["items"][0]
    assert "image_paths" in item0
    assert isinstance(item0["image_paths"], list)
    assert "ai_visual_description" in item0
    assert isinstance(item0["ai_visual_description"], str)
    assert item0["ai_visual_description"].strip() != ""


def test_xlsx_without_images_returns_empty_visual_fields(monkeypatch):
    import app.services.document_invoice_analyze as dia

    monkeypatch.setenv("GEMINI_API_KEY", "test-key")
    _install_fake_genai(monkeypatch, dia, mode="negative")

    xlsx_data = _build_xlsx_without_image()

    import asyncio

    result = asyncio.run(
        dia.analyze_invoice_file(
            data=xlsx_data,
            filename="packing_no_photo.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    )

    assert result["status"] == "OK"
    item0 = result["items"][0]
    assert item0.get("image_paths") == []
    assert item0.get("ai_visual_description") in ("", None)


def test_xlsx_conflict_text_vs_photo_prioritizes_visual(monkeypatch):
    import app.services.document_invoice_analyze as dia

    monkeypatch.setenv("GEMINI_API_KEY", "test-key")
    _install_fake_genai(monkeypatch, dia, mode="conflict")

    xlsx_data = _build_conflict_xlsx()

    import asyncio

    result = asyncio.run(
        dia.analyze_invoice_file(
            data=xlsx_data,
            filename="packing_conflict.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    )

    assert result["status"] == "OK"
    item0 = result["items"][0]
    hs = str(item0.get("suggested_hs_code") or "")
    tech = str(item0.get("technical_description") or "")
    assert hs.startswith("84"), f"Expected 84xx bearing-like code, got {hs}"
    assert "не соответствует изображению" in tech.lower()
    assert "подшипник" in tech.lower()

    # Diagnostic log for conflict resolution behavior.
    print(f"Conflict resolved -> hs={hs}; technical_description={tech}")

