#!/usr/bin/env python3
"""
Стресс-тест НДС 22% / льгота 10%: data/raw_invoices/vat_test_22.xlsx → таблица в консоль.

  cd customs-clear/backend
  python3 scripts/vat_22_stress_validate.py
  python3 scripts/vat_22_stress_validate.py --freight-usd 2500 --incoterms FOB
  python3 scripts/vat_22_stress_validate.py --freight-usd 2500 --incoterms CIF

Требуется GEMINI_API_KEY или GOOGLE_API_KEY для подбора кода и экспертизы НДС.
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))

from dotenv import load_dotenv

load_dotenv(_ROOT / ".env")
load_dotenv()


def _ensure_vat_test_xlsx(path: Path) -> None:
    if path.is_file():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    import pandas as pd

    df = pd.DataFrame(
        [
            {
                "俄语品名": "Детская коляска",
                "用途/功能": "перевозка детей, прогулочная",
                "单价": "100",
                "总数量": "1",
            },
            {
                "俄语品名": "Офисное кресло",
                "用途/功能": "офисная мебель для взрослых",
                "单价": "200",
                "总数量": "1",
            },
            {
                "俄语品名": "Медицинский тонометр",
                "用途/功能": "измерение артериального давления, бытовой",
                "单价": "50",
                "总数量": "1",
            },
            {
                "俄语品名": "Промышленный насос",
                "用途/功能": "перекачка жидкостей, стационарная установка",
                "单价": "500",
                "总数量": "1",
            },
        ]
    )
    df.to_excel(path, index=False, engine="openpyxl")


def main() -> None:
    import pandas as pd

    from app.services.currency_sync import CurrencyService
    from app.services.invoice_analyzer import (
        DEFAULT_VAT_RATE,
        VAT_IMPORT_MULTIPLIER_STANDARD,
        InvoiceAnalyzer,
        enrich_with_customs_data,
        apply_smart_net_weight_to_line_item,
        iter_item_rows,
        load_specification_table,
        map_columns,
        suggest_hs_code,
        _parse_number,
    )

    _INCOTERMS_CHOICES = (
        "EXW",
        "FCA",
        "FOB",
        "FAS",
        "CFR",
        "CIF",
        "CPT",
        "CIP",
        "DAP",
        "DPU",
        "DDP",
    )
    ap = argparse.ArgumentParser(description="Стресс-тест НДС + проверка Incoterms / фрахта")
    ap.add_argument("--freight-usd", type=float, default=0.0, help="Фрахт USD (распределение по строкам)")
    ap.add_argument(
        "--incoterms",
        type=str,
        default="EXW",
        choices=_INCOTERMS_CHOICES,
        help="Базис: FOB — фрахт в таможенной стоимости; CIF — фрахт в цене, не в базе пошлины",
    )
    args = ap.parse_args()

    raw_path = _ROOT / "data" / "raw_invoices" / "vat_test_22.xlsx"
    _ensure_vat_test_xlsx(raw_path)

    if not (os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")):
        print("Нет GEMINI_API_KEY / GOOGLE_API_KEY — пропуск вызовов Gemini.", file=sys.stderr)
        sys.exit(2)

    df, _imgs = load_specification_table(raw_path)
    mapped = map_columns(df)
    usd_rate = CurrencyService.get_usd_rate()
    mapped = InvoiceAnalyzer(freight_usd=float(args.freight_usd), incoterms=str(args.incoterms)).apply_financial_columns(
        mapped, usd_rate
    )

    rows_out: list[dict] = []
    print()
    print(
        f"Incoterms={args.incoterms}, freight_usd={args.freight_usd}, usd_rate={usd_rate:.4f} "
        f"(фрахт в руб. {float(args.freight_usd) * usd_rate:.2f})"
    )
    print()
    print(
        f"{'Товар':<28} | {'customs_value':>14} | {'alloc_fr_rub':>12} | {'НДС %':>6} | "
        f"{'vat_amount':>10} | {'unit_landed':>12} | {'Ожидание':<24}"
    )
    print("-" * 130)

    for i, item in enumerate(iter_item_rows(mapped, images_by_excel_row=None)):
        name = (item.get("name_ru") or "").strip()[:30]
        hs_payload = suggest_hs_code(item)
        hs = (hs_payload.get("hs_code") or "").strip()
        apply_smart_net_weight_to_line_item(item, hs, df=mapped, row_index=i)
        logic = (hs_payload.get("vat_logic") or "").strip().replace("\n", " ")[:58]
        pref = hs_payload.get("preferential_vat_group")
        vat_override = None
        if pref is not None:
            vf = hs_payload.get("vat_rate_final")
            try:
                if vf is not None and float(vf) in (10.0, float(DEFAULT_VAT_RATE)):
                    vat_override = float(vf)
            except (TypeError, ValueError):
                pass
        enr = enrich_with_customs_data(hs, item, vat_import_override=vat_override) if hs else {}
        vat_pct = enr.get("vat_import_rate")
        vat_s = f"{float(vat_pct):g}" if isinstance(vat_pct, (int, float)) else "—"
        cv = _parse_number(item.get("customs_value"))
        af = _parse_number(item.get("allocated_freight_rub"))
        addon = _parse_number(item.get("landed_cost_freight_addon")) or 0.0
        va = enr.get("vat_amount")
        bd = enr.get("base_duty_amount")
        qn = _parse_number(item.get("quantity")) or 1.0
        unit_landed = ""
        if (
            cv is not None
            and isinstance(bd, (int, float))
            and isinstance(va, (int, float))
            and qn > 0
        ):
            unit_landed = round((float(cv) + float(bd) + float(va) + float(addon)) / float(qn), 4)
        exp = ""
        if "коляск" in name.lower():
            exp = "ожид. 10%"
        elif "кресл" in name.lower():
            exp = "ожид. 22%"
        elif "тонометр" in name.lower():
            exp = "ожид. 10%"
        elif "насос" in name.lower():
            exp = "ожид. 22%"
        cv_s = f"{cv:g}" if cv is not None else "—"
        af_s = f"{af:g}" if af is not None else "—"
        va_s = f"{float(va):.2f}" if isinstance(va, (int, float)) else "—"
        ul_s = str(unit_landed) if unit_landed != "" else "—"
        print(f"{name:<28} | {cv_s:>14} | {af_s:>12} | {vat_s:>6} | {va_s:>10} | {ul_s:>12} | {exp:<24}")
        rows_out.append(
            {
                "name_ru": name,
                "suggested_hs_code": hs,
                "quantity": item.get("quantity") or "1",
                "duty_rate": enr.get("duty_rate") if enr.get("duty_rate") is not None else "",
                "vat_rate": vat_s,
                "item_price_rub": item.get("item_price_rub") or "",
                "allocated_freight_rub": item.get("allocated_freight_rub") or "",
                "customs_value": item.get("customs_value") or "",
                "landed_cost_freight_addon": item.get("landed_cost_freight_addon") or "",
                "incoterms": item.get("incoterms") or str(args.incoterms),
                "vat_amount": enr.get("vat_amount") if enr.get("vat_amount") is not None else "",
                "customs_value_base": enr.get("customs_value_base") if enr.get("customs_value_base") is not None else "",
                "base_duty_amount": enr.get("base_duty_amount") if enr.get("base_duty_amount") is not None else "",
                "unit_landed_cost": unit_landed if unit_landed != "" else "",
                "ОБОСНОВАНИЕ_НДС": logic,
            }
        )

    # Строка с явными числами — чтобы в Excel гарантированно появились формулы vat / duty / landed
    rows_out.append(
        {
            "name_ru": "КОНТРОЛЬ_ФОРМУЛЫ",
            "suggested_hs_code": "0000000000",
            "quantity": "1",
            "duty_rate": 5,
            "vat_rate": "22",
            "item_price_rub": 100.0,
            "allocated_freight_rub": 0.0,
            "customs_value": 100.0,
            "landed_cost_freight_addon": 0.0,
            "incoterms": str(args.incoterms),
            "vat_amount": 0,
            "customs_value_base": 100.0,
            "base_duty_amount": 5.0,
            "unit_landed_cost": 128.1,
            "ОБОСНОВАНИЕ_НДС": "(100+5)*0.22=23.1",
        }
    )

    print("-" * 130)
    print(
        f"Проверка множителя: VAT_IMPORT_MULTIPLIER_STANDARD = {VAT_IMPORT_MULTIPLIER_STANDARD} "
        f"(ставка по умолчанию {DEFAULT_VAT_RATE}%)."
    )

    out_dir = _ROOT / "data" / "processed_invoices"
    out_dir.mkdir(parents=True, exist_ok=True)
    from datetime import datetime, timezone

    from app.services.invoice_analyzer import write_invoice_report_excel

    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    out_xlsx = out_dir / f"processed_vat_test_22_{args.incoterms}_{ts}.xlsx"
    write_invoice_report_excel(pd.DataFrame(rows_out), out_xlsx, include_summary_row=True)
    print(f"\nМини-отчёт (Incoterms {args.incoterms}): {out_xlsx}")

    from openpyxl import load_workbook

    wb = load_workbook(out_xlsx, data_only=False)
    try:
        ws = wb.active
        hdr = {str(c.value).strip(): c.column for c in ws[2] if c.value}
        va_col = hdr.get("vat_amount")
        ul_col = hdr.get("unit_landed_cost")
        if va_col:
            found = None
            for r in range(3, ws.max_row + 1):
                f = ws.cell(row=r, column=va_col).value
                if isinstance(f, str) and f.startswith("=") and "0.22" in f.replace(" ", ""):
                    found = (r, f)
                    break
            if found:
                print(f"Формула vat_amount (строка {found[0]}): {found[1]}")
                print("OK: в формуле присутствует множитель 0.22.")
            else:
                print("Предупреждение: не найдена ячейка vat_amount с формулой, содержащей 0.22.", file=sys.stderr)
        if ul_col:
            for r in range(3, min(ws.max_row + 1, 8)):
                nm = ws.cell(row=r, column=hdr.get("name_ru", 1)).value
                if nm and ("насос" in str(nm).lower() or "коляск" in str(nm).lower()):
                    uf = ws.cell(row=r, column=ul_col).value
                    print(f"unit_landed_cost [{nm}]: {uf}")
                    break
    finally:
        wb.close()


if __name__ == "__main__":
    main()
