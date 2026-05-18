from __future__ import annotations

import asyncio
import io
import json
import os
import sys
import types
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.services.document_invoice_analyze import analyze_invoice_file
from app.services.export_service import generate_final_customs_excel
from app.services.payment_profile_builder import build_full_payment_profile


def _build_conflict_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Packing"
    ws["A1"] = "name"
    ws["A2"] = "Plastic Toy Boat"

    png_buf = io.BytesIO()
    img = Image.new("RGB", (64, 64), color=(180, 180, 180))
    for x in range(64):
        for y in range(64):
            dx = x - 32
            dy = y - 32
            r2 = dx * dx + dy * dy
            if 18 * 18 < r2 < 28 * 28:
                img.putpixel((x, y), (110, 110, 110))
            elif r2 <= 10 * 10:
                img.putpixel((x, y), (220, 220, 220))
    img.save(png_buf, format="PNG")
    png_buf.seek(0)
    ws.add_image(XLImage(png_buf), "B2")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _install_fake_genai() -> None:
    class _Resp:
        def __init__(self, text: str):
            self.text = text

    class _FakeModel:
        def __init__(self, *_args, **_kwargs):
            pass

        def generate_content(self, parts, generation_config=None):  # noqa: ARG002
            if isinstance(parts, list) and parts and isinstance(parts[0], str) and "табличные данные из Excel" in parts[0]:
                payload = {
                    "items": [
                        {
                            "name": "Steel bearing unit",
                            "suggested_hs_code": "8482109000",
                            "price": 500,
                            "net_weight_kg": 0.4,
                            "currency": "USD",
                        }
                    ]
                }
                return _Resp(json.dumps(payload, ensure_ascii=False))
            return _Resp(
                "Техническое описание: на фото металлический подшипник качения; "
                "текст 'Plastic Toy Boat' не соответствует изображению."
            )

    fake_genai_mod = types.ModuleType("google.generativeai")
    fake_genai_mod.GenerativeModel = _FakeModel
    fake_genai_mod.configure = lambda **_kwargs: None
    fake_google_mod = types.ModuleType("google")
    fake_google_mod.generativeai = fake_genai_mod
    sys.modules["google"] = fake_google_mod
    sys.modules["google.generativeai"] = fake_genai_mod


async def _run() -> None:
    os.environ["GEMINI_API_KEY"] = "test-key"
    _install_fake_genai()

    xlsx_data = _build_conflict_xlsx()
    analyzed = await analyze_invoice_file(
        data=xlsx_data,
        filename="packing_conflict.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    items = list(analyzed.get("items") or [])
    if not items:
        raise RuntimeError("Конфликт-тест не вернул items")

    export_rows: list[dict] = []
    for it in items:
        hs_code = str(it.get("suggested_hs_code") or "").strip()
        if not hs_code:
            continue
        payload = {
            "hs_code": hs_code,
            "customs_value": float(it.get("price") or 1000.0),
            "freight": 100.0,
            "country": "US",
            "quantity": 1.0,
            "net_weight_kg": float(it.get("net_weight_kg") or 1.0),
        }
        profile = build_full_payment_profile(
            payload=payload,
            hs_code=hs_code,
            country="US",
            item_data=it,
        )
        export_rows.append(
            {
                "item_description": str(it.get("name") or ""),
                "ai_technical_description": str(it.get("ai_visual_description") or it.get("technical_description") or ""),
                "payment_profile": profile.model_dump(),
                "duty_rate": None,
                "vat_rate": None,
            }
        )

    if not export_rows:
        raise RuntimeError("Нет строк для экспорта")

    xlsx_bytes = generate_final_customs_excel(export_rows)
    out_path = ROOT / "test_output.xlsx"
    out_path.write_bytes(xlsx_bytes)
    print(f"OK: saved {out_path}")


if __name__ == "__main__":
    asyncio.run(_run())
