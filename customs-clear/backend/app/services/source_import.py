from __future__ import annotations

import csv
import io
import json
import re
from typing import Any
from xml.etree import ElementTree as ET

from loguru import logger

from .normative_store import normalize_hs_duty_rate_string, upsert_hs_rate, upsert_source_status


def _to_float(v: Any, default: float = 0.0) -> float:
    try:
        if v is None or v == "":
            return default
        return float(str(v).replace(",", "."))
    except Exception:
        return default


def _to_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in {"1", "true", "yes", "y", "да"}


def _normalize_row(raw: dict[str, Any]) -> dict[str, Any] | None:
    row = dict(raw)
    hc_raw = row.get("hs_code")
    if hc_raw is not None and str(hc_raw).strip():
        digits = re.sub(r"\D", "", str(hc_raw))[:10]
        if len(digits) >= 4:
            row["hs_code"] = digits
            if not str(row.get("hs_prefix") or "").strip():
                row["hs_prefix"] = digits[:4]

    hs_prefix = str(row.get("hs_prefix") or row.get("hs_code") or "").strip()
    if not hs_prefix:
        return None
    return {
        "hs_code": str(row.get("hs_code") or hs_prefix).strip(),
        "hs_prefix": hs_prefix,
        "duty_rate": normalize_hs_duty_rate_string(row.get("duty_rate")),
        "vat_import_rate": _to_float(row.get("vat_import_rate"), 22.0),
        "vat_rule": str(row.get("vat_rule") or "none").strip(),
        "vat_rule_basis": str(row.get("vat_rule_basis") or "").strip(),
        "excise_type": str(row.get("excise_type") or "none").strip(),
        "excise_value": _to_float(row.get("excise_value"), 0.0),
        "excise_basis": str(row.get("excise_basis") or "").strip(),
        "has_antidumping": _to_bool(row.get("has_antidumping")),
        "antidumping_type": str(row.get("antidumping_type") or "none").strip(),
        "antidumping_value": _to_float(row.get("antidumping_value"), 0.0),
        "antidumping_condition": str(row.get("antidumping_condition") or "").strip(),
        "antidumping_countries": str(row.get("antidumping_countries") or "").strip(),
        "valid_from": str(row.get("valid_from") or "").strip(),
        "valid_to": str(row.get("valid_to") or "").strip(),
        "source_url": str(row.get("source_url") or "").strip(),
        "source_revision": str(row.get("source_revision") or "import").strip(),
    }


def _rows_from_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(r) for r in reader]


def _rows_from_json(content: bytes) -> tuple[list[dict[str, Any]], str]:
    data = json.loads(content.decode("utf-8", errors="ignore"))
    if isinstance(data, list):
        return [dict(x) for x in data if isinstance(x, dict)], "import-json"
    if isinstance(data, dict):
        rev = str(data.get("revision") or "import-json")
        rows = data.get("rows") or data.get("data") or []
        return [dict(x) for x in rows if isinstance(x, dict)], rev
    return [], "import-json"


def _rows_from_xml(content: bytes) -> list[dict[str, Any]]:
    root = ET.fromstring(content)
    rows: list[dict[str, Any]] = []
    for row in root.findall(".//row"):
        item: dict[str, Any] = {}
        for ch in row:
            item[ch.tag] = ch.text
        rows.append(item)
    return rows


def _norm_xlsx_header(cell: Any) -> str:
    t = str(cell or "").strip().lower().replace("ё", "е")
    return " ".join(t.split())


def _guess_field_from_header(h: str) -> str | None:
    if not h:
        return None
    if any(x in h for x in ("тнвэд", "тн вэд", "tnved")):
        return "hs_code"
    if h in ("код", "code") or h.startswith("код "):
        return "hs_code"
    if "hs_code" in h or h == "hs code":
        return "hs_code"
    if "ндс" in h or "vat" in h:
        return "vat_import_rate"
    if "пошлин" in h or "duty" in h or "тариф" in h or "етт" in h or "ввозн" in h:
        return "duty_rate"
    return None


def _rows_from_xlsx(content: bytes) -> list[dict[str, Any]]:
    from io import BytesIO

    from openpyxl import load_workbook

    wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        try:
            header_row = next(it)
        except StopIteration:
            return []

        headers = [_norm_xlsx_header(c) for c in header_row]
        idx_to_field: dict[int, str] = {}
        for i, h in enumerate(headers):
            field = _guess_field_from_header(h)
            if field:
                idx_to_field[i] = field

        if not idx_to_field:
            if len(headers) >= 3:
                idx_to_field = {0: "hs_code", 2: "duty_rate"}
            elif len(headers) >= 2:
                idx_to_field = {0: "hs_code", 1: "duty_rate"}
            elif len(headers) >= 1:
                idx_to_field = {0: "hs_code"}

        rows: list[dict[str, Any]] = []
        for row in it:
            if not row:
                continue
            item: dict[str, Any] = {}
            for i, field in idx_to_field.items():
                if i < len(row):
                    val = row[i]
                    if val is not None and str(val).strip() != "":
                        item[field] = val
            if item:
                rows.append(item)
        return rows
    finally:
        wb.close()


def import_normative_file(
    filename: str,
    content: bytes,
    source_code: str = "MANUAL_IMPORT",
    source_name: str = "Ручной импорт нормативных ставок",
) -> dict[str, Any]:
    name = (filename or "").lower()
    revision = "import-file"

    if name.endswith(".csv"):
        rows = _rows_from_csv(content)
    elif name.endswith(".json"):
        try:
            parsed = json.loads(content.decode("utf-8", errors="ignore"))
        except json.JSONDecodeError:
            parsed = None
        if isinstance(parsed, dict):
            from .normative_bundle import _is_bundle_payload, import_normative_bundle_dict

            if _is_bundle_payload(parsed):
                return import_normative_bundle_dict(
                    parsed,
                    filename=filename,
                    source_code=source_code,
                    source_name=source_name or "Пакет ТН ВЭД / ЕТТ / нетарифка",
                )
        rows, revision = _rows_from_json(content)
    elif name.endswith(".xml"):
        rows = _rows_from_xml(content)
    elif name.endswith((".xlsx", ".xlsm")):
        rows = _rows_from_xlsx(content)
        revision = "import-xlsx"
    else:
        raise ValueError("Поддерживаются .csv, .json, .xml, .xlsx, .xlsm")

    imported = 0
    skipped = 0
    for raw in rows:
        row = _normalize_row(raw)
        if not row:
            skipped += 1
            continue
        upsert_hs_rate(row)
        imported += 1

    note_src = "local-file" if source_code == "MANUAL_IMPORT" else filename
    upsert_source_status(
        source_code=source_code,
        source_name=source_name,
        source_url=note_src,
        revision=revision,
        is_stale=False,
        note=f"Файл: {filename}, импортировано: {imported}, пропущено: {skipped}",
    )
    logger.info(f"Normative import: file={filename}, imported={imported}, skipped={skipped}")
    return {"status": "OK", "imported": imported, "skipped": skipped, "revision": revision}

