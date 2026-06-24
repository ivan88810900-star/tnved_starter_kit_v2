"""Универсальный парсер пакинг-листов с автоопределением колонок."""

from __future__ import annotations

import base64
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .invoice_analyzer import extract_images_from_xlsx

COLUMN_ALIASES: dict[str, list[str]] = {
    "article": [
        "票号", "货号", "编号", "article", "sku", "артикул",
        "item no", "item#", "part no", "ref",
    ],
    "name": [
        "品名", "名称", "商品名称", "description", "product name",
        "наименование", "товар", "goods", "item name", "产品名称",
    ],
    "material": [
        "材质", "材料", "material", "материал", "composition",
    ],
    "qty_boxes": [
        "件数", "箱数", "cartons", "ctns", "boxes", "кол-во мест",
        "кол-во коробок", "pcs/ctn", "外箱数",
    ],
    "qty_per_box": [
        "装箱数", "每箱数量", "pcs/box", "qty/box", "штук в коробке",
        "pieces per carton", "每箱件数",
    ],
    "qty_total": [
        "总数量", "数量", "quantity", "total qty", "总件数",
        "количество", "кол-во", "total pcs", "总计",
    ],
    "weight_gross": [
        "总毛重", "毛重", "gross weight", "g.w.", "вес брутто",
        "total gross", "weight kg", "总重量",
    ],
    "weight_net": [
        "净重", "总净重", "net weight", "n.w.", "вес нетто",
    ],
    "volume": [
        "立方", "体积", "cbm", "volume", "объём", "куб", "m3",
    ],
    "value": [
        "保值", "价值", "单价", "价格", "value", "price",
        "стоимость", "цена", "amount", "总价", "金额",
    ],
    "image": [
        "图片", "照片", "photo", "image", "picture",
        "фото", "изображение", "img", "产品图片",
    ],
}

_TOTAL_ROW_MARKERS = (
    "total", "итого", "合计", "总计", "sum", "всего", "小计",
)


@dataclass
class PackingRow:
    row_num: int
    article: str | None = None
    name_cn: str | None = None
    material: str | None = None
    box_count: int | None = None
    pcs_per_box: int | None = None
    total_qty: int | None = None
    weight_gross: float | None = None
    weight_net: float | None = None
    volume_cbm: float | None = None
    value_usd: float | None = None
    image_path: Path | None = None
    image_base64: str | None = None

    def to_dict(self, *, include_image: bool = False) -> dict[str, Any]:
        out = asdict(self)
        out["image_path"] = str(self.image_path) if self.image_path else None
        out["has_image"] = bool(self.image_base64 or self.image_path)
        if not include_image:
            out.pop("image_base64", None)
        return out


def _norm_header(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip().lower())


def _header_matches(header_text: str, aliases: list[str]) -> bool:
    h = _norm_header(header_text)
    if not h:
        return False
    best_len = 0
    for alias in aliases:
        a = alias.lower().strip()
        if not a:
            continue
        if h == a or a in h:
            if len(a) > best_len:
                best_len = len(a)
    return best_len > 0


def detect_columns(ws) -> dict[str, int]:
    """
    Автоматически определяет колонки по заголовкам.
    Возвращает dict: field_name → column_number (1-based).
    Поддерживает заголовки в строках 1-3 (на случай объединённых ячеек).
    """
    detected: dict[str, int] = {}
    used_cols: set[int] = set()
    max_col = ws.max_column or 1
    header_rows: list[list[tuple[int, str]]] = []
    for row_num in range(1, 4):
        row_values: list[tuple[int, str]] = []
        for col in range(1, max_col + 1):
            val = ws.cell(row=row_num, column=col).value
            row_values.append((col, str(val or "").strip()))
        header_rows.append(row_values)

    # Более специфичные поля — раньше (длинные алиасы внутри _header_matches).
    field_order = sorted(
        (f for f in COLUMN_ALIASES if f != "image"),
        key=lambda f: max((len(a) for a in COLUMN_ALIASES[f]), default=0),
        reverse=True,
    )
    for field in field_order:
        aliases = COLUMN_ALIASES[field]
        for row_vals in header_rows:
            best_col: int | None = None
            best_score = 0
            for col_num, header_text in row_vals:
                if col_num in used_cols:
                    continue
                h = _norm_header(header_text)
                for alias in aliases:
                    a = alias.lower().strip()
                    if not a:
                        continue
                    if h == a or a in h:
                        score = len(a) + (10 if h == a else 0)
                        if score > best_score:
                            best_score = score
                            best_col = col_num
            if best_col is not None:
                detected[field] = best_col
                used_cols.add(best_col)
                break
    return detected


def find_data_start_row(ws, col_map: dict[str, int]) -> int:
    """
    Находит строку, с которой начинаются данные.
    Ищет первую строку после заголовка, где есть текст или число в ключевой колонке.
    """
    name_col = col_map.get("name") or col_map.get("article")
    max_row = ws.max_row or 1
    scan_to = min(max_row, 20)
    if not name_col:
        for row_num in range(2, scan_to + 1):
            for col in range(1, (ws.max_column or 1) + 1):
                val = ws.cell(row=row_num, column=col).value
                if val is not None and str(val).strip() and not str(val).startswith("="):
                    return row_num
        return 2

    for row_num in range(2, scan_to + 1):
        val = ws.cell(row=row_num, column=name_col).value
        if val is None:
            continue
        s = str(val).strip()
        if not s or s.startswith("="):
            continue
        if _norm_header(s) in _TOTAL_ROW_MARKERS:
            continue
        return row_num
    return 2


def _is_total_row(ws, row_num: int, col_map: dict[str, int]) -> bool:
    for col in col_map.values():
        val = ws.cell(row=row_num, column=col).value
        if val is None:
            continue
        s = _norm_header(str(val))
        if any(m in s for m in _TOTAL_ROW_MARKERS):
            return True
    return False


def _to_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _to_int(v: Any) -> int | None:
    if v is None:
        return None
    try:
        s = str(v).strip().replace(",", "").replace(" ", "")
        if not s:
            return None
        return int(float(s))
    except (TypeError, ValueError):
        return None


def _to_float(v: Any) -> float | None:
    if v is None:
        return None
    try:
        s = str(v).strip().replace(",", "").replace(" ", "")
        if not s:
            return None
        return float(s)
    except (TypeError, ValueError):
        return None


def parse_packing_list(file_path: str | Path) -> tuple[list[PackingRow], dict[str, Any]]:
    """
    Универсальный парсер пакинг-листов.

    Возвращает:
    - list[PackingRow]: строки данных
    - dict: мета-информация (detected_columns, data_start_row, etc.)
    """
    file_path = Path(file_path)
    images_by_row = extract_images_from_xlsx(file_path, any_column=True)

    wb = load_workbook(file_path, data_only=True)
    try:
        ws = wb.active
        col_map = detect_columns(ws)
        data_start = find_data_start_row(ws, col_map)

        def get(row_num: int, field: str) -> Any:
            col = col_map.get(field)
            if not col:
                return None
            val = ws.cell(row=row_num, column=col).value
            if isinstance(val, str) and val.strip().startswith("="):
                return None
            return val

        rows: list[PackingRow] = []
        for r in range(data_start, (ws.max_row or 0) + 1):
            if _is_total_row(ws, r, col_map):
                continue

            name = get(r, "name")
            article = get(r, "article")
            if not name and not article:
                continue

            total_qty = _to_int(get(r, "qty_total"))
            if total_qty is None:
                boxes = _to_int(get(r, "qty_boxes"))
                per_box = _to_int(get(r, "qty_per_box"))
                if boxes and per_box:
                    total_qty = boxes * per_box

            img_path = images_by_row.get(r)
            img_b64: str | None = None
            if img_path and img_path.exists():
                img_b64 = base64.b64encode(img_path.read_bytes()).decode()

            rows.append(
                PackingRow(
                    row_num=r,
                    article=_to_str(article),
                    name_cn=_to_str(name),
                    material=_to_str(get(r, "material")),
                    box_count=_to_int(get(r, "qty_boxes")),
                    pcs_per_box=_to_int(get(r, "qty_per_box")),
                    total_qty=total_qty,
                    weight_gross=_to_float(get(r, "weight_gross")),
                    weight_net=_to_float(get(r, "weight_net")),
                    volume_cbm=_to_float(get(r, "volume")),
                    value_usd=_to_float(get(r, "value")),
                    image_path=img_path,
                    image_base64=img_b64,
                )
            )

        meta = {
            "detected_columns": col_map,
            "data_start_row": data_start,
            "total_rows": len(rows),
            "rows_with_images": sum(1 for row in rows if row.image_base64),
            "columns_found": list(col_map.keys()),
            "columns_missing": [f for f in COLUMN_ALIASES if f not in col_map and f != "image"],
            "sheet_name": ws.title,
        }
        return rows, meta
    finally:
        wb.close()
