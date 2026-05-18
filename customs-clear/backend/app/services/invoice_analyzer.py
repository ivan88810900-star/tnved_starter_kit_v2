"""Разбор спецификаций (Excel/CSV): маппинг колонок, подбор ТН ВЭД через Gemini, обогащение из БД."""

from __future__ import annotations

import copy
import base64
import hashlib
import sqlite3
import json
import mimetypes
import os
import re
import time
import shutil
import threading
from collections import deque
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import Path
from typing import Any, Iterator, Sequence

import pandas as pd
import httpx
from loguru import logger
from pydantic import BaseModel, Field
from sqlalchemy import or_
from sqlalchemy.exc import DBAPIError, OperationalError, ProgrammingError
from sqlalchemy.orm import Session

from ..db import SessionLocal
from ..models import CountrySpecificRule, EuSanctionsList, HsRate, OfacSdnList, SanctionImportRisk

from .normative_store import (
    find_geo_duty_override_row,
    find_geo_embargo_match,
    format_applied_special_duty_label,
    get_country_risk_by_iso,
)
from ..models.tnved import Commodity, IntellectualProperty, NonTariffMeasure
from .gemini_genai_configure import (
    configure_google_generativeai,
    gemini_generate_content_rest_url,
    resolved_gemini_model_name,
)
from .trois_registry_sync import normalize_trademark_for_registry, query_trois_matches_for_trademark
from .normative_store import find_classification_precedents_for_invoice_item
from .vat_preferential_reference import match_preferential_vat_group
from .state_registry_match import format_registry_check_excel, lookup_state_registries
from .compliance_resolver import apply_compliance_resolution_to_enrichment, resolve_vat_rate_for_hs

_IMG_PATH_UNSET = object()

_REGISTRY_CHECK_EMPTY: dict[str, Any] = {
    "status": "Не найдено",
    "document_number": "—",
    "date_status": "—",
    "recommendation": "—",
    "sgr": {
        "status": "Не применимо",
        "document_number": "—",
        "date_status": "—",
        "recommendation": "—",
    },
}

_REGISTRY_DB_UNAVAILABLE_MSG = "Проверка временно недоступна (сбой БД)"


def _registry_check_db_unavailable_payload() -> dict[str, Any]:
    """Единый ответ при блокировке/сбое БД при сверке реестров (не роняем весь анализ инвойса)."""
    m = _REGISTRY_DB_UNAVAILABLE_MSG
    return {
        "status": m,
        "document_number": "—",
        "date_status": "—",
        "recommendation": m,
        "sgr": {
            "status": m,
            "document_number": "—",
            "date_status": "—",
            "recommendation": m,
        },
    }


def _compute_registry_check_payload(
    item_data: dict[str, Any],
    attrs: dict[str, str],
    hs_code: str,
    db_session: Session | None,
) -> dict[str, Any]:
    try:
        own = False
        s = db_session
        if s is None:
            s = SessionLocal()
            own = True
        try:
            return lookup_state_registries(s, item_data, attrs, hs_code=str(hs_code or ""))
        finally:
            if own and s is not None:
                s.close()
    except (
        OperationalError,
        DBAPIError,
        ProgrammingError,
        sqlite3.OperationalError,
        OSError,
    ) as e:
        logger.warning("_compute_registry_check_payload: сбой БД/файла БД (сверка реестров): {}", e)
        return _registry_check_db_unavailable_payload()
    except Exception as e:
        logger.warning("_compute_registry_check_payload: {}", e)
        return copy.deepcopy(_REGISTRY_CHECK_EMPTY)

# Базовая ставка НДС при импорте для этого проекта (отчёты, расчёты без льготы).
DEFAULT_VAT_RATE: int = 22


class HsClassificationAttributes(BaseModel):
    """Атрибуты товара в ответе классификатора (вложенный объект JSON)."""

    name: str = ""
    manufacturer: str = ""
    trademark: str = ""
    model: str = ""
    material: str = ""
    purpose_and_tech: str = ""


class HsElectronicsCompliance(BaseModel):
    """Оценка РЧЦ / ФСБ и радиочастот для электроники (главы 84, 85, 90)."""

    has_wireless_tech: bool = False
    has_encryption: bool = False
    frequencies: list[str] = Field(default_factory=list)
    rf_license_required: bool = False
    fss_notification_required: bool = False
    compliance_justification: str = ""


class HsClassificationResult(BaseModel):
    """
    Схема JSON-ответа Gemini для классификации ТН ВЭД.

    Поле ``opi_reasoning_steps`` объявлено первым: в structured output это задаёт порядок полей в
    схеме и подталкивает модель сначала выстроить логику по ОПИ 1-6, затем зафиксировать код.
    """

    opi_reasoning_steps: list[str] = Field(
        ...,
        description="Пошаговые рассуждения по ОПИ 1-6 перед выбором кода ТН ВЭД",
    )
    suggested_hs_code: str = Field(..., description="Ровно 10 цифр кода ТН ВЭД ЕАЭС")
    justification: str = Field(default="", description="Краткое обоснование по ОПИ ТН ВЭД")
    attributes: HsClassificationAttributes
    confidence_score: int = Field(..., description="Уверенность в выбранном коде, 0–100")
    compliance_warnings: list[str] = Field(default_factory=list)
    vision_insights: str = Field(default="", description="Краткий вывод по фото")
    box_31_description: str = Field(default="", description="Текст для графы 31 декларации")
    missing_info: list[str] | None = Field(
        default=None,
        description="Чего не хватает в описании для 100% уверенности; null или пустой список, если достаточно данных",
    )
    supplier_question_en: str = Field(
        default="",
        description="Точный технический вопрос поставщику на английском при нехватке данных; иначе пустая строка",
    )
    electronics_compliance: HsElectronicsCompliance | None = Field(
        default=None,
        description="Для глав 84/85/90 — заполнить; иначе null",
    )


GEMINI_ELECTRONICS_RULES = """=== ПРАВИЛА ДЛЯ ЭЛЕКТРОНИКИ (Группы 84, 85, 90) ===
Если товар попадает в эти группы (первые две цифры кода ТН ВЭД 84, 85 или 90), ты ОБЯЗАН:
1. Извлечь точное название бренда и модели из текста инвойса или с фотографии товара/шильдика.
2. Использовать встроенный Google Поиск, чтобы найти официальные технические характеристики именно этой модели
   (частоты, мощность, наличие Wi‑Fi/Bluetooth/NFC, радиомодулей). Запрещено опираться на «похожие» модели — только точное совпадение бренда и модели; если модель не найдена в открытых источниках, снизь уверенность и опиши это в missing_info и compliance_justification.
3. Сверить найденные частоты и признак шифрования с разделом «НЕТАРИФНЫЕ МЕРЫ» в контексте RAG (если он есть) и с общими правилами ввоза.
4. Если в изделии есть модули беспроводной связи или криптография, отрази необходимость нотификации ФСБ и/или заключения РЧЦ / Минпромторга в объекте electronics_compliance; учитывай типичные исключения
   (например, Bluetooth малой мощности часто не требует отдельного заключения РЧЦ, но может требовать нотификацию ФСБ при криптографии — аргументируй в compliance_justification).
5. Реестры нотификаций ФСБ и РЭС (Роскомнадзор): после ответа бэкенд сверяет бренд и модель/артикул с локальной копией реестров.
   Если точный артикул из инвойса не совпадает с моделью в строке реестра, но для того же бренда есть действующая нотификация на аналогичный тип товара (с тем же назначением: например, сетевое оборудование, антенна),
   в justification или compliance_justification явно укажи, что это возможность использовать уже существующее разрешение при подтверждении соответствия модели и сроков; не утверждай факт подтверждения без данных.
Заполни electronics_compliance строго по JSON-схеме. Если глава не 84/85/90 — верни для этого поля null."""

# Множители для суммы НДС: (таможенная стоимость + пошлина) * множитель.
VAT_IMPORT_MULTIPLIER_STANDARD: float = 0.22  # DEFAULT_VAT_RATE / 100
VAT_IMPORT_MULTIPLIER_PREFERENTIAL: float = 0.10

# Оценка нетто из брутто по префиксу ТН ВЭД (если в инвойсе нет нетто). Длиннейший совпавший префикс побеждает (8544 раньше 85).
WEIGHT_NET_COEFF_DEFAULT: float = 0.92
WEIGHT_COEFFICIENTS: dict[str, float] = {
    # Рулоны / ткани / плёнки / бумага / кабель на катушках
    "3920": 0.98,
    "3921": 0.98,
    "4801": 0.98,
    "4802": 0.98,
    "8544": 0.97,
    "50": 0.98,
    "51": 0.98,
    "52": 0.98,
    "53": 0.98,
    "54": 0.98,
    "55": 0.98,
    "56": 0.98,
    "57": 0.98,
    # Оборудование (глава 84–85; 8544 исключён — см. выше)
    "84": 0.90,
    "85": 0.90,
}


def _coefficient_for_hs_prefix(hs_code: str) -> float:
    """Коэффициент брутто→нетто по префиксу кода (2 или 4 знака)."""
    p = re.sub(r"\D", "", (hs_code or ""))[:10]
    if len(p) < 2:
        return WEIGHT_NET_COEFF_DEFAULT
    best_len = 0
    best = WEIGHT_NET_COEFF_DEFAULT
    for prefix, coeff in WEIGHT_COEFFICIENTS.items():
        if p.startswith(prefix) and len(prefix) > best_len:
            best_len = len(prefix)
            best = float(coeff)
    return best


# Инструкция для Gemini (жёсткое правило проекта).
GEMINI_PROJECT_VAT_RULES = (
    "Стандартный НДС = 22%. Льготный НДС = 10%. "
    "Твоя задача — найти основания для применения 10% (ПП РФ 908/688), иначе применяй 22%.\n\n"
)

GEMINI_GEOPOLITICAL_RULES = (
    "Ты — эксперт по санкционному и таможенному праву в контексте ввоза в РФ. При подборе кода ТН ВЭД учитывай "
    "страну происхождения товара (если она указана в данных). Если товар из США, ЕС или Великобритании, "
    "проверь описание на вхождение в типичные категории повышенных пошлин и санкционных ограничений "
    "(в т.ч. двойного назначения, военной продукции, электроники, авиационной техники). "
    "Сообщай пользователю, какой сертификат происхождения (СТ-1, форма А по правилам происхождения ЕАЭС, "
    "непреференциальный сертификат и др.) может минимизировать риски при наличии преференциального соглашения; "
    "не выдавай юридически обязывающих заключений — только ориентиры для проверки с юристом.\n\n"
)

GEMINI_EAEU_BASE_RULES = """=== БАЗОВЫЕ ПРАВИЛА ЕАЭС (КРИТИЧЕСКИ ВАЖНО) ===
ПРАВИЛО 1 (Материалы): Всегда строго различай натуральное и искусственное. Если в описании есть слова «эко», «искусственный», «поли-», «PU», «PVC», «ПВХ», «пластик», «сетка», «синтетика» — это ИСКУССТВЕННЫЕ материалы. КАТЕГОРИЧЕСКИ ЗАПРЕЩЕНО классифицировать их кодами для натуральной кожи, меха, натурального дерева или живых растений.

ПРАВИЛО 2 (Ветеринарный и Фитосанитарный контроль): Ветеринарные сертификаты требуются ТОЛЬКО для сырого мяса, живых животных и необработанных натуральных шкур/кожи. Фитосанитарные — ТОЛЬКО для необработанного дерева, семян и сырых овощей/фруктов. Для изделий из синтетики, пластика, металла, резины или обработанного текстиля эти сертификаты НЕ ТРЕБУЮТСЯ НИКОГДА. Не упоминай их.

ПРАВИЛО 3 (Сертификация): Для большинства потребительских товаров (одежда, обувь, мебель, электроника) пиши просто: «Декларация/Сертификат соответствия ТР ТС». Не выдумывай экзотические разрешения, если товар не является криптографией, оружием или спецтехникой.
"""

GEMINI_VISUAL_DEEP_ANALYSIS_BLOCK = """=== ВИЗУАЛЬНЫЙ АНАЛИЗ ===
К запросу прикреплено фото товара. Выполни технический анализ:

Шильдики: Прочитай все заводские таблички (мощность, вольты, фазы, давление).

Конструкция: Опиши внешний вид (например, для двигателя - наличие ребер охлаждения, тип крепления).
Используй эти визуальные данные для заполнения недостающих атрибутов (материал, модель, характеристики) и подбора максимально точного кода ТН ВЭД.

Поле vision_insights в JSON: одной короткой строкой на русском опиши главное, что удалось прочитать с фото (например: «Обнаружен шильдик 0.75kW, 3 фазы»). Если фото нет или нечитаемо — пустая строка.
"""


GEMINI_VISUAL_TRADEMARK_PACKING_BLOCK = """=== ВИЗУАЛЬНЫЙ АНАЛИЗ ТОВАРНОГО ЗНАКА ===
Вам предоставлена фотография товара или шильдика из упаковочного листа. Выполните глубокий анализ:

Поиск логотипов и надписей: Внимательно изучите изображение на наличие известных логотипов, названий брендов, товарных знаков. Обратите внимание на шильдики, наклейки, гравировки или тиснения.

Извлечение: Если товарный знак найден, извлеките его в JSON-матрицу атрибутов в поле trademark.

Игнорирование шума: Отсекайте общие названия, маркировки моделей или артикулы, если они не являются товарным знаком.
"""


GEMINI_HS_CLASSIFY_SYSTEM_INSTRUCTION = (
    "Ты — строгий таможенный эксперт. Твоя задача — извлечь атрибуты товара из инвойса и подобрать код ТН ВЭД. "
    "Если данных для какого-то атрибута не хватает, пиши «НЕИЗВЕСТЕН». Для товарного знака при отсутствии данных "
    "пиши «ОТСУТСТВУЕТ». Не выдумывай изготовителя, марку и материал — только факты из входных данных и фото. "
    "Если к запросу прикреплено фото, извлеки технические данные (шильдик и др.) в manufacturer, model, "
    "purpose_and_tech; видимый товарный знак/бренд с фото — в attributes.trademark (иначе ОТСУТСТВУЕТ). "
    "Не подставляй в trademark сомнительные артикулы без признаков знака. "
    "Ответ — ТОЛЬКО один валидный JSON по схеме (response_mime_type application/json): в первую очередь "
    "opi_reasoning_steps (пошаговая логика по ОПИ), затем suggested_hs_code, justification, "
    "attributes { name, manufacturer, trademark, model, material, purpose_and_tech }, confidence_score (целое 0–100), "
    "compliance_warnings, vision_insights, box_31_description, missing_info (массив строк или null), supplier_question_en. "
    "Без markdown, без текста до или после JSON.\n\n"
    "=== ОБЯЗАТЕЛЬНЫЙ АЛГОРИТМ ОПИ 1-6 (как Senior Декларант) ===\n"
    "Ты ОБЯЗАН сначала рассуждать, применяя Основные правила интерпретации ТН ВЭД (ОПИ 1-6). "
    "Ты ОБЯЗАН рассуждать пошагово: примени ОПИ 1-6, обоснуй выбор, опираясь на предоставленные прецеденты "
    "и тексты товарных позиций. Результат рассуждений помещай в opi_reasoning_steps до финального кода. "
    "В поле opi_reasoning_steps ОБЯЗАТЕЛЬНО отрази, какие ОПИ применены и почему. "
    "Минимум 4 шага, каждый шаг конкретный. Только после этого выдавай suggested_hs_code.\n"
    "Шаг 1: ОПИ 1 — анализируй тексты товарных позиций и примечания к разделам/группам из контекста. "
    "Если по ОПИ 1 код определяется однозначно — зафиксируй это.\n"
    "Шаг 2: Если товар незавершенный, разобранный, в виде заготовки или смеси — проверь ОПИ 2а/2б и опиши, "
    "как это меняет классификацию.\n"
    "Шаг 3: Если остается конкуренция позиций — применяй ОПИ 3а, 3б, 3в строго по порядку. "
    "При применении ОПИ 3б обязательно укажи признак «основного свойства» (material/function/value/role).\n"
    "Шаг 4: Для подпозиций используй ОПИ 6 (сравнение на одном уровне детализации) и выбери 10 знаков.\n"
    "Если ни одно правило не дает уверенный результат — укажи это и снизь confidence_score.\n\n"
    "=== FEW-SHOT ЭТАЛОНЫ (ОПИ 3б) ===\n"
    "Эталон 1: «Набор для ухода за обувью: крем, щетка, салфетка в одном футляре. "
    "По ОПИ 3б основной характер задает крем (функционально главный компонент и основная стоимость), "
    "поэтому набор классифицирован по позиции средства для ухода».\n"
    "Эталон 2: «Подарочный набор для кофе: кружка + 250 г кофе. "
    "По ОПИ 3б основной характер задает кофе (потребительская цель набора), "
    "поэтому классификация по товарной позиции кофе, а не по посуде».\n\n"
    "Оцени свою уверенность в коде от 0 до 100 в поле confidence_score (целое число). "
    "Если данных в описании товара недостаточно (например, нет материала, назначения, типа питания, подошвы и т.п.), "
    "снижай оценку и обязательно перечисли в missing_info, что именно нужно запросить у поставщика для точной "
    "классификации; при полной достаточности данных укажи пустой массив [] или null.\n"
    "Если описание слишком скудное для точной классификации до 10 знаков (нет вольтажа, материала, принципа действия), "
    "сгенерируй в supplier_question_en точный технический вопрос на английском для иностранного поставщика. "
    "Если данных хватает — supplier_question_en должен быть пустой строкой.\n\n"
    "=== НЕТАРИФНЫЕ МЕРЫ: СТРОГИЙ ФОРМАТ ===\n"
    "СТРОГО ЗАПРЕЩЕНО использовать размытые формулировки в предупреждениях (compliance_warnings): "
    "«по необходимости», «возможно потребуется», «при необходимости», «может требоваться» и аналогичные.\n"
    "Если мера из контекста/базы имеет условия применения (например, вольтаж, возраст, наличие шифрования, "
    "радиочастоты, назначение товара), ты обязан вывести условие явно и конкретно.\n"
    "Формат каждого предупреждения: \"[Название документа] - ЕСЛИ [Условие из примечания]\".\n"
    "Примеры: \"СГР - ЕСЛИ товар предназначен для детей\", "
    "\"Нотификация ФСБ - ЕСЛИ устройство содержит функции шифрования\".\n"
    "Если данных инвойса не хватает для точного условия, не пиши размытую фразу: добавь конкретные недостающие "
    "параметры в missing_info (какие характеристики нужно уточнить у поставщика).\n\n"
    "Если к запросу прикреплено изображение товара, внимательно изучи его. Визуальные характеристики (форма, "
    "материал верха и подошвы, наличие шнурков, длина голенища и иные наблюдаемые признаки) имеют ВЫСШИЙ приоритет "
    "при классификации. Если текст инвойса противоречит фотографии (например, в тексте «кожа», а на фото "
    "текстильная сетка), опирайся на фотографию и обязательно укажи это в поле opi_reasoning_steps.\n\n"
    "На основе фактов из инвойса и стиля текстов из блоков «ОФИЦИАЛЬНАЯ БАЗА ЗНАНИЙ» — "
    "[ОФИЦИАЛЬНЫЕ ПРЕДВАРИТЕЛЬНЫЕ РЕШЕНИЯ ТАМОЖНИ (ВЫСШИЙ ПРИОРИТЕТ)] и [ПРИМЕРЫ ИЗ ПРАКТИКИ (Реальные декларации)] — "
    "сгенерируй идеальное описание товара для графы 31 таможенной декларации. Описание должно быть строгим, "
    "профессиональным, включать все обязательные характеристики (бренд, состав, назначение) и использовать "
    "официальную терминологию (например, «текстильный материал» вместо «ткань»). Запиши результат в поле "
    "box_31_description (одна связная строка или абзац без JSON внутри).\n\n"
    "=== Экосбор и вес упаковки (брутто − нетто) ===\n"
    "При анализе веса упаковки (брутто минус нетто), если материал упаковки не указан в инвойсе явно, применяй "
    "следующие правила: 1) Если товар поставляется в рулонах (ткань, плёнка, кабель и т.п. — по тексту наименования "
    "или описания), считай материалом упаковки «полимерную пленку». 2) Для всех остальных товаров по умолчанию "
    "считай материалом упаковки «картонную коробку» (гофрированный картон). Эти правила должны согласовываться с "
    "логикой бэкенда при передаче материала в расчёт экосбора (полимерная плёнка / гофрокартон).\n\n"
    "=== ЭТАЛОННЫЙ ПРЕЦЕДЕНТ (EXACT MATCH) — САМЫЙ ВЫСОКИЙ ПРИОРИТЕТ ===\n"
    "Если в «ОФИЦИАЛЬНАЯ БАЗА ЗНАНИЙ» присутствует раздел [ЭТАЛОННЫЙ ПРЕЦЕДЕНТ (EXACT MATCH)] с непустыми строками, "
    "ЭТОТ РАЗДЕЛ ИМЕЕТ НАИВЫСШИЙ ПРИОРИТЕТ: он содержит подтверждённые человеком или таможней соответствия "
    "«описание товара -> 10-значный код ТН ВЭД». "
    "Если в строке указано AUTO_USE=TRUE и товар из инвойса (с учётом фото и атрибутов) смысл-в-смысл совпадает "
    "с этим прецедентом, ты ОБЯЗАН выдать в поле suggested_hs_code ровно тот 10-значный код, что указан в поле КОД, "
    "и в первом же шаге opi_reasoning_steps сослаться на этот прецедент с указанием source и score. "
    "Отступить от этого кода допустимо ТОЛЬКО если из инвойса/фото явно следует иной материал/функция/назначение, "
    "причём различие нужно описать словами в opi_reasoning_steps. Без такого явного различия выбирать другой код ЗАПРЕЩЕНО.\n\n"
    "Если в блоке «ОФИЦИАЛЬНАЯ БАЗА ЗНАНИЙ» есть раздел [ОФИЦИАЛЬНЫЕ ПРЕДВАРИТЕЛЬНЫЕ РЕШЕНИЯ ТАМОЖНИ (ВЫСШИЙ ПРИОРИТЕТ)] "
    "и в нём есть строка, по смыслу однозначно подходящая к описанию товара из инвойса/фото, ты ОБЯЗАН выбрать "
    "ровно тот 10-значный код ТН ВЭД, который указан в этой строке (поле «Код»), и обосновать выбор ссылкой на "
    "это решение: оно имеет высшую юридическую силу относительно иных подсказок в контексте.\n"
    "Примеры декларирования из раздела [ПРИМЕРЫ ИЗ ПРАКТИКИ] используй только как вспомогательный материал "
    "(стиль, терминология), если подходящих предварительных решений нет или ни одно из них не подходит по описанию товара.\n\n"
    "Для глав 84 и 85: если по описанию возможны нотификация ФСБ или учёт в реестре РЭС, а точная модель из инвойса "
    "может не совпасть с одной строкой реестра, всё равно отрази в обосновании сценарий «действующая нотификация на бренд "
    "для аналогичного типа изделия» как возможность использования существующего разрешения после юридической/технической "
    "проверки соответствия (без утверждения, что разрешение уже применимо).\n\n"
    "Для товаров, подлежащих санитарно-эпидемиологическому контролю, при отсутствии точного СГР, ищи в базе СГР аналогичную "
    "продукцию того же бренда или производителя. Выводи номер этого СГР как возможный вариант для таможенного оформления "
    "(отрази это в opi_reasoning_steps или justification; косметика, бытовая химия и др. — по коду и нетарифным мерам, в т.ч. "
    "Решение КТС №299). Не утверждай юридическую применимость номера без проверки декларантом.\n\n"
    "Если в системной инструкции присутствует блок «ОФИЦИАЛЬНАЯ БАЗА ЗНАНИЙ» с разделом [ПРИМЕРЫ ИЗ ПРАКТИКИ] "
    "и при этом нет подходящего предварительного решения, изучи эти примеры: используй их стиль, структуру описания "
    "и терминологию при формировании поля justification и при выборе 10-значного кода ТН ВЭД "
    "(в рамках общих правил и фактов из инвойса/фото).\n\n"
    "Если присутствует раздел [ВЕКТОРНЫЕ ПРЕЦЕДЕНТЫ (SEMANTIC SEARCH)], "
    "обязательно сопоставь товар с найденными семантически похожими кейсами, укажи это в opi_reasoning_steps "
    "и используй как аргумент при применении ОПИ 1-6 по порядку. "
    "Обоснуй код ТН ВЭД, опираясь на эти прецеденты и ОПИ 1-6.\n\n"
    "Если в блоке знаний есть раздел [ПРЕЦЕДЕНТЫ И РАЗЪЯСНЕНИЯ (СУДЫ/ЕЭК/ПОЯСНЕНИЯ ТН ВЭД)], "
    "используй его как юридически значимый аргумент: сопоставь факты товара с описанием кейса, "
    "покажи это в opi_reasoning_steps и укажи, усиливает ли кейс твой итоговый выбор.\n\n"
    + GEMINI_EAEU_BASE_RULES
)

# Схема ответа для Gemini Structured Outputs (подмножество OpenAPI Schema).
# Порядок свойств: opi_reasoning_steps первым — совпадает с :class:`HsClassificationResult`.
_HS_CLASSIFY_JSON_SCHEMA: dict[str, Any] = {
    "type": "object",
    "properties": {
        "opi_reasoning_steps": {
            "type": "array",
            "items": {"type": "string"},
            "description": "Пошаговые рассуждения по ОПИ 1-6 (сначала логика, затем код в suggested_hs_code); минимум один шаг",
        },
        "suggested_hs_code": {"type": "string", "description": "Ровно 10 цифр кода ТН ВЭД ЕАЭС"},
        "justification": {
            "type": "string",
            "description": "Краткое обоснование по ОПИ ТН ВЭД с опорой на прецеденты из промпта, если они даны",
        },
        "attributes": {
            "type": "object",
            "properties": {
                "name": {"type": "string", "description": "Наименование товара"},
                "manufacturer": {"type": "string", "description": "Изготовитель или НЕИЗВЕСТЕН"},
                "trademark": {
                    "type": "string",
                    "description": "Товарный знак с фото/текста или ОТСУТСТВУЕТ; при фото — только уверенно читаемый знак",
                },
                "model": {"type": "string", "description": "Марка/модель/артикул"},
                "material": {"type": "string", "description": "Материал"},
                "purpose_and_tech": {
                    "type": "string",
                    "description": "Назначение и технические характеристики",
                },
            },
            "required": [
                "name",
                "manufacturer",
                "trademark",
                "model",
                "material",
                "purpose_and_tech",
            ],
        },
        "confidence_score": {
            "type": "integer",
            "description": "Уверенность в выбранном коде ТН ВЭД, целое число 0–100",
        },
        "compliance_warnings": {"type": "array", "items": {"type": "string"}},
        "vision_insights": {
            "type": "string",
            "description": "Краткий вывод по фото (шильдик, конструкция); пустая строка если фото нет",
        },
        "box_31_description": {
            "type": "string",
            "description": "Готовый текст для графы 31 декларации: стиль как в примерах/предв. решениях из RAG, факты из инвойса",
        },
        "missing_info": {
            "type": "array",
            "items": {"type": "string"},
            "description": "Чего не хватает в данных для 100% уверенности; пустой массив, если запросить нечего",
        },
        "supplier_question_en": {
            "type": "string",
            "description": "Технический вопрос поставщику на английском для уточнения недостающих характеристик; если данных хватает — пустая строка",
        },
        "electronics_compliance": {
            "type": "object",
            "description": "Для глав 84/85/90; для остальных товаров — null (ключ опустить или null)",
            "properties": {
                "has_wireless_tech": {
                    "type": "boolean",
                    "description": "Wi‑Fi, Bluetooth, радиомодуль, NFC и т.п.",
                },
                "has_encryption": {
                    "type": "boolean",
                    "description": "Криптография / шифрование на устройстве",
                },
                "frequencies": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Рабочие частоты, например 2.4 GHz, 433 MHz",
                },
                "rf_license_required": {
                    "type": "boolean",
                    "description": "Нужно заключение РЧЦ / Минпромторга",
                },
                "fss_notification_required": {
                    "type": "boolean",
                    "description": "Нужна нотификация ФСБ",
                },
                "compliance_justification": {
                    "type": "string",
                    "description": "Аргументация со ссылкой на спецификации и нетарифные меры из контекста",
                },
            },
            "required": [
                "has_wireless_tech",
                "has_encryption",
                "frequencies",
                "rf_license_required",
                "fss_notification_required",
                "compliance_justification",
            ],
        },
    },
    "required": [
        "opi_reasoning_steps",
        "suggested_hs_code",
        "justification",
        "attributes",
        "confidence_score",
        "compliance_warnings",
        "vision_insights",
        "box_31_description",
        "missing_info",
        "supplier_question_en",
    ],
}

_LOW_CONFIDENCE_HS_SUFFIX = " [LOW CONFIDENCE] Требуется ручная проверка кода декларантом."

_BOX31_SKIP_VALUES = frozenset({"", "НЕИЗВЕСТЕН", "ОТСУТСТВУЕТ"})


def _env_int(name: str, default: int) -> int:
    raw = os.getenv(name)
    if raw is None:
        return default
    try:
        return int(str(raw).strip())
    except Exception:
        return default


def _env_float(name: str, default: float) -> float:
    raw = os.getenv(name)
    if raw is None:
        return default
    try:
        return float(str(raw).strip())
    except Exception:
        return default


# Границы ретраев Gemini при 429/ResourceExhausted (чтобы пайплайн не зависал на десятки минут).
_GEMINI_RATE_LIMIT_MAX_RETRIES = max(1, _env_int("GEMINI_RATE_LIMIT_MAX_RETRIES", 8))
_GEMINI_RATE_LIMIT_MAX_SLEEP_SEC = max(1.0, _env_float("GEMINI_RATE_LIMIT_MAX_SLEEP_SEC", 30.0))
_GEMINI_CLIENT_MIN_INTERVAL_SEC = max(0.0, _env_float("GEMINI_CLIENT_MIN_INTERVAL_SEC", 0.35))
_GEMINI_CLIENT_MAX_CALLS_PER_MIN = max(0, _env_int("GEMINI_CLIENT_MAX_CALLS_PER_MIN", 90))
_GEMINI_CLIENT_WINDOW_SEC = 60.0
_GEMINI_CLIENT_RATE_TS: deque[float] = deque()
_GEMINI_CLIENT_RATE_LOCK = threading.Lock()

# Главы ТН ВЭД (первые 2 цифры кода), разрешённые в промпте suggest_hs_code.
HS_SUGGEST_ALLOWED_CHAPTERS: frozenset[str] = frozenset(
    {"04", "33", "39", "40", "61", "62", "63", "73", "82", "83", "84", "85", "87", "90", "94", "95", "96"}
)

# Главы с расширенной логикой электроники (РЧЦ, ФСБ, Google Search).
HS_ELECTRONICS_CHAPTERS: frozenset[str] = frozenset({"84", "85", "90"})


def _backend_data_dir() -> Path:
    """Каталог `backend/data` (родитель `app/`)."""
    return Path(__file__).resolve().parents[2] / "data"


def temp_invoice_images_dir() -> Path:
    """Временные изображения строк инвойса: `backend/data/temp_images`."""
    return _backend_data_dir() / "temp_images"


def cleanup_temp_invoice_images() -> None:
    """Удаляет `backend/data/temp_images` после разбора файла (или для сброса кэша)."""
    d = temp_invoice_images_dir()
    if d.is_dir():
        shutil.rmtree(d, ignore_errors=True)


def raw_invoice_sidecar_images_dir() -> Path:
    """Папка с фото по артикулу / номеру строки: ``backend/data/raw_invoices/images/``."""
    return _backend_data_dir() / "raw_invoices" / "images"


_SIDE_CAR_IMAGE_EXTENSIONS: tuple[str, ...] = (".jpg", ".jpeg", ".png", ".webp")


def _sanitize_article_image_stem(article: str) -> str:
    """Безопасное имя файла из значения артикула (RS-2026 → RS-2026)."""
    s = (article or "").strip()
    if not s:
        return ""
    s = re.sub(r'[<>:"/\\|?\x00-\x1f]', "_", s)
    s = s.strip(" ._") or "item"
    return s[:240]


def _resolve_raw_invoice_sidecar_image(item_data: dict[str, Any]) -> Path | None:
    """
    Сторонние изображения в ``data/raw_invoices/images/``:

    - ``{Артикул}.jpg`` / ``.png`` / … (значение из строки после ``map_columns`` → ``article``);
    - или ``row_{N}.jpg``, где *N* — 1-based номер строки листа Excel (см. ``_invoice_excel_row`` в :func:`iter_item_rows`).
    """
    root = raw_invoice_sidecar_images_dir()
    if not root.is_dir():
        return None

    article = (
        str(item_data.get("article") or "").strip()
        or str(item_data.get("sku") or "").strip()
        or str(item_data.get("SKU") or "").strip()
        or str(item_data.get("Артикул") or "").strip()
    )
    if article:
        stem = _sanitize_article_image_stem(article)
        if stem:
            for ext in _SIDE_CAR_IMAGE_EXTENSIONS:
                cand = root / f"{stem}{ext}"
                if cand.is_file():
                    return cand.resolve()
                cand_u = root / (stem + ext.upper())
                if cand_u.is_file():
                    return cand_u.resolve()

    row_raw = item_data.get("_invoice_excel_row")
    if row_raw is not None and str(row_raw).strip():
        try:
            n = int(str(row_raw).strip())
        except ValueError:
            n = 0
        if n > 0:
            for ext in _SIDE_CAR_IMAGE_EXTENSIONS:
                cand = root / f"row_{n}{ext}"
                if cand.is_file():
                    return cand.resolve()
                cand_u = root / f"row_{n}{ext.upper()}"
                if cand_u.is_file():
                    return cand_u.resolve()
    return None


def _photo_analysis_report(path: Path | None) -> str:
    """Подпись для отчёта Excel: «Да (файл)» или «Нет»."""
    if path is None:
        return "Нет"
    try:
        if path.is_file():
            return f"Да ({path.name})"
    except OSError:
        pass
    return "Нет"


def extract_images_from_xlsx(file_path: str | Path, *, sheet: int | str = 0) -> dict[int, Path]:
    """
    Обходит `sheet._images` (openpyxl), сопоставляет якорь с номером строки Excel (1-based),
    сохраняет снимок в `backend/data/temp_images/{row_number}.jpg` (RGB JPEG).

    Учитываются вложения в колонке с заголовком «产品图片» / аналогах (как в спецификациях).
    Перед заполнением очищает каталог temp_images.
    """
    p = Path(file_path)
    if p.suffix.lower() != ".xlsx":
        return {}
    cleanup_temp_invoice_images()
    out_dir = temp_invoice_images_dir()
    out_dir.mkdir(parents=True, exist_ok=True)

    from openpyxl import load_workbook
    from PIL import Image

    wb = load_workbook(p, data_only=True, read_only=False)
    try:
        if isinstance(sheet, int):
            ws = wb.worksheets[sheet]
        else:
            ws = wb[str(sheet)]

        img_col: int | None = None
        for cell in ws[1]:
            if cell.value and _header_is_product_image_column(str(cell.value)):
                img_col = int(cell.column) - 1
                break
        if img_col is None:
            return {}

        result: dict[int, Path] = {}
        for im in getattr(ws, "_images", None) or []:
            try:
                frm = im.anchor._from
                if int(frm.col) != img_col:
                    continue
                excel_row = int(frm.row) + 1
                raw = im._data()
                if not raw:
                    continue
                out_path = (out_dir / f"{excel_row}.jpg").resolve()
                with Image.open(BytesIO(raw)) as pil_im:
                    rgb = pil_im.convert("RGB")
                    rgb.save(out_path, format="JPEG", quality=90)
                result[excel_row] = out_path
            except Exception as e:
                logger.debug("extract_images_from_xlsx: пропуск вложения: {}", e)
        return result
    finally:
        wb.close()


# Соответствие заголовков (китайский / русский) внутренним полям строки товара.
_HEADER_RULES: list[tuple[str, str]] = [
    ("Наименование", "name_ru"),
    ("Страна", "country_origin"),
    ("Код", "declared_hs_code"),
    ("俄语品名", "name_ru"),
    ("材质/成分", "material"),
    ("材质", "material"),
    ("成分", "material"),
    ("用途/功能", "usage"),
    ("用途", "usage"),
    ("功能", "usage"),
    ("净重", "weight_net"),
    ("毛重", "weight_gross"),
    ("原产国", "country_origin"),
    ("产地", "country_origin"),
    ("Origin_Country", "country_origin"),
    ("产品名称", "name_cn"),
    ("品牌", "brand"),
    ("总数量", "quantity"),
    ("单价", "unit_price"),
    ("Image_Path", "image_path"),
    ("Артикул", "article"),
    ("артикул", "article"),
    ("SKU", "article"),
    ("sku", "article"),
    ("Артикул/SKU", "article"),
]


# Подстроки в заголовке столбца цены/суммы (латиница + китайский).
_PRICE_HEADER_MARKERS = ("price", "amount", "单价", "金额")


def _header_is_product_image_column(header: str) -> bool:
    """Колонка с фото товара (产品图片 и аналоги)."""
    h = _norm_header(header)
    if not h:
        return False
    if h == _norm_header("产品图片"):
        return True
    if "图片" in h and ("产品" in h or "商品" in h):
        return True
    low = h.lower()
    if "product" in low and "image" in low:
        return True
    if "photo" in low and ("product" in low or "goods" in low):
        return True
    return False


def _norm_header(h: Any) -> str:
    s = str(h).strip() if h is not None else ""
    s = re.sub(r"[\n\r]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s


def _normalize_header_for_mapping(h: Any) -> str:
    """Ключ для сопоставления колонок: lower, без \\n/\\r, схлопывание пробелов."""
    s = "" if h is None else str(h)
    s = re.sub(r"[\n\r]+", " ", s, flags=re.MULTILINE)
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def _dedupe_normalized_column_names(names: list[str]) -> list[str]:
    """Гарантирует уникальные имена колонок после нормализации."""
    seen: dict[str, int] = {}
    out: list[str] = []
    for n in names:
        if n in seen:
            seen[n] += 1
            out.append(f"{n}__dup{seen[n]}")
        else:
            seen[n] = 0
            out.append(n)
    return out


def _cell_str_invoice(v: Any) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    s = str(v).strip()
    return "" if s.lower() == "nan" else s


def _header_suggests_price(header: str) -> bool:
    h = _norm_header(header)
    if not h:
        return False
    low = h.lower()
    for m in _PRICE_HEADER_MARKERS:
        if m.isascii():
            if m in low:
                return True
        elif m in h:
            return True
    return False


# JSON Schema для семантического маппинга колонок (Gemini structured output).
_COLUMN_MAP_JSON_SCHEMA: dict[str, Any] = {
    "type": "object",
    "properties": {
        "name_col": {
            "type": "string",
            "description": "Точное исходное имя колонки наименования товара; пустая строка если нет",
        },
        "brand_col": {
            "type": "string",
            "description": "Точное имя колонки бренда; пустая строка если нет",
        },
        "material_col": {
            "type": "array",
            "items": {"type": "string"},
            "description": "Список точных имён колонок про материал (верх, подошва и т.д.); пустой массив если нет",
        },
        "gross_weight_col": {
            "type": "string",
            "description": "Точное имя колонки веса брутто; пустая строка если нет",
        },
        "qty_col": {
            "type": "string",
            "description": "Точное имя колонки общего количества штук/пар (не коробок); пустая строка если нет",
        },
        "price_col": {
            "type": "string",
            "description": "Точное имя колонки цены или суммы; пустая строка если нет",
        },
        "image_col": {
            "type": "string",
            "description": "Точное имя колонки фото/ссылки на картинку; пустая строка если нет",
        },
    },
    "required": [
        "name_col",
        "brand_col",
        "material_col",
        "gross_weight_col",
        "qty_col",
        "price_col",
        "image_col",
    ],
}

_GEMINI_COLUMN_MAP_USER_INSTRUCTION = (
    "Ты — эксперт по обработке международных инвойсов и упаковочных листов. Тебе даны заголовки и первые 3 строки таблицы.\n"
    "Твоя задача — сопоставить оригинальные колонки из таблицы со стандартными системными ключами.\n"
    "Будь очень внимателен: отличай «количество мест/коробок» от «общего количества штук/пар». Нам нужно именно общее количество товара.\n"
    "Верни JSON по схеме (response_mime_type application/json): значения — это точные оригинальные названия колонков из таблицы "
    "(как в первой строке CSV и в массиве заголовков). Если колонки нет — пустая строка для строковых полей или пустой массив для material_col.\n"
    "Ключи ответа: name_col (наименование товара), brand_col (торговая марка), material_col (список колонок про материал; можно несколько), "
    "gross_weight_col (вес брутто), qty_col (общее количество штук/пар, не коробок), price_col (цена или стоимость), image_col (фото или ссылки).\n"
    "Колонки с префиксом «[RAW] » — технические дубликаты для отчёта; в ответе указывай только оригинальные имена колонок без префикса «[RAW] ».\n"
    "Не добавляй пояснений вне JSON."
)

# Поля, уже заданные эвристикой/ИИ; для остальных применяем _HEADER_RULES.
_INVOICE_COLUMN_MAP_RESERVED: frozenset[str] = frozenset(
    {
        "name_cn",
        "name_ru",
        "brand",
        "material",
        "weight_gross",
        "quantity",
        "image_path",
        "article",
        "total_cost_estimate",
        "Description_Full",
    }
)


def _parse_loose_json_object(raw: str) -> dict[str, Any] | None:
    t = (raw or "").strip()
    if not t:
        return None
    if t.startswith("```"):
        t = re.sub(r"^```(?:json)?\s*", "", t, flags=re.IGNORECASE).strip()
        t = re.sub(r"\s*```\s*$", "", t).strip()
    try:
        obj = json.loads(t)
    except json.JSONDecodeError:
        return None
    return obj if isinstance(obj, dict) else None


def _gemini_invoice_column_map_request(user_prompt: str) -> dict[str, Any] | None:
    """Один запрос к Gemini: маппинг колонок → dict или None."""
    key = (os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY") or "").strip()
    if not key:
        return None
    try:
        import google.generativeai as genai
    except ImportError:
        return None
    configure_google_generativeai(genai, api_key=key)
    gen_cfg = genai.GenerationConfig(
        temperature=0.05,
        max_output_tokens=1024,
        response_mime_type="application/json",
        response_schema=_COLUMN_MAP_JSON_SCHEMA,
    )
    model = genai.GenerativeModel(
        resolved_gemini_model_name(),
        system_instruction=_GEMINI_COLUMN_MAP_USER_INSTRUCTION,
    )
    max_retries = 8
    attempt = 0
    while True:
        try:
            resp = model.generate_content(user_prompt, generation_config=gen_cfg)
            try:
                raw_text = (resp.text or "").strip()
            except ValueError:
                logger.warning("_gemini_invoice_column_map_request: пустой ответ Gemini")
                return None
            data = _parse_loose_json_object(raw_text)
            if not isinstance(data, dict):
                return None
            return data
        except Exception as e:
            if _is_gemini_resource_exhausted(e) and attempt < max_retries:
                logger.warning(
                    "Gemini 429 (column map), пауза 45 с (попытка {}/{}): {}",
                    attempt + 1,
                    max_retries,
                    e,
                )
                time.sleep(45)
                attempt += 1
                continue
            logger.debug("_gemini_invoice_column_map_request: {}", e)
            return None


def _ai_map_columns(df: pd.DataFrame) -> dict[str, Any] | None:
    """
    Семантический маппинг заголовков через Gemini (заголовки + первые 3 строки CSV).

    Используется из ``map_columns``; сценарий ``scripts/test_invoice_parsing.py`` вызывает ``map_columns``.

    Возвращает dict с ключами name_col, brand_col, material_col (list[str]), gross_weight_col,
    qty_col, price_col, image_col — значения: точные имена колонок из ``df`` или пустые строки / [].
    """
    if df is None or df.empty:
        return None
    headers = [str(c) for c in df.columns]
    try:
        sample = df.head(3).to_csv(index=False, lineterminator="\n")
    except Exception as e:
        logger.debug("_ai_map_columns: to_csv: {}", e)
        return None
    if len(sample) > 16000:
        sample = sample[:16000] + "\n...[truncated]"
    headers_json = json.dumps(headers, ensure_ascii=False)
    user_prompt = (
        "Заголовки колонок (JSON-массив строк, в точности как в таблице):\n"
        f"{headers_json}\n\n"
        "Первые 3 строки данных (CSV):\n"
        f"{sample}\n"
    )
    return _gemini_invoice_column_map_request(user_prompt)


def _resolve_df_column_name(ref: Any, originals: list[str]) -> str | None:
    """Сопоставляет ответ ИИ с реальной колонкой ``df`` (точное или нормализованное совпадение)."""
    if ref is None or (isinstance(ref, float) and pd.isna(ref)):
        return None
    s = str(ref).strip()
    if not s or s.lower() in ("null", "none", "nan", ""):
        return None
    for o in originals:
        if str(o) == s:
            return str(o)
    for o in originals:
        if str(o).strip() == s.strip():
            return str(o)
    if re.match(r"^\[RAW\]\s*", s, flags=re.I):
        s_alt = re.sub(r"^\[RAW\]\s*", "", s, count=1, flags=re.I).strip()
        if s_alt:
            for o in originals:
                if str(o) == s_alt or str(o).strip() == s_alt.strip():
                    return str(o)
            ns_alt = _normalize_header_for_mapping(s_alt)
            for o in originals:
                if _normalize_header_for_mapping(str(o)) == ns_alt:
                    return str(o)
    ns = _normalize_header_for_mapping(s)
    for o in originals:
        if _normalize_header_for_mapping(str(o)) == ns:
            return str(o)
    return None


def _is_raw_mirror_column(name: Any) -> bool:
    """Колонка-копия исходника для отчёта (не трогать эвристикой переименования)."""
    return str(name).startswith("[RAW] ")


def _apply_header_rules_remaining_df(df: pd.DataFrame) -> pd.DataFrame:
    """Точные совпадения _HEADER_RULES для колонок, ещё не приведённых к внутренним именам."""
    out = df.copy()
    rename_map: dict[str, str] = {}
    for col in list(out.columns):
        if _is_raw_mirror_column(col):
            continue
        if col in _INVOICE_COLUMN_MAP_RESERVED:
            continue
        nk = _normalize_header_for_mapping(col)
        for needle, field in _HEADER_RULES:
            if field in _INVOICE_COLUMN_MAP_RESERVED:
                continue
            if field in out.columns:
                continue
            if nk == _normalize_header_for_mapping(needle):
                rename_map[col] = field
                break
    return out.rename(columns=rename_map) if rename_map else out


def _maybe_assign_price_column(df: pd.DataFrame) -> pd.DataFrame:
    if "total_cost_estimate" in df.columns:
        return df
    out = df.copy()
    for col in list(out.columns):
        if _is_raw_mirror_column(col):
            continue
        if _header_suggests_price(str(col)):
            return out.rename(columns={col: "total_cost_estimate"})
    return out


def _try_apply_ai_column_map(df: pd.DataFrame, ai: dict[str, Any]) -> tuple[pd.DataFrame, bool]:
    """Применяет ответ Gemini: переименование, ``Description_Full``, ``material`` при списке колонок."""
    originals = [str(c) for c in df.columns]
    out_pre = df.copy()

    name_c = _resolve_df_column_name(ai.get("name_col"), originals)
    brand_c = _resolve_df_column_name(ai.get("brand_col"), originals)
    gross_c = _resolve_df_column_name(ai.get("gross_weight_col"), originals)
    qty_c = _resolve_df_column_name(ai.get("qty_col"), originals)
    price_c = _resolve_df_column_name(ai.get("price_col"), originals)
    image_c = _resolve_df_column_name(ai.get("image_col"), originals)

    raw_mats = ai.get("material_col")
    mats_raw: list[Any]
    if raw_mats is None:
        mats_raw = []
    elif isinstance(raw_mats, str):
        mats_raw = [raw_mats] if str(raw_mats).strip() else []
    elif isinstance(raw_mats, list):
        mats_raw = list(raw_mats)
    else:
        mats_raw = []

    mats_resolved: list[str] = []
    seen_m: set[str] = set()
    for m in mats_raw:
        r = _resolve_df_column_name(m, originals)
        if r and r not in seen_m:
            seen_m.add(r)
            mats_resolved.append(r)

    scalar_srcs = {x for x in (name_c, brand_c, gross_c, qty_c, price_c, image_c) if x}
    mats_resolved = [m for m in mats_resolved if m not in scalar_srcs]

    if not scalar_srcs and not mats_resolved:
        return df.copy(), False

    def _desc_row(row: pd.Series) -> str:
        chunks: list[str] = []
        if name_c and name_c in row.index:
            s = _cell_str_invoice(row.get(name_c))
            if s:
                chunks.append(s)
        if brand_c and brand_c in row.index:
            b = _cell_str_invoice(row.get(brand_c))
            if b:
                chunks.append(f"Бренд: {b}")
        for m in mats_resolved:
            if m in row.index:
                v = _cell_str_invoice(row.get(m))
                if v:
                    chunks.append(f"{_norm_header(m)}: {v}")
        return ", ".join(chunks).strip()

    desc_series = out_pre.apply(_desc_row, axis=1) if len(out_pre) else pd.Series(dtype=object)

    material_joined = pd.Series([""] * len(out_pre), index=out_pre.index, dtype=object)
    if len(mats_resolved) > 1:

        def _mat_join_row(row: pd.Series) -> str:
            bits: list[str] = []
            for m in mats_resolved:
                if m not in row.index:
                    continue
                v = _cell_str_invoice(row.get(m))
                if v:
                    bits.append(f"{_norm_header(m)}: {v}")
            return ", ".join(bits)

        material_joined = out_pre.apply(_mat_join_row, axis=1)

    rename_m: dict[str, str] = {}
    used_src: set[str] = set()
    used_tgt: set[str] = set()

    def take(src: str | None, tgt: str) -> None:
        if not src or src not in out_pre.columns or src in used_src or tgt in used_tgt:
            return
        rename_m[src] = tgt
        used_src.add(src)
        used_tgt.add(tgt)

    take(name_c, "name_ru")
    take(brand_c, "brand")
    take(gross_c, "weight_gross")
    take(qty_c, "quantity")
    take(price_c, "total_cost_estimate")
    take(image_c, "image_path")
    if len(mats_resolved) == 1:
        take(mats_resolved[0], "material")

    out = out_pre.rename(columns=rename_m) if rename_m else out_pre.copy(deep=False)

    if len(mats_resolved) > 1:
        out["material"] = material_joined.reindex(out.index).fillna("").astype(str)
        out = out.drop(columns=[c for c in mats_resolved if c in out.columns], errors="ignore")

    if len(desc_series):
        out["Description_Full"] = desc_series.reindex(out.index).fillna("")
    else:
        out["Description_Full"] = ""

    desc_nonempty = out["Description_Full"].fillna("").astype(str).str.strip().ne("")
    if desc_nonempty.any():
        if "name_ru" not in out.columns:
            out["name_ru"] = ""
        out.loc[desc_nonempty, "name_ru"] = out.loc[desc_nonempty, "Description_Full"]

    out = _apply_header_rules_remaining_df(out)
    out = _maybe_assign_price_column(out)
    return out, True


def _promote_actual_headers(df: pd.DataFrame) -> pd.DataFrame:
    """
    Если pandas прочитал первую строку как данные (много ``Unnamed:`` в колонках),
    ищет в первых 10 строках реальную шапку (по ``notna``, ≥3 ячеек) и поднимает её в ``df.columns``,
    затем удаляет полностью пустые строки.
    """
    if df is None or df.empty:
        return df
    cols = list(df.columns)
    ncols = len(cols)
    if ncols == 0:
        return df.dropna(how="all")
    unnamed_count = sum(1 for c in cols if str(c).startswith("Unnamed:"))
    if unnamed_count <= ncols * 0.30:
        return df.dropna(how="all")

    out = df.copy(deep=False)
    n_scan = min(10, len(out))
    header_pos: int | None = None
    for pos in range(n_scan):
        row = out.iloc[pos]
        if int(row.notna().sum()) >= 3:
            header_pos = pos
            break
    if header_pos is None:
        return out.dropna(how="all")

    header_row = out.iloc[header_pos]
    new_names: list[str] = []
    for i in range(ncols):
        val = header_row.iloc[i]
        s = _cell_str_invoice(val).strip()
        new_names.append(f"empty_col_{i + 1}" if not s else s)

    out.columns = new_names
    out = out.iloc[header_pos + 1 :].reset_index(drop=True)
    return out.dropna(how="all")


def map_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Маппинг колонок: сначала семантический через Gemini (``_ai_map_columns``), при ошибке —
    эвристика (нормализация заголовков, ``str.contains`` / ``_HEADER_RULES``).

    Итоговые имена для ``InvoiceAnalyzer``: ``name_ru``, ``quantity``, ``weight_gross``, ``image_path``,
    ``total_cost_estimate``, ``brand``, ``material``, ``Description_Full`` (аналог латиницы Qty / Weight_Gross_kg в отчётах).

    После поднятия шапки в таблицу добавляются колонки ``[RAW] …`` — копии исходных полей для отчёта декларанту;
    порядок колонок в Excel — см. ``write_invoice_report_excel`` (``[RAW]`` в начале).
    """
    if df is None or df.empty:
        return pd.DataFrame()
    df = _promote_actual_headers(df)
    if df is None or df.empty:
        return pd.DataFrame()
    _raw_snapshot_cols = list(df.columns)
    for col in _raw_snapshot_cols:
        df[f"[RAW] {col}"] = df[col]
    ai = _ai_map_columns(df)
    if ai:
        out, ok = _try_apply_ai_column_map(df, ai)
        if ok:
            return out
    return _map_columns_fallback(df)


def _map_columns_fallback(df: pd.DataFrame) -> pd.DataFrame:
    """
    Нормализует заголовки (lower, без переносов строк), сопоставляет колонки по вхождению подстрок
    (русский / китайский / латиница), собирает ``Description_Full`` и заполняет ``name_ru`` для ИИ,
    переименовывает вес/количество/фото в поля пайплайна (``weight_gross``, ``quantity``, ``image_path``).
    """
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    norm_cols = _dedupe_normalized_column_names(
        [
            str(c) if _is_raw_mirror_column(c) else _normalize_header_for_mapping(c)
            for c in out.columns
        ]
    )
    out.columns = norm_cols
    cols = list(out.columns)
    consumed: set[str] = set()

    def _first(pred: Any) -> str | None:
        for c in cols:
            if c in consumed:
                continue
            try:
                if pred(c):
                    return c
            except Exception:
                continue
        return None

    def _take(src: str | None, internal: str, *, allow_replace: bool = False) -> None:
        nonlocal consumed
        if src is None:
            return
        if internal in seen_internal and not allow_replace:
            return
        if src in consumed:
            return
        rename_map[src] = internal
        seen_internal.add(internal)
        consumed.add(src)

    rename_map: dict[str, str] = {}
    seen_internal: set[str] = set()

    name_cn_c = _first(lambda c: "品名" in c)
    name_ru_c = _first(lambda c: "наименование" in c or "название груза" in c)
    brand_c = _first(lambda c: any(x in c for x in ("бренд", "марка", "品牌")))
    mat_up_c = _first(
        lambda c: ("материал" in c and "верх" in c)
        or ("材质" in c and "面" in c)
        or ("材料" in c and "面" in c)
    )
    mat_sol_c = _first(
        lambda c: ("материал" in c and "подошва" in c)
        or ("底" in c and any(t in c for t in ("材质", "材料", "материал", "鞋", "底材")))
    )
    gross_c = _first(
        lambda c: "брутто" in c.replace("ё", "е")
        or "общий вес" in c
        or "重量kg" in c.replace(" ", "")
        or "gross" in c
        or ("毛重" in c)
    )
    qty_c = _first(
        lambda c: "количество" in c
        or "双" in c
        or "总数量" in c
        or "数量" in c
    )
    if qty_c is None:
        qty_c = _first(
            lambda c: "total" in c
            and not any(x in c for x in ("金额", "价格", "price", "cost", "usd", "руб", "сумм", "amount"))
        )
    img_c = _first(lambda c: ("фото" in c or "照片" in c) and "产品图片" not in c and "商品图片" not in c)
    if img_c is None:
        img_c = _first(lambda c: ("图片" in c or "photo" in c) and ("路径" in c or "path" in c or "文件" in c))
    if img_c is None:
        img_c = _first(lambda c: c in ("照片", "图片", "фото") or c.endswith("照片") or c.endswith("图片"))
    article_c = _first(
        lambda c: _normalize_header_for_mapping(c) in ("артикул", "sku", "артикул/sku", "код товара", "vendor code")
        or ("артикул" in _normalize_header_for_mapping(c) and "наименование" not in _normalize_header_for_mapping(c))
    )

    # Сборка полного описания для ИИ (до переименования — ключи уже нормализованы)
    def _desc_for_row(row: pd.Series) -> str:
        chunks: list[str] = []
        for key in (name_ru_c, name_cn_c):
            if key and key in row.index:
                s = _cell_str_invoice(row.get(key))
                if s and s not in chunks:
                    chunks.append(s)
        if brand_c and brand_c in row.index:
            b = _cell_str_invoice(row.get(brand_c))
            if b:
                chunks.append(f"Бренд: {b}")
        if mat_up_c and mat_up_c in row.index:
            u = _cell_str_invoice(row.get(mat_up_c))
            if u:
                chunks.append(f"Верх: {u}")
        if mat_sol_c and mat_sol_c in row.index:
            so = _cell_str_invoice(row.get(mat_sol_c))
            if so:
                chunks.append(f"Подошва: {so}")
        return ", ".join(chunks).strip()

    desc_series = out.apply(_desc_for_row, axis=1) if len(out) else pd.Series(dtype=object)

    _take(name_cn_c, "name_cn")
    _take(name_ru_c, "name_ru")
    _take(brand_c, "brand")
    _take(gross_c, "weight_gross")
    _take(qty_c, "quantity")
    _take(img_c, "image_path")
    _take(article_c, "article")

    for col in cols:
        if _is_raw_mirror_column(col):
            continue
        if col in consumed:
            continue
        nk = _normalize_header_for_mapping(col)
        for needle, field in _HEADER_RULES:
            if field in ("name_cn", "name_ru", "brand", "weight_gross", "quantity", "image_path", "article"):
                continue
            if nk == _normalize_header_for_mapping(needle):
                _take(col, field)
                break

    if "total_cost_estimate" not in seen_internal:
        for col in cols:
            if _is_raw_mirror_column(col):
                continue
            if col in consumed:
                continue
            if _header_suggests_price(col):
                _take(col, "total_cost_estimate")
                break

    out = out.rename(columns=rename_map)

    if len(desc_series):
        out["Description_Full"] = desc_series.reindex(out.index).fillna("")
    else:
        out["Description_Full"] = ""

    desc_nonempty = out["Description_Full"].fillna("").astype(str).str.strip().ne("")
    if desc_nonempty.any():
        if "name_ru" not in out.columns:
            out["name_ru"] = ""
        out.loc[desc_nonempty, "name_ru"] = out.loc[desc_nonempty, "Description_Full"]

    # Материал: верх + подошва в одно поле material (если колонка попала в rename_map — берём новое имя)
    mat_up_key = rename_map[mat_up_c] if mat_up_c and mat_up_c in rename_map else mat_up_c
    mat_sol_key = rename_map[mat_sol_c] if mat_sol_c and mat_sol_c in rename_map else mat_sol_c
    if mat_up_key or mat_sol_key:
        def _merge_mat(row: pd.Series) -> str:
            u = _cell_str_invoice(row.get(mat_up_key)) if mat_up_key and mat_up_key in row.index else ""
            s = _cell_str_invoice(row.get(mat_sol_key)) if mat_sol_key and mat_sol_key in row.index else ""
            bits: list[str] = []
            if u:
                bits.append(f"Верх: {u}")
            if s:
                bits.append(f"Подошва: {s}")
            return ", ".join(bits)

        mat_combined = out.apply(_merge_mat, axis=1) if len(out) else pd.Series(dtype=object)
        mat_combined = mat_combined.fillna("").astype(str).str.strip()
        if "material" in out.columns:
            base_m = out["material"].fillna("").astype(str).str.strip()
            out["material"] = [
                (b + "; " + m) if b and m else (m or b)
                for b, m in zip(base_m.tolist(), mat_combined.tolist())
            ]
        else:
            out["material"] = mat_combined
        drop_m = [c for c in (mat_up_key, mat_sol_key) if c and c in out.columns]
        if drop_m:
            out = out.drop(columns=drop_m, errors="ignore")

    return out


def extract_images_from_excel(path: str | Path, *, sheet: int | str = 0) -> dict[int, Path]:
    """Совместимость: то же, что extract_images_from_xlsx (JPEG в backend/data/temp_images)."""
    return extract_images_from_xlsx(path, sheet=sheet)


def load_specification_table(
    path: str | Path, *, sheet: int | str = 0
) -> tuple[pd.DataFrame, dict[int, Path]]:
    """
    Читает CSV или XLS/XLSX в DataFrame.
    Для .xlsx дополнительно извлекает картинки (extract_images_from_xlsx → data/temp_images/{row}.jpg).
    """
    p = Path(path)
    if not p.is_file():
        raise FileNotFoundError(p)
    suf = p.suffix.lower()
    if suf == ".csv":
        return pd.read_csv(p, encoding="utf-8-sig", dtype=str), {}
    if suf == ".xlsx":
        df = pd.read_excel(p, sheet_name=sheet, dtype=str)
        imgs = extract_images_from_xlsx(p, sheet=sheet)
        return df, imgs
    if suf == ".xls":
        return pd.read_excel(p, sheet_name=sheet, dtype=str), {}
    raise ValueError(f"Неподдерживаемый формат: {suf}")


def iter_item_rows(
    df: pd.DataFrame,
    *,
    images_by_excel_row: dict[int, Path] | None = None,
    header_rows: int = 1,
) -> Iterator[dict[str, Any]]:
    """Строки как словари; пустые значения — пустые строки. При наличии карты — путь к фото в _image_path."""
    mapped = map_columns(df)
    for idx, (_, row) in enumerate(mapped.iterrows()):
        d: dict[str, Any] = {}
        for c in mapped.columns:
            v = row.get(c)
            d[str(c)] = "" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v).strip()
        excel_row = idx + 1 + header_rows
        d["_invoice_excel_row"] = excel_row
        if images_by_excel_row:
            ip = images_by_excel_row.get(excel_row)
            if ip is not None and Path(ip).is_file():
                d["_image_path"] = str(Path(ip).resolve())
        yield d


def _parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and pd.isna(value):
            return None
        return float(value)
    s = str(value).strip().replace("\xa0", " ").replace(" ", "").replace(",", ".")
    if not s:
        return None
    s = re.sub(r"[^\d.\-]", "", s)
    if not s or s in ("-", ".", "-."):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _calculate_smart_net_weight(gross_weight: Any, hs_code: str) -> float | None:
    """
    Оценка веса нетто из брутто и специфики ТН ВЭД (коэффициент из ``WEIGHT_COEFFICIENTS`` / дефолт 0.92).
    Возвращает ``None``, если брутто не задано или не положительное.
    """
    g = _parse_number(gross_weight)
    if g is None or g <= 0:
        return None
    coeff = _coefficient_for_hs_prefix(hs_code)
    return round(float(g) * coeff, 3)


def apply_smart_net_weight_to_line_item(
    item: dict[str, Any],
    hs_code: str,
    *,
    df: pd.DataFrame | None = None,
    row_index: int | None = None,
) -> str:
    """
    Если нетто пустое (0 / NaN / отсутствует), а брутто есть — заполняет ``weight_net`` (и алиасы ``Weight_Net_kg``)
    по коэффициенту для ``hs_code``. Опционально синхронизирует строку ``df`` по ``row_index``.

    Возвращает текст для ``Vision_Insights`` / ``Calculation_Notes`` или пустую строку.
    """
    gross_raw: Any = None
    for k in ("weight_gross", "Weight_Gross_kg", "weight_gross_kg"):
        v = item.get(k)
        if v is not None and str(v).strip() != "" and not (isinstance(v, float) and pd.isna(v)):
            gross_raw = v
            break
    if gross_raw is None:
        gross_raw = item.get("weight_gross")

    wn = _parse_number(item.get("weight_net"))
    if wn is None or wn <= 0:
        wn = _parse_number(item.get("Weight_Net_kg"))
    if wn is None or wn <= 0:
        wn = _parse_number(item.get("weight_net_kg"))
    if wn is not None and wn > 0:
        return ""

    net_val = _calculate_smart_net_weight(gross_raw, hs_code)
    if net_val is None:
        return ""

    coeff = _coefficient_for_hs_prefix(hs_code)
    g_num = _parse_number(gross_raw)
    s_net = f"{net_val:.3f}".rstrip("0").rstrip(".") if isinstance(net_val, float) else str(net_val)
    item["weight_net"] = s_net
    item["Weight_Net_kg"] = s_net

    if df is not None and row_index is not None and row_index >= 0:
        ri = int(row_index)
        for col in ("weight_net", "Weight_Net_kg"):
            if col in df.columns:
                try:
                    df.iloc[ri, df.columns.get_loc(col)] = s_net
                except Exception as e:
                    logger.debug("apply_smart_net_weight_to_line_item: не записали {} в df: {}", col, e)

    g_show = g_num if g_num is not None else gross_raw
    return (
        f"[AUTO-NETTO] Рассчитано по коэффициенту {coeff} (Брутто: {g_show} -> Нетто: {net_val})."
    )


def _customs_value_from_line(item_data: dict[str, Any] | None) -> float | None:
    """
    Таможенная стоимость для расчёта пошлины/НДС.

    Если в строке уже есть `customs_value` (после InvoiceAnalyzer._calculate_financials), используем её;
    иначе — unit_price * quantity.
    """
    if not item_data:
        return None
    cv = _parse_number(item_data.get("customs_value"))
    if cv is not None and cv >= 0:
        return round(float(cv), 2)
    qty = _parse_number(item_data.get("quantity"))
    unit = _parse_number(item_data.get("unit_price"))
    if qty is None or qty <= 0 or unit is None or unit < 0:
        return None
    return round(unit * qty, 2)


_DUTY_AD_VALOREM_RE = re.compile(r"(\d+(?:\.\d+)?)\s*%", re.IGNORECASE)
_DUTY_SPECIFIC_EUR_RE = re.compile(
    r"(\d+(?:\.\d+)?)\s*(?:евро\s*/\s*кг|евро\s*за|евро/кг|евро\b|eur\s*/\s*кг|eur/кг|eur\b|€\s*/\s*кг|€)",
    re.IGNORECASE,
)


def _parse_duty_rate(rate_string: str | Any) -> dict[str, Any]:
    """
    Разбор текстовой ставки ввозной пошлины (ЕТТ / справочник): адвалорная %, специфика €/кг, правило MAX/ADD/STANDARD.
    """
    if rate_string is None or (isinstance(rate_string, float) and pd.isna(rate_string)):
        return {"ad_valorem": 0.0, "specific_eur": 0.0, "rule": "STANDARD", "raw_text": ""}
    if isinstance(rate_string, (int, float)) and not isinstance(rate_string, bool):
        raw_text = f"{float(rate_string):g}%"
    else:
        raw_text = str(rate_string).strip()
    if not raw_text:
        return {"ad_valorem": 0.0, "specific_eur": 0.0, "rule": "STANDARD", "raw_text": raw_text}

    m_pct = _DUTY_AD_VALOREM_RE.search(raw_text)
    ad_valorem = float(m_pct.group(1)) if m_pct else 0.0
    m_eur = _DUTY_SPECIFIC_EUR_RE.search(raw_text)
    specific_eur = float(m_eur.group(1)) if m_eur else 0.0

    if ad_valorem == 0.0 and specific_eur == 0.0:
        m_plain = re.match(r"^\s*(\d+(?:\.\d+)?)\s*$", raw_text)
        if m_plain:
            ad_valorem = float(m_plain.group(1))

    low = raw_text.lower()
    if "не менее" in low or "не меньше" in low or re.search(r"\bmax\b", low):
        rule = "MAX"
    elif "+" in raw_text or "плюс" in low:
        rule = "ADD"
    else:
        rule = "STANDARD"

    return {
        "ad_valorem": ad_valorem,
        "specific_eur": specific_eur,
        "rule": rule,
        "raw_text": raw_text,
    }


def _format_geopolitical_duty_for_parse(override: Any) -> str:
    """Строка ставки для ``_parse_duty_rate`` после геополитической подмены (число → с «%»)."""
    s = str(override).strip() if override is not None else ""
    if not s:
        return "0"
    low = s.lower()
    if "%" in s or "евро" in low or "eur" in low or "€" in s:
        return s
    try:
        return f"{float(s.replace(',', '.')):g}%"
    except (TypeError, ValueError):
        return s


def _compute_base_duty_rub(
    parsed: dict[str, Any],
    customs_value_rub: float,
    net_weight_kg: float | None,
    weight_gross_kg: float | None,
    eur_rate: float,
) -> tuple[float, list[str]]:
    """
    Базовая сумма пошлины в рублях: адвалор от таможенной стоимости; специфика — масса нетто × €/кг × курс EUR.
    Возвращает (сумма, список предупреждений для декларанта/отчёта).
    """
    ad_pct = float(parsed.get("ad_valorem") or 0.0)
    spec_eur = float(parsed.get("specific_eur") or 0.0)
    rule = str(parsed.get("rule") or "STANDARD").strip().upper() or "STANDARD"
    warnings: list[str] = []

    cv = max(0.0, float(customs_value_rub))
    er = max(0.0, float(eur_rate))
    ad_valorem_amt = cv * (ad_pct / 100.0)

    wn = net_weight_kg if net_weight_kg is not None and net_weight_kg > 0 else None
    wg = weight_gross_kg if weight_gross_kg is not None and weight_gross_kg > 0 else None

    weight_for_specific: float | None = None
    if spec_eur > 0:
        if wn is not None:
            weight_for_specific = float(wn)
        else:
            base_msg = "[MATH ERROR] Нет веса нетто для расчета специфической пошлины"
            if wg is not None:
                logger.warning(
                    "{} — используем вес брутто {:.6f} кг как временный fallback",
                    base_msg,
                    float(wg),
                )
                weight_for_specific = float(wg)
                warnings.append(
                    f"{base_msg} — применён вес брутто {float(wg):g} кг (временный fallback для €/кг)."
                )
            else:
                logger.warning("{}, специфическая часть пошлины принята за 0", base_msg)
                weight_for_specific = 0.0
                warnings.append(f"{base_msg}; вес брутто отсутствует — специфическая часть = 0.")

    specific_amt = float(weight_for_specific or 0.0) * spec_eur * er if spec_eur > 0 else 0.0

    if rule == "MAX":
        duty_amount = max(ad_valorem_amt, specific_amt)
    elif rule == "ADD":
        duty_amount = ad_valorem_amt + specific_amt
    else:
        if ad_pct > 0:
            duty_amount = ad_valorem_amt
        elif spec_eur > 0:
            duty_amount = specific_amt
        else:
            duty_amount = 0.0

    return round(max(0.0, float(duty_amount)), 2), warnings


# Incoterms: фрахт в таможенную стоимость строки добавляем только для «поставка без фрахта в цене».
_INCOTERMS_FREIGHT_ADD_TO_CUSTOMS: frozenset[str] = frozenset({"EXW", "FCA", "FOB", "FAS"})
_INCOTERMS_FREIGHT_IN_INVOICE_PRICE: frozenset[str] = frozenset({"CFR", "CIF", "CPT", "CIP", "DAP", "DPU", "DDP"})

_SANCTION_RANK = {"forbidden": 3, "ban": 3, "risk": 2, "warn": 2, "safe": 1, "ok": 1, "": 0}


def _scan_sanction_import_risks(hs_digits: str) -> tuple[str, str]:
    """Возвращает (sanction_risk текст, Sanction_Status на русском)."""
    p = re.sub(r"\D", "", hs_digits or "")[:10]
    sanction_risk = ""
    sanction_status = "Безопасно"
    worst = 0
    if len(p) < 4:
        return sanction_risk, sanction_status
    with SessionLocal() as db:
        rows = db.query(SanctionImportRisk).all()
    for sr in rows:
        px = (sr.hs_code_prefix or "").strip()
        if not px or not p.startswith(px):
            continue
        lvl = (sr.risk_level or "risk").strip().lower()
        rank = _SANCTION_RANK.get(lvl, 2)
        if rank > worst:
            worst = rank
            sanction_risk = (
                f"КРИТИЧНО: префикс {px} ({sr.jurisdiction}): {(sr.description or '')[:400]}"
            ).strip()
    if worst >= _SANCTION_RANK["forbidden"]:
        sanction_status = "Запрещено"
    elif worst >= _SANCTION_RANK["risk"]:
        sanction_status = "Риск"
    return sanction_risk, sanction_status


def _norm_sanction_name(raw: Any) -> str:
    s = re.sub(r"[^A-Za-zА-Яа-я0-9]+", " ", str(raw or "").upper())
    return re.sub(r"\s+", " ", s).strip()


def _parse_aliases(raw: Any) -> list[str]:
    txt = str(raw or "").strip()
    if not txt:
        return []
    try:
        arr = json.loads(txt)
        if isinstance(arr, list):
            out = []
            for x in arr:
                t = str(x or "").strip()
                if t:
                    out.append(t)
            return out
    except Exception:
        pass
    return [x.strip() for x in re.split(r"[;,|\n]+", txt) if x.strip()]


def check_sanction_risks(
    hs_code: str,
    manufacturer_name: str,
    origin_country: str | None,
    db_session: Session | None = None,
) -> dict[str, Any]:
    """
    Жесткая проверка санкционных рисков:
    - OFAC SDN: fuzzy-совпадение по name/aliases.
    - EU sanctions: попадание HS-кода в санкционные записи.
    - Country specific rules: локальные правила по стране происхождения.
    """
    hs = re.sub(r"\D", "", str(hs_code or ""))[:10]
    name_raw = str(manufacturer_name or "").strip()
    name_norm = _norm_sanction_name(name_raw)
    iso = _normalize_country_iso(origin_country)

    out: dict[str, Any] = {
        "ofac_hit": False,
        "ofac_score": 0.0,
        "ofac_match_name": "",
        "eu_hit": False,
        "eu_matches": [],
        "country_rule_hit": False,
        "country_rules": [],
        "alerts": [],
    }

    own = False
    s = db_session
    if s is None:
        s = SessionLocal()
        own = True
    try:
        # --- OFAC fuzzy ---
        if len(name_norm) >= 3:
            tokens = [t for t in name_norm.split() if len(t) >= 3][:3]
            q = s.query(OfacSdnList)
            if tokens:
                ors = []
                for t in tokens:
                    pat = f"%{t}%"
                    ors.append(OfacSdnList.name.ilike(pat))
                    ors.append(OfacSdnList.aliases.ilike(pat))
                q = q.filter(or_(*ors))
            cands = q.limit(5000).all()
            best_score = 0.0
            best_name = ""
            for row in cands:
                names = [str(row.name or "").strip(), *_parse_aliases(row.aliases)]
                for cand in names:
                    cand_norm = _norm_sanction_name(cand)
                    if not cand_norm:
                        continue
                    sc = SequenceMatcher(None, name_norm, cand_norm).ratio()
                    if sc > best_score:
                        best_score = sc
                        best_name = cand
            if best_score >= 0.82:
                out["ofac_hit"] = True
                out["ofac_score"] = round(best_score, 4)
                out["ofac_match_name"] = best_name[:1024]
                out["alerts"].append(
                    f"[КРИТИЧЕСКИЙ РИСК OFAC] Производитель найден в санкционном списке США (совпадение {best_score * 100:.0f}%)!"
                )

        # --- EU HS restrictions ---
        eu_rows: list[EuSanctionsList] = []
        if len(hs) >= 4:
            hs_prefixes = []
            for L in (10, 8, 6, 4):
                if len(hs) >= L:
                    hs_prefixes.append(hs[:L])
            q_eu = s.query(EuSanctionsList)
            ors = []
            for p in hs_prefixes:
                ors.append(EuSanctionsList.hs_code == p)
                ors.append(EuSanctionsList.hs_code.like(f"{p}%"))
            if ors:
                q_eu = q_eu.filter(or_(*ors))
            eu_rows = q_eu.limit(40).all()
            if eu_rows:
                out["eu_hit"] = True
                out["eu_matches"] = [
                    {
                        "hs_code": (r.hs_code or "")[:10],
                        "entity_name": (r.entity_name or "")[:512],
                        "description": (r.description or "")[:800],
                    }
                    for r in eu_rows[:5]
                ]
                out["alerts"].append(
                    "[САНКЦИИ ЕС] Код попадает под ограничения экспорта товаров двойного назначения."
                )

        # --- Country specific rules ---
        if iso:
            c_rows = (
                s.query(CountrySpecificRule)
                .filter(CountrySpecificRule.country_code == iso)
                .order_by(CountrySpecificRule.id.asc())
                .limit(20)
                .all()
            )
            if c_rows:
                out["country_rule_hit"] = True
                out["country_rules"] = [
                    f"[{(r.rule_type or 'other').upper()}] {(r.description or '').strip()[:400]}"
                    for r in c_rows
                    if (r.description or "").strip()
                ][:10]
                if out["country_rules"]:
                    out["alerts"].append(
                        f"[СТРАНОВОЕ ОГРАНИЧЕНИЕ] Для {iso} действует усиленный документальный контроль."
                    )
    finally:
        if own and s is not None:
            s.close()
    return out


def _normalize_country_iso(raw: Any) -> str | None:
    s = str(raw or "").strip().upper()
    if not s:
        return None
    if len(s) == 2 and s.isalpha():
        return s
    m = re.search(r"\b([A-Z]{2})\b", s)
    if m:
        return m.group(1)
    compact = re.sub(r"[\s._-]+", "", s)
    aliases = {"USA": "US", "UK": "GB", "РОССИЯ": "RU", "РФ": "RU", "КИТАЙ": "CN", "США": "US"}
    return aliases.get(compact) if compact in aliases else None


def _extract_country_origin(item_data: dict[str, Any] | None) -> str | None:
    if not item_data:
        return None
    for key in ("country_origin", "origin_country", "country_of_origin", "Origin_Country"):
        v = item_data.get(key)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        t = str(v).strip()
        if t:
            return t
    return None


def _product_blob_for_rag(item_data: dict[str, Any] | None) -> str:
    """Текст строки инвойса для поиска прецедентов / префиксов нетарифки."""
    if not item_data:
        return ""
    parts = [
        str(item_data.get("name_ru") or ""),
        str(item_data.get("name_cn") or ""),
        str(item_data.get("material") or ""),
        str(item_data.get("usage") or ""),
        str(item_data.get("brand") or ""),
    ]
    return " ".join(p for p in parts if p).strip()


def _nt_prefix_for_rag(product_description: str, suggested_chapter: str | None) -> str | None:
    """Префикс ТН ВЭД (2–6 знаков) для выборки non_tariff_measures."""
    if suggested_chapter:
        ch = re.sub(r"\D", "", str(suggested_chapter).strip())[:2].zfill(2)
        if len(ch) == 2 and ch.isdigit():
            return ch
    blob = re.sub(r"[^\d]+", " ", product_description or "")
    runs = re.findall(r"\d{4,10}", blob)
    for r in sorted(runs, key=len, reverse=True):
        if len(r) >= 6:
            return r[:6]
        if len(r) >= 4:
            return r[:4]
    return None


def _get_rag_context(
    db_session: Session,
    product_description: str,
    suggested_chapter: str | None = None,
) -> str:
    """
    Сборка RAG для классификации: см. :func:`app.services.rag_retriever.build_rag_context`.
    """
    from .rag_retriever import build_rag_context

    pref = _nt_prefix_for_rag(product_description, suggested_chapter) or ""
    if not pref and suggested_chapter:
        ch = re.sub(r"\D", "", str(suggested_chapter).strip())[:2].zfill(2)
        if len(ch) == 2 and ch.isdigit():
            pref = ch
    return build_rag_context(db_session, pref, product_description)


class InvoiceAnalyzer:
    """Финансово-логистический слой: фрахт, распределение, таможенная стоимость по Incoterms."""

    def __init__(
        self,
        *,
        freight_usd: float = 0.0,
        incoterms: str = "EXW",
        db_session: Session | None = None,
    ) -> None:
        self.freight_usd = float(freight_usd)
        self.incoterms = (incoterms or "EXW").strip().upper()
        self._db_session = db_session

    def suggest_hs_code_for_item(
        self,
        item_data: dict[str, Any],
        *,
        image_path: Any = _IMG_PATH_UNSET,
        db_session: Session | None = None,
        fast_mode: bool = False,
    ) -> dict[str, Any]:
        """
        Подбор 10-значного ТН ВЭД, текстов для графы 31 (сборка из атрибутов и ``box_31_description`` от ИИ) и (при фото) ``Vision_Insights``.
        Путь к изображению: колонка ``Image_Path`` в строке, вложение ``_image_path`` или явный аргумент ``image_path``.
        ``db_session``: переопределение сессии БД для RAG (по умолчанию — сессия из конструктора).
        """
        return suggest_hs_code(
            item_data,
            image_path=image_path,
            db_session=db_session if db_session is not None else self._db_session,
            fast_mode=fast_mode,
        )

    def apply_financial_columns(
        self,
        df: pd.DataFrame,
        usd_rate: float,
        eur_rate: float | None = None,
    ) -> pd.DataFrame:
        """Обёртка: подставляет freight/incoterms из конструктора; курс EUR — для расчёта специфических пошлин."""
        return self._calculate_financials(df, self.freight_usd, self.incoterms, float(usd_rate), eur_rate)

    def check_geopolitical_risks(self, hs_code: str, country_origin: str | None) -> dict[str, Any]:
        """
        Страна (CountryRisk), заградительные ставки (geo_special_duties / ПП РФ №2140),
        санкционные префиксы (sanction_import_risks).
        """
        p = re.sub(r"\D", "", hs_code or "")[:10]
        iso = _normalize_country_iso(country_origin)
        sanction_risk, sanction_status = _scan_sanction_import_risks(p)
        neutral = {
            "Country_Risk_Status": "Нейтральная",
            "Applied_Special_Duty": "",
            "Required_Certificates": "Сертификат происхождения — по общему порядку при необходимости.",
            "Sanction_Status": sanction_status,
            "sanction_risk": sanction_risk[:2000],
            "duty_rate_override": None,
            "geopolitical_duty_note": "",
            "embargo_blocked": False,
            "embargo_document_basis": "",
            "embargo_document_link": "",
        }
        if not iso:
            return neutral

        cr = get_country_risk_by_iso(iso)
        if cr is None:
            return neutral

        unf = bool(cr.is_unfriendly)
        pref = bool(cr.has_preference)
        if unf:
            origin_label = "Недружественная"
        elif pref:
            origin_label = "Преференция"
        else:
            origin_label = "Нейтральная"

        required = (cr.required_cert or "").strip() or neutral["Required_Certificates"]
        if pref and "СТ-1" not in required.upper() and "FORM" not in required.upper():
            required = (required + "; возможна преференция 0% — СТ-1 / Form A (EAV)").strip("; ")

        duty_override_raw: str | None = None
        basis_full = ""
        embargo_row = None
        if len(p) >= 4:
            embargo_row = find_geo_embargo_match(p, iso, country_is_unfriendly=unf)
        if embargo_row is not None:
            basis_emb = (embargo_row.document_basis or "").strip()
            link_emb = (embargo_row.document_link or "").strip()
            crit = f"[CRITICAL RISK] Ввоз запрещен на основании {basis_emb}."
            merged_risk = f"{crit} {sanction_risk}".strip()[:2000]
            return {
                "Country_Risk_Status": origin_label,
                "Applied_Special_Duty": format_applied_special_duty_label(basis_emb),
                "Required_Certificates": (required + "; ввоз запрещён нормативными мерами").strip("; ")[:2000],
                "Sanction_Status": "Запрещено",
                "sanction_risk": merged_risk,
                "duty_rate_override": None,
                "geopolitical_duty_note": basis_emb[:500],
                "embargo_blocked": True,
                "embargo_document_basis": basis_emb[:512],
                "embargo_document_link": link_emb[:2000],
            }

        if len(p) >= 4:
            spec = find_geo_duty_override_row(p, iso, country_is_unfriendly=unf)
            if spec is not None:
                dr_geo = str(spec.duty_rate).strip() if spec.duty_rate is not None else ""
                duty_override_raw = dr_geo if dr_geo else None
                basis_full = (spec.document_basis or "").strip()

        applied_short = format_applied_special_duty_label(basis_full) if duty_override_raw is not None else ""

        if sanction_status != "Безопасно":
            required = (
                required + "; расширенный комплаенс по санкционным спискам ЕС/США/UK"
            ).strip("; ")

        if unf and duty_override_raw is not None:
            required = (
                required + "; непреференциальный сертификат при отсутствии льготной зоны происхождения"
            ).strip("; ")

        return {
            "Country_Risk_Status": origin_label,
            "Applied_Special_Duty": applied_short,
            "Required_Certificates": required[:2000],
            "Sanction_Status": sanction_status,
            "sanction_risk": sanction_risk[:2000],
            "duty_rate_override": duty_override_raw,
            "geopolitical_duty_note": basis_full[:500],
            "embargo_blocked": False,
            "embargo_document_basis": "",
            "embargo_document_link": "",
        }

    def _calculate_financials(
        self,
        df: pd.DataFrame,
        freight_usd: float,
        incoterms: str,
        usd_rate: float,
        eur_rate: float | None = None,
    ) -> pd.DataFrame:
        """
        Шаги А–Д: фрахт в руб., база распределения (вес брутто / иначе стоимость строки),
        доля фрахта по строке, customs_value по Incoterms, колонки в DataFrame.
        Курс EUR (ЦБ РФ) пишется в ``eur_rate_cb`` для согласованности с расчётом специфической пошлины.
        """
        try:
            if df is None or df.empty:
                return df
            out = coerce_invoice_gross_weight_columns(df.copy())
            if eur_rate is None:
                try:
                    from .currency_sync import CurrencyService

                    eur_rate = float(CurrencyService.get_eur_rate())
                except Exception:
                    eur_rate = 100.0
            eur_rate = max(0.0, float(eur_rate))
            inc = (incoterms or "EXW").strip().upper()
            if inc not in _INCOTERMS_FREIGHT_ADD_TO_CUSTOMS | _INCOTERMS_FREIGHT_IN_INVOICE_PRICE:
                logger.warning(
                    "InvoiceAnalyzer: неизвестный Incoterms {!r} — считаем как EXW (фрахт добавляется к таможенной стоимости)",
                    inc,
                )
                inc = "EXW"

            total_freight_rub = max(0.0, float(freight_usd)) * max(0.0, float(usd_rate))

            if inc in _INCOTERMS_FREIGHT_IN_INVOICE_PRICE and float(freight_usd) > 0:
                logger.warning(
                    "Incoterms {}: фрахт в цене инвойса — к таможенной стоимости строки allocated_freight_rub не "
                    "добавляется (только распределение для landed cost).",
                    inc,
                )

            n = len(out)
            line_usd: list[float] = []
            weights: list[float] = []
            for _, row in out.iterrows():
                qty = _parse_number(row.get("quantity"))
                unit = _parse_number(row.get("unit_price"))
                total_est = _parse_number(row.get("total_cost_estimate"))
                line = 0.0
                if qty is not None and unit is not None and qty > 0 and unit >= 0:
                    line = float(unit * qty)
                elif total_est is not None and total_est >= 0:
                    line = float(total_est)
                line_usd.append(line)
                wg = _parse_number(row.get("weight_gross"))
                if wg is None or wg <= 0:
                    wg = _parse_number(row.get("Weight_Gross_kg"))
                if wg is None or wg <= 0:
                    wg = _parse_number(row.get("weight_gross_kg"))
                wn = _parse_number(row.get("weight_net"))
                w = 0.0
                if wg is not None and wg > 0:
                    w = float(wg)
                elif wn is not None and wn > 0:
                    w = float(wn)
                weights.append(w)

            sum_w = sum(weights)
            sum_line = sum(line_usd)
            use_weight = sum_w > 0
            if use_weight:
                denom = sum_w
                keys = weights
            else:
                denom = sum_line if sum_line > 0 else 0.0
                keys = line_usd

            allocated: list[float] = []
            if total_freight_rub <= 0 or n == 0:
                allocated = [0.0] * n
            elif denom <= 0:
                share = total_freight_rub / float(n)
                allocated = [round(share, 2)] * n
            else:
                raw = [total_freight_rub * (k / denom) for k in keys]
                rounded = [round(x, 2) for x in raw]
                drift = round(total_freight_rub - sum(rounded), 2)
                if rounded and abs(drift) >= 0.01:
                    rounded[-1] = round(rounded[-1] + drift, 2)
                allocated = rounded

            item_prices_rub: list[float] = []
            customs_values: list[float] = []
            landed_addons: list[float] = []

            add_freight_to_cv = inc in _INCOTERMS_FREIGHT_ADD_TO_CUSTOMS
            freight_in_invoice = inc in _INCOTERMS_FREIGHT_IN_INVOICE_PRICE

            for i in range(n):
                ip_rub = round(line_usd[i] * max(0.0, float(usd_rate)), 2)
                item_prices_rub.append(ip_rub)
                af = allocated[i] if i < len(allocated) else 0.0
                if add_freight_to_cv:
                    cv = round(ip_rub + af, 2)
                else:
                    cv = round(ip_rub, 2)
                customs_values.append(cv)
                # Для себестоимости на складе: если фрахт уже в customs_value, не дублируем.
                landed_addons.append(round(af, 2) if freight_in_invoice else 0.0)

            out["item_price_rub"] = item_prices_rub
            out["allocated_freight_rub"] = allocated
            out["customs_value"] = customs_values
            out["landed_cost_freight_addon"] = landed_addons
            out["incoterms"] = [inc] * n
            out["eur_rate_cb"] = [round(eur_rate, 4)] * n
            return out
        except Exception as e:
            logger.warning("InvoiceAnalyzer._calculate_financials: ошибка, возвращаем исходный df без изменений: {}", e)
            return df


def format_warning_cell(text: str | Sequence[Any] | None) -> str:
    """Текст предупреждения для Excel — с заметным префиксом."""
    if text is None:
        return ""
    if isinstance(text, (list, tuple)):
        parts = [str(x).strip() for x in text if str(x).strip()]
        body = "; ".join(parts)
    else:
        body = str(text).strip()
    if not body:
        return ""
    if body.upper().startswith("ВНИМАНИЕ"):
        return body
    return f"ВНИМАНИЕ: {body}"


def _extract_json_object(raw: str) -> dict[str, Any]:
    text = (raw or "").strip()
    text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*```$", "", text)
    i = text.find("{")
    j = text.rfind("}")
    if i == -1 or j <= i:
        return {}
    try:
        obj = json.loads(text[i : j + 1])
    except json.JSONDecodeError:
        return {}
    return obj if isinstance(obj, dict) else {}


def _merge_purpose_into_purpose_and_tech(attributes: Any) -> dict[str, Any] | None:
    """Совместимость: старые ответы с ``purpose`` → ``purpose_and_tech``."""
    if not isinstance(attributes, dict):
        return None
    out = dict(attributes)
    pt = str(out.get("purpose_and_tech") or "").strip()
    if not pt:
        legacy = out.get("purpose")
        if legacy is not None and str(legacy).strip():
            out["purpose_and_tech"] = str(legacy).strip()
    return out


def _with_merged_attributes(data: dict[str, Any]) -> dict[str, Any]:
    out = dict(data or {})
    if "opi_reasoning_steps" not in out and isinstance(out.get("reasoning_steps"), list):
        out["opi_reasoning_steps"] = list(out.get("reasoning_steps") or [])
    if "reasoning_steps" not in out and isinstance(out.get("opi_reasoning_steps"), list):
        out["reasoning_steps"] = list(out.get("opi_reasoning_steps") or [])
    if "suggested_hs_code" not in out and out.get("hs_code") is not None:
        out["suggested_hs_code"] = str(out.get("hs_code") or "")
    if "hs_code" not in out and out.get("suggested_hs_code") is not None:
        out["hs_code"] = str(out.get("suggested_hs_code") or "")
    att = out.get("attributes")
    merged = _merge_purpose_into_purpose_and_tech(att)
    if merged is not None:
        out["attributes"] = merged
    return out


def _parse_hs_classify_response_json(raw_text: str) -> dict[str, Any] | None:
    """
    Разбор JSON ответа классификации: ``json.loads`` → :func:`_extract_json_object` → повторная попытка по срезу ``{...}``.
    При битом JSON возвращает ``None`` (вызывающий может подставить сырой текст).
    """
    s = (raw_text or "").strip()
    if not s:
        return None
    try:
        obj = json.loads(s)
        if isinstance(obj, dict):
            return _with_merged_attributes(obj)
    except json.JSONDecodeError:
        pass
    ex = _extract_json_object(s)
    if isinstance(ex, dict) and ex:
        return _with_merged_attributes(ex)
    i, j = s.find("{"), s.rfind("}")
    if i != -1 and j > i:
        try:
            obj = json.loads(s[i : j + 1])
            if isinstance(obj, dict):
                return _with_merged_attributes(obj)
        except json.JSONDecodeError:
            pass
    logger.warning("Ответ классификации ТН ВЭД: не удалось разобрать JSON (ни json.loads, ни выделение объекта).")
    return None


def _extract_json_array(raw: str) -> list[Any]:
    text = (raw or "").strip()
    text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*```$", "", text)
    i = text.find("[")
    j = text.rfind("]")
    if i == -1 or j <= i:
        return []
    try:
        arr = json.loads(text[i : j + 1])
    except json.JSONDecodeError:
        return []
    return arr if isinstance(arr, list) else []


def _parse_countries_with_llm(legal_text: str) -> list[str]:
    """
    Извлечение ISO-2 из фрагмента постановления через Gemini (развёртка ЕС в коды стран — по промпту).
    """
    excerpt = (legal_text or "").strip()[:8000]
    if not excerpt:
        return []
    eu_hint = (
        "Страны ЕС (ISO-2, ориентир): AT, BE, BG, HR, CY, CZ, DK, EE, FI, FR, DE, GR, HU, IE, IT, LV, LT, LU, "
        "MT, NL, PL, PT, RO, SK, SI, ES, SE."
    )
    prompt = (
        "Извлеки из этого юридического текста все упомянутые страны и верни их в виде JSON-списка строгих ISO-кодов "
        "(например, [\"US\", \"GB\", \"DE\", \"FR\"]). Если упоминается «Европейский союз», разверни его в список "
        "всех ISO-кодов стран-участниц.\n"
        f"{eu_hint}\n"
        "Ответь только валидным JSON-массивом строк из двух латинских букв в верхнем регистре, без пояснений.\n\n"
        f"Текст:\n{excerpt}"
    )
    try:
        text = _gemini_generate(prompt, max_output_tokens=2048, temperature=0.0)
    except Exception as e:
        logger.warning("_parse_countries_with_llm: Gemini недоступен: {}", e)
        return []
    codes: list[str] = []
    for x in _extract_json_array(text):
        s = str(x).strip().upper()
        if len(s) == 2 and s.isalpha():
            codes.append(s)
    return sorted(set(codes))


def _is_gemini_resource_exhausted(exc: BaseException) -> bool:
    """429 / quota / ResourceExhausted от Google Generative AI."""
    try:
        from google.api_core import exceptions as ga_exc

        if isinstance(exc, ga_exc.ResourceExhausted):
            return True
    except Exception:
        pass
    s = str(exc).lower()
    return (
        "429" in s
        or "resource exhausted" in s
        or "too many requests" in s
        or "quota" in s
        or "rate limit" in s
    )


def _gemini_rate_limit_sleep(attempt: int) -> float:
    """Экспоненциальный backoff + jitter для 429, с верхним пределом."""
    # 1, 2, 4, 8, ... + jitter, но не дольше _GEMINI_RATE_LIMIT_MAX_SLEEP_SEC.
    base = min(_GEMINI_RATE_LIMIT_MAX_SLEEP_SEC, float(2 ** max(0, attempt)))
    jitter = min(5.0, base * 0.2)
    delay = min(_GEMINI_RATE_LIMIT_MAX_SLEEP_SEC, base + (jitter * 0.5))
    time.sleep(delay)
    return delay


def _gemini_global_rate_limit(stage: str) -> None:
    """
    Глобальный клиентский лимитер Gemini (process-wide):
    - минимальный интервал между запросами;
    - максимум запросов в скользящем окне 60 секунд.
    """
    if _GEMINI_CLIENT_MIN_INTERVAL_SEC <= 0 and _GEMINI_CLIENT_MAX_CALLS_PER_MIN <= 0:
        return
    while True:
        wait_sec = 0.0
        with _GEMINI_CLIENT_RATE_LOCK:
            now = time.monotonic()
            while _GEMINI_CLIENT_RATE_TS and now - _GEMINI_CLIENT_RATE_TS[0] >= _GEMINI_CLIENT_WINDOW_SEC:
                _GEMINI_CLIENT_RATE_TS.popleft()

            if _GEMINI_CLIENT_MIN_INTERVAL_SEC > 0 and _GEMINI_CLIENT_RATE_TS:
                wait_sec = max(wait_sec, (_GEMINI_CLIENT_RATE_TS[-1] + _GEMINI_CLIENT_MIN_INTERVAL_SEC) - now)
            if _GEMINI_CLIENT_MAX_CALLS_PER_MIN > 0 and len(_GEMINI_CLIENT_RATE_TS) >= _GEMINI_CLIENT_MAX_CALLS_PER_MIN:
                wait_sec = max(wait_sec, (_GEMINI_CLIENT_RATE_TS[0] + _GEMINI_CLIENT_WINDOW_SEC) - now)

            if wait_sec <= 0:
                _GEMINI_CLIENT_RATE_TS.append(now)
                return
        if wait_sec > 0.0:
            if wait_sec >= 1.0:
                logger.debug("Gemini client rate limit: stage={}, sleep={:.2f}s", stage, wait_sec)
            time.sleep(wait_sec)


def _gemini_generate(
    prompt: str,
    *,
    max_output_tokens: int = 1024,
    temperature: float = 0.15,
    image_path: Path | None = None,
    generation_config: dict[str, Any] | None = None,
) -> str:
    """
    Текстовый запрос к Gemini; при ``image_path`` — мультимодально (текст + Pillow Image в ``generate_content``).
    ``generation_config`` при передаче объединяется с temperature/max_output_tokens (можно задать свои ключи).
    """
    key = (os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY") or "").strip()
    if not key:
        raise RuntimeError("Не задан GEMINI_API_KEY или GOOGLE_API_KEY")
    try:
        import google.generativeai as genai
    except ImportError as e:
        raise RuntimeError("Нужен пакет google-generativeai") from e
    configure_google_generativeai(genai, api_key=key)
    model = genai.GenerativeModel(resolved_gemini_model_name())
    cfg: dict[str, Any] = dict(generation_config or {})
    cfg.setdefault("temperature", temperature)
    cfg.setdefault("max_output_tokens", max_output_tokens)
    max_retries = _GEMINI_RATE_LIMIT_MAX_RETRIES
    attempt = 0
    while True:
        try:
            _gemini_global_rate_limit("generate")
            pil_img = None
            if image_path is not None and image_path.is_file():
                try:
                    from PIL import Image
                except ImportError as e:
                    raise RuntimeError("Нужен пакет Pillow для мультимодального запроса") from e
                try:
                    # validate image first to avoid passing a broken stream to Gemini
                    with Image.open(image_path) as probe:
                        probe.verify()
                    pil_img = Image.open(image_path)
                except Exception as e:
                    logger.warning(
                        "_gemini_generate: изображение не прочитано ({}), запрос без фото: {}",
                        image_path,
                        e,
                    )
            if pil_img is not None:
                try:
                    resp = model.generate_content(
                        [prompt, pil_img],
                        generation_config=cfg,
                    )
                finally:
                    pil_img.close()
            else:
                resp = model.generate_content(prompt, generation_config=cfg)
            return (getattr(resp, "text", "") or "").strip()
        except Exception as e:
            if _is_gemini_resource_exhausted(e) and attempt < max_retries:
                slept = _gemini_rate_limit_sleep(attempt)
                logger.warning(
                    "Gemini 429/rate-limit (generate), пауза {:.1f} с (попытка {}/{}): {}",
                    slept,
                    attempt + 1,
                    max_retries,
                    e,
                )
                attempt += 1
                continue
            raise


def _gemini_generate_multimodal(
    prompt: str,
    image_path: Path,
    *,
    max_output_tokens: int = 1024,
    temperature: float = 0.15,
) -> str:
    """Совместимость: делегирует в ``_gemini_generate`` с изображением."""
    return _gemini_generate(
        prompt,
        max_output_tokens=max_output_tokens,
        temperature=temperature,
        image_path=image_path,
    )


def _resolve_invoice_row_image_path(item_data: dict[str, Any]) -> Path | None:
    """
    Путь к фото строки: колонка ``Image_Path`` / ``_image_path`` (XLSX), затем каталог
    ``data/raw_invoices/images/`` — по артикулу или ``row_{номер строки Excel}``.
    """
    from .vision_extractor import DocumentVisionExtractor

    ext = DocumentVisionExtractor()
    for key in ("image_path", "_image_path"):
        v = item_data.get(key)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        s = str(v).strip()
        if not s:
            continue
        got = ext.extract_image_for_row(s)
        if got is not None:
            return got
    side = _resolve_raw_invoice_sidecar_image(item_data)
    if side is not None:
        return side
    return None


def _validate_product_image_path(path: Path | None) -> Path | None:
    """Проверяет, что файл — читаемое изображение Pillow; иначе WARNING и ``None``."""
    if path is None:
        return None
    try:
        from PIL import Image

        with Image.open(path) as im:
            im.load()
    except Exception as e:
        logger.warning(
            "Не удалось открыть или прочитать изображение товара ({}). Продолжаем анализ только по тексту: {}",
            path,
            e,
        )
        return None
    return path


def _parse_vat_rate_final_value(raw: Any) -> float | None:
    """Итог экспертизы НДС: 10% (льгота) или базовая ставка проекта (22%). 20% трактуем как устаревший ответ → 22%."""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        v = float(raw)
        if abs(v - 20.0) < 1e-9:
            v = float(DEFAULT_VAT_RATE)
        return v if v in (10.0, float(DEFAULT_VAT_RATE)) else None
    s = str(raw).strip().replace(",", ".")
    m = re.search(r"\b(10|20|22)\b", s)
    if not m:
        return None
    got = float(m.group(1))
    if abs(got - 20.0) < 1e-9:
        got = float(DEFAULT_VAT_RATE)
    return got if got in (10.0, float(DEFAULT_VAT_RATE)) else None


def gemini_vat_expertise_preferential(
    hs_code: str,
    item_data: dict[str, Any],
    *,
    group: dict[str, Any],
    image_path: Path | None = None,
) -> dict[str, Any]:
    """
    Второй запрос к Gemini для кодов из справочника ПП РФ №908 / №688.
    Возвращает vat_rate_final (10 или 22) и vat_logic.
    """
    out: dict[str, Any] = {"vat_rate_final": None, "vat_logic": ""}
    if not (os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY") or "").strip():
        return out

    hs = re.sub(r"\D", "", hs_code or "")[:10]
    if len(hs) != 10:
        return out

    name_ru = (item_data.get("name_ru") or "").strip()
    name_cn = (item_data.get("name_cn") or "").strip()
    material = (item_data.get("material") or "").strip()
    usage = (item_data.get("usage") or "").strip()
    brand = (item_data.get("brand") or "").strip()
    act = str(group.get("act") or "")
    cat = str(group.get("category") or "")
    pref = str(group.get("prefix") or "")

    expertise_block = (
        f"ЭКСПЕРТИЗА НДС: Этот код ({hs}) может претендовать на ставку 10%. Проанализируй описание и фото:\n\n"
        "Является ли товар изделием именно для детей? (маркировка, дизайн, размер).\n\n"
        "Соответствует ли он техническим критериям (например, для мебели — ростовые группы, для одежды — рост/обхват).\n\n"
        "Если это мед. изделие — есть ли признаки профессионального использования?\n\n"
        "Верни JSON поле vat_rate_final (10 или 22) и vat_logic (краткое пояснение, почему выбрана эта ставка со ссылкой на критерии закона)."
    )

    prompt = (
        GEMINI_PROJECT_VAT_RULES
        + "Ты налоговый и таможенный эксперт по НДС РФ при импорте.\n"
        f"Справочная группа льготы (по префиксу {pref}): {act} — {cat}.\n"
        f"Код ТН ВЭД: {hs}\n"
        f"Наименование (RU): {name_ru or '—'}\n"
        f"Наименование (CN): {name_cn or '—'}\n"
        f"Материал: {material or '—'}\n"
        f"Назначение: {usage or '—'}\n"
        f"Бренд: {brand or '—'}\n\n"
        f"{expertise_block}\n\n"
        'Верни СТРОГО один JSON-объект без markdown: {"vat_rate_final":10,"vat_logic":"..."} '
        f"(vat_rate_final — 10 при льготе по ПП РФ, иначе {DEFAULT_VAT_RATE} как базовая ставка проекта)."
    )

    resolved_img = image_path if image_path and image_path.is_file() else _resolve_invoice_row_image_path(item_data)
    resolved_img = _validate_product_image_path(resolved_img)

    try:
        text = _gemini_generate(
            prompt,
            max_output_tokens=1024,
            temperature=0.1,
            image_path=resolved_img,
        )
    except Exception as e:
        logger.warning("gemini_vat_expertise_preferential: {}", e)
        return out

    data = _extract_json_object(text)
    vf = _parse_vat_rate_final_value(data.get("vat_rate_final"))
    logic = str(data.get("vat_logic") or "").strip()
    if vf is not None:
        out["vat_rate_final"] = int(vf) if vf == int(vf) else vf
    out["vat_logic"] = logic[:8000]
    return out


def _normalize_compliance_warnings(raw: Any) -> list[str]:
    if raw is None:
        return []
    if isinstance(raw, list):
        out: list[str] = []
        for el in raw:
            if isinstance(el, dict):
                out.append("; ".join(f"{k}: {v}" for k, v in el.items() if v))
            else:
                s = str(el).strip()
                if s:
                    out.append(s)
        return out
    if isinstance(raw, str) and raw.strip():
        return [raw.strip()]
    return []


# Маркеры «искусственного» состава / синтетики в описании (нижний регистр для поиска).
_SYNTHETIC_PROFILE_MARKERS: tuple[str, ...] = (
    "эко",
    "искусств",
    "pu",
    "сетка",
    "резина",
    "пластик",
    "pvc",
    "пвх",
    "поли-",
    "полиуретан",
    "синтетик",
    "искусственн",
    "металл",
    "metal",
    "резин",
)
_FORBIDDEN_VET_PHYTO_SUBSTR: tuple[str, ...] = ("ветеринарн", "фитосанитарн")
_STANDARD_TR_TS_CERT_LINE = "Стандартная сертификация ТР ТС"

_INVOICE_GROSS_WEIGHT_COLUMN_ALIASES: tuple[str, ...] = (
    "Weight_Gross_kg",
    "weight_gross",
    "weight_gross_kg",
)


def coerce_invoice_gross_weight_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Универсально приводит колонки брутто к float по всем строкам (как ``pd.to_numeric`` после нормализации запятой).
    Вызывается до расчёта фрахта и до обогащения пошлинами.
    """
    if df is None or df.empty:
        return df
    out = df.copy()
    for col in _INVOICE_GROSS_WEIGHT_COLUMN_ALIASES:
        if col not in out.columns:
            continue
        ser = (
            out[col]
            .astype(str)
            .str.replace("\xa0", " ", regex=False)
            .str.replace(" ", "", regex=False)
            .str.replace(",", ".", regex=False)
        )
        out[col] = pd.to_numeric(ser, errors="coerce").fillna(0.0)
    return out


def _item_synthetic_profile_blob(
    item_data: dict[str, Any] | None,
    attrs: dict[str, Any] | None = None,
) -> str:
    parts: list[str] = []
    if item_data:
        for k in ("name_ru", "name_cn", "material", "usage", "brand", "Description_Full", "description"):
            v = item_data.get(k)
            if v is not None and str(v).strip():
                parts.append(str(v))
    if attrs:
        for k in ("name", "material", "purpose_and_tech", "model"):
            v = attrs.get(k)
            if v is not None and str(v).strip():
                parts.append(str(v))
    return " ".join(parts).lower()


def _profile_indicates_synthetic_materials(blob: str) -> bool:
    b = (blob or "").lower()
    return any(m in b for m in _SYNTHETIC_PROFILE_MARKERS)


def _text_mentions_forbidden_vet_phyto(s: str) -> bool:
    low = (s or "").lower()
    return any(tok in low for tok in _FORBIDDEN_VET_PHYTO_SUBSTR)


def _sanitize_compliance_warnings_for_synthetic_items(
    warnings: list[str],
    *,
    item_data: dict[str, Any] | None,
    attrs: dict[str, Any] | None = None,
) -> list[str]:
    """Если в описании синтетика/ПУ и т.п., вычищаем из compliance вет/фито-формулировки модели."""
    blob = _item_synthetic_profile_blob(item_data, attrs)
    if not _profile_indicates_synthetic_materials(blob):
        return list(warnings or [])
    out: list[str] = []
    for line in warnings or []:
        t = str(line).strip()
        if not t:
            continue
        if _text_mentions_forbidden_vet_phyto(t):
            out.append(_STANDARD_TR_TS_CERT_LINE)
        else:
            out.append(t)
    return out


def _sanitize_risk_notes_for_synthetic_items(notes: str, item_data: dict[str, Any] | None) -> str:
    if not (notes or "").strip():
        return notes
    blob = _item_synthetic_profile_blob(item_data, None)
    if not _profile_indicates_synthetic_materials(blob):
        return notes
    if not _text_mentions_forbidden_vet_phyto(notes):
        return notes
    lines = str(notes).split("\n")
    new_lines: list[str] = []
    for ln in lines:
        if _text_mentions_forbidden_vet_phyto(ln):
            new_lines.append(_STANDARD_TR_TS_CERT_LINE)
        else:
            new_lines.append(ln)
    return "\n".join(new_lines)


def _sanitize_enrichment_certs_and_non_tariff(out: dict[str, Any], item_data: dict[str, Any] | None) -> None:
    """Пост-фильтр БД/ИИ: вет/фито при синтетическом профиле → стандартная формулировка ТР ТС."""
    if not item_data:
        return
    blob = _item_synthetic_profile_blob(item_data, None)
    if not _profile_indicates_synthetic_materials(blob):
        return
    rc = str(out.get("Required_Certificates") or "")
    if _text_mentions_forbidden_vet_phyto(rc):
        out["Required_Certificates"] = _STANDARD_TR_TS_CERT_LINE[:2000]
    for nt in out.get("non_tariff") or []:
        if not isinstance(nt, dict):
            continue
        for key in ("document_required", "description"):
            val = str(nt.get(key) or "")
            if _text_mentions_forbidden_vet_phyto(val):
                lim = 255 if key == "document_required" else 500
                nt[key] = _STANDARD_TR_TS_CERT_LINE[:lim]


def _merge_trois_registry_risks(compliance_warnings: list[str], attrs: dict[str, Any]) -> tuple[list[str], str]:
    """
    Проверка ``trois_registry`` по ``attributes["trademark"]``; дополняет compliance_warnings и текст для Excel.
    """
    tm_raw = str(attrs.get("trademark") or "").strip()
    norm = normalize_trademark_for_registry(tm_raw)
    if not norm or norm in ("НЕИЗВЕСТЕН", "ОТСУТСТВУЕТ"):
        return list(compliance_warnings), ""
    base = list(compliance_warnings or [])
    try:
        with SessionLocal() as db:
            hits = query_trois_matches_for_trademark(db, norm)
            ip_hits = _query_intellectual_property_trademark(db, norm)
            ip_holders = [str(x.right_holder or "").strip() for x in ip_hits if str(x.right_holder or "").strip()]
    except Exception as e:
        logger.debug("TROIS registry: пропуск проверки ({})", e)
        return base, ""
    msgs: list[str] = []
    seen: set[str] = set()
    for h in hits:
        key = (h.reg_number or "").strip() or str(h.id)
        if key in seen:
            continue
        seen.add(key)
        rh = (h.right_holder or "").strip() or (ip_holders[0] if ip_holders else "—")
        score = float(getattr(h, "_trois_match_score", 1.0) or 0.0)
        quality = ""
        if score < 0.95:
            quality = f" Нечеткое совпадение: {score:.2f}, требуется ручная проверка."
        valid_until = (h.valid_until or "").strip() if hasattr(h, "valid_until") else ""
        reps = (h.representatives or "").strip() if hasattr(h, "representatives") else ""
        brand = (getattr(h, "brand", "") or h.trademark or norm).strip()
        extra_bits: list[str] = []
        if (h.reg_number or "").strip():
            extra_bits.append(f"рег. №{h.reg_number.strip()}")
        if valid_until:
            extra_bits.append(f"срок: {valid_until}")
        if reps:
            extra_bits.append(f"представители: {reps}")
        extra_txt = f" ({'; '.join(extra_bits)})" if extra_bits else ""
        msgs.append(
            f'[ТРОИС RISK] Бренд в ТРОИС! Требуется разрешение правообладателя: {rh}. '
            f'Бренд: "{brand}".{extra_txt}{quality}'
        )
    if not msgs:
        return base, ""
    return [*base, *msgs], " | ".join(msgs)


def _query_intellectual_property_trademark(db: Session, tm_upper: str) -> list[IntellectualProperty]:
    """Поиск в локальной БД ТРОИС (таблица intellectual_properties): точное совпадение brand_name, затем LIKE."""
    from sqlalchemy import func

    def _norm(s: str) -> str:
        x = normalize_trademark_for_registry(s)
        x = re.sub(r"[^A-ZА-ЯЁ0-9]+", " ", x)
        return re.sub(r"\s+", " ", x).strip()

    def _score(a: str, b: str) -> float:
        return SequenceMatcher(None, _norm(a), _norm(b)).ratio()

    if len(tm_upper) < 2:
        return []
    bu = IntellectualProperty.brand_name
    exact = (
        db.query(IntellectualProperty)
        .filter(func.upper(func.trim(bu)) == tm_upper)
        .limit(20)
        .all()
    )
    if exact:
        for row in exact:
            setattr(row, "_trois_match_score", 1.0)
        return exact
    if len(tm_upper) >= 3:
        pat = f"%{tm_upper}%"
        like_rows = (
            db.query(IntellectualProperty)
            .filter(func.upper(func.trim(bu)).like(pat))
            .limit(200)
            .all()
        )
        ranked: list[tuple[float, IntellectualProperty]] = []
        for row in like_rows:
            s = _score(tm_upper, row.brand_name or "")
            if s >= 0.74:
                ranked.append((s, row))
        ranked.sort(key=lambda x: x[0], reverse=True)
        out = [r for _, r in ranked[:20]]
        for s, row in ranked[:20]:
            setattr(row, "_trois_match_score", round(float(s), 4))
        return out
    return []


def _append_photo_intellectual_property_trademark_risks(
    compliance_warnings: list[str],
    trois_cell: str,
    attrs: dict[str, Any],
    *,
    had_image_in_request: bool,
) -> tuple[list[str], str]:
    """
    Если к классификации было приложено фото: проверка attributes.trademark по ``intellectual_properties`` (LIKE).
    Сообщения — про обнаружение на фото; колонка Trois_Control дополняется блоком «по фото».
    """
    if not had_image_in_request:
        return compliance_warnings, trois_cell
    tm_raw = str(attrs.get("trademark") or "").strip()
    norm = normalize_trademark_for_registry(tm_raw)
    if not norm or norm in ("НЕИЗВЕСТЕН", "ОТСУТСТВУЕТ"):
        return compliance_warnings, trois_cell
    try:
        with SessionLocal() as db:
            hits = _query_intellectual_property_trademark(db, norm)
    except Exception as e:
        logger.debug("IntellectualProperty (фото/ТМ): пропуск ({})", e)
        return compliance_warnings, trois_cell
    msgs: list[str] = []
    seen: set[str] = set()
    for h in hits:
        key = (h.reg_number or "").strip() or f"{h.brand_name}|{h.hs_code_prefix}"
        if key in seen:
            continue
        seen.add(key)
        rh = (h.right_holder or "").strip() or "—"
        score = float(getattr(h, "_trois_match_score", 1.0) or 0.0)
        quality = ""
        if score < 0.95:
            quality = f" Нечеткое совпадение: {score:.2f}, требуется ручная проверка."
        msgs.append(
            f'[ТРОИС RISK] Бренд в ТРОИС! Требуется разрешение правообладателя: {rh}. '
            f'На фото обнаружен знак "{norm}" (реестр IP).{quality}'
        )
    if not msgs:
        return compliance_warnings, trois_cell
    base_cw = list(compliance_warnings or [])
    for m in msgs:
        if m not in base_cw:
            base_cw.append(m)
    photo_tc = " | ".join(msgs)
    if trois_cell.strip():
        trois_out = f"{trois_cell.strip()} || [по фото, реестр ТРОИС]: {photo_tc}"
    else:
        trois_out = f"[по фото, реестр ТРОИС]: {photo_tc}"
    return base_cw, trois_out


def _build_box_31(attributes: dict[str, Any] | None) -> str:
    """
    31-я графа по шаблону ФТС: нумеруются только блоки наименование / изготовитель / знак / модель;
    «Материал» и «Характеристики» без номера. Пункты со значениями «НЕИЗВЕСТЕН» / «ОТСУТСТВУЕТ» пропускаются.
    """
    if not attributes or not isinstance(attributes, dict):
        return ""

    def _omit(v: str) -> bool:
        s = (v or "").strip()
        return not s or s in _BOX31_SKIP_VALUES

    name = str(attributes.get("name") or "").strip()
    manufacturer = str(attributes.get("manufacturer") or "").strip()
    trademark = str(attributes.get("trademark") or "").strip()
    model = str(attributes.get("model") or "").strip()
    material = str(attributes.get("material") or "").strip()
    purpose_and_tech = str(
        attributes.get("purpose_and_tech") or attributes.get("purpose") or ""
    ).strip()

    numbered: list[str] = []
    if not _omit(name):
        numbered.append(f"{len(numbered) + 1}. {name}.")
    if not _omit(manufacturer):
        numbered.append(f"{len(numbered) + 1}. Изготовитель: {manufacturer}.")
    if not _omit(trademark):
        numbered.append(f"{len(numbered) + 1}. Товарный знак: {trademark}.")
    if not _omit(model):
        numbered.append(f"{len(numbered) + 1}. Марка/Модель: {model}.")

    tail: list[str] = []
    if not _omit(material):
        tail.append(f"Материал: {material}.")
    if not _omit(purpose_and_tech):
        tail.append(f"Характеристики: {purpose_and_tech}.")

    return " ".join([*numbered, *tail]).strip()


def _build_box_31_description(attributes: dict[str, Any] | None) -> str:
    """Совместимость со старым именем; см. :func:`_build_box_31`."""
    return _build_box_31(attributes)


def _default_hs_attributes_from_item(item_data: dict[str, Any]) -> dict[str, str]:
    name_ru = (item_data.get("name_ru") or "").strip()
    name_cn = (item_data.get("name_cn") or "").strip()
    brand = (item_data.get("brand") or "").strip()
    material = (item_data.get("material") or "").strip()
    usage = (item_data.get("usage") or "").strip()
    name = name_ru or name_cn or "НЕИЗВЕСТЕН"
    tm = "ОТСУТСТВУЕТ"
    if brand and brand not in ("无", "—", "-"):
        tm = brand
    return {
        "name": name,
        "manufacturer": "НЕИЗВЕСТЕН",
        "trademark": tm,
        "model": "НЕИЗВЕСТЕН",
        "material": material or "НЕИЗВЕСТЕН",
        "purpose_and_tech": usage or "НЕИЗВЕСТЕН",
    }


def _normalize_hs_attributes_from_llm(raw: Any) -> dict[str, str]:
    keys = ("name", "manufacturer", "trademark", "model", "material", "purpose_and_tech")
    out: dict[str, str] = {k: "НЕИЗВЕСТЕН" for k in keys}
    out["trademark"] = "ОТСУТСТВУЕТ"
    if isinstance(raw, dict):
        src = dict(raw)
        if "purpose_and_tech" not in src and src.get("purpose"):
            src["purpose_and_tech"] = src.get("purpose")
        for k in keys:
            v = str(src.get(k) or "").strip()
            if v:
                out[k] = v
    return out


def _parse_confidence_score(raw: Any) -> float:
    try:
        v = float(raw)
    except (TypeError, ValueError):
        return 0.0
    return max(0.0, min(100.0, v))


def _normalize_reasoning_steps(raw: Any) -> list[str]:
    if not isinstance(raw, list):
        return []
    out: list[str] = []
    for x in raw:
        s = str(x).strip()
        if s:
            out.append(s)
    return out


def _normalize_missing_info_list(raw: Any) -> list[str]:
    if raw is None:
        return []
    if not isinstance(raw, list):
        return []
    out: list[str] = []
    for x in raw:
        s = str(x).strip()
        if s:
            out.append(s)
    return out


def _chapter_two_from_hs(hs_code: str) -> str:
    return re.sub(r"\D", "", (hs_code or "").strip())[:2]


def _is_aggregate_hs10(hs_code: str) -> bool:
    d = re.sub(r"\D", "", (hs_code or "").strip())[:10]
    return len(d) == 10 and d.endswith("000000")


def _is_allowed_suggest_hs10(hs_code: str) -> bool:
    d = re.sub(r"\D", "", (hs_code or "").strip())[:10]
    if len(d) != 10:
        return False
    if _is_aggregate_hs10(d):
        return False
    return _chapter_two_from_hs(d) in HS_SUGGEST_ALLOWED_CHAPTERS


def _normalize_model_hs10(raw: Any) -> str:
    d = re.sub(r"\D", "", str(raw or ""))[:10]
    return d if _is_allowed_suggest_hs10(d) else ""


def _resolve_existing_hs10_with_session(db: Session, raw_hs: Any) -> tuple[str, str]:
    """
    Возвращает только существующий в ``tnved_commodities`` 10-значный код.

    Стратегия:
    1) exact match;
    2) ближайший "родитель" через zero-padding (2202991500 -> 2202990000);
    3) самый близкий код по префиксу (длиннейший префикс, затем минимальный код).
    """
    hs = re.sub(r"\D", "", str(raw_hs or ""))[:10]
    if len(hs) != 10:
        return "", "invalid_format"

    exact = db.query(Commodity.code).filter(Commodity.code == hs).first()
    if exact and re.sub(r"\D", "", str(exact[0] or ""))[:10] == hs:
        return hs, "exact"

    for ln in range(9, 1, -1):
        cand = hs[:ln] + ("0" * (10 - ln))
        row = db.query(Commodity.code).filter(Commodity.code == cand).first()
        if row:
            c = re.sub(r"\D", "", str(row[0] or ""))[:10]
            if len(c) == 10:
                return c, f"parent_zero_pad_{ln}"

    for ln in range(9, 1, -1):
        pref = hs[:ln]
        row = (
            db.query(Commodity.code)
            .filter(Commodity.code.like(f"{pref}%"))
            .order_by(Commodity.code.asc())
            .first()
        )
        if row:
            c = re.sub(r"\D", "", str(row[0] or ""))[:10]
            if len(c) == 10:
                return c, f"prefix_match_{ln}"

    return "", "not_found"


def _default_electronics_compliance_dict() -> dict[str, Any]:
    return {
        "has_wireless_tech": False,
        "has_encryption": False,
        "frequencies": [],
        "rf_license_required": False,
        "fss_notification_required": False,
        "compliance_justification": "",
    }


def _normalize_electronics_compliance_payload(raw: Any, *, hs_two: str) -> dict[str, Any] | None:
    """Нормализует блок electronics_compliance; ``None`` если глава не 84/85/90."""
    h2 = (hs_two or "")[:2]
    if h2 not in HS_ELECTRONICS_CHAPTERS:
        return None
    if raw is None:
        d = _default_electronics_compliance_dict()
        d["compliance_justification"] = "Блок electronics_compliance отсутствует в ответе модели."
        return d
    if not isinstance(raw, dict):
        d = _default_electronics_compliance_dict()
        d["compliance_justification"] = "Некорректный формат electronics_compliance в JSON."
        return d
    out = _default_electronics_compliance_dict()
    out["has_wireless_tech"] = bool(raw.get("has_wireless_tech"))
    out["has_encryption"] = bool(raw.get("has_encryption"))
    fq = raw.get("frequencies")
    if isinstance(fq, list):
        out["frequencies"] = [str(x).strip() for x in fq if str(x).strip()][:50]
    out["rf_license_required"] = bool(raw.get("rf_license_required"))
    out["fss_notification_required"] = bool(raw.get("fss_notification_required"))
    out["compliance_justification"] = str(raw.get("compliance_justification") or "").strip()[:8000]
    return out


def _gemini_google_search_retrieval_tools() -> list[Any] | None:
    """Инструмент Google Search Retrieval для ``google.generativeai`` (если доступен)."""
    try:
        import google.generativeai.protos as protos
        from google.generativeai.types import Tool

        return [Tool(google_search_retrieval=protos.GoogleSearchRetrieval())]
    except Exception as e:
        logger.warning("Google Search Retrieval tool недоступен: {}", e)
        return None


def _normalize_hs_structured_payload(data: dict[str, Any]) -> dict[str, Any]:
    if "opi_reasoning_steps" not in data:
        legacy_rs = data.get("reasoning_steps")
        data["opi_reasoning_steps"] = legacy_rs if isinstance(legacy_rs, list) else []
    if "reasoning_steps" not in data:
        data["reasoning_steps"] = data.get("opi_reasoning_steps") or []
    if "suggested_hs_code" not in data:
        data["suggested_hs_code"] = str(data.get("hs_code") or "")
    if "hs_code" not in data:
        data["hs_code"] = str(data.get("suggested_hs_code") or "")
    if "vision_insights" not in data:
        data["vision_insights"] = ""
    if "missing_info" not in data or data.get("missing_info") is None:
        data["missing_info"] = []
    elif not isinstance(data.get("missing_info"), list):
        data["missing_info"] = []
    if "supplier_question_en" not in data or data.get("supplier_question_en") is None:
        data["supplier_question_en"] = ""
    else:
        data["supplier_question_en"] = str(data.get("supplier_question_en") or "").strip()[:2000]
    return data


def _gemini_extract_text_from_rest_response(data: dict[str, Any]) -> str:
    text = ""
    for cand in data.get("candidates") or []:
        content = cand.get("content") or {}
        for part in content.get("parts") or []:
            if isinstance(part, dict) and part.get("text"):
                text += str(part.get("text") or "")
        if text.strip():
            break
    return text.strip()


def _gemini_hs_classify_google_search_raw(
    user_prompt: str,
    *,
    system_instruction: str,
    image_path: Path | None,
    max_output_tokens: int,
    temperature: float,
) -> dict[str, Any] | None:
    """
    Вызов Gemini через raw REST с инструментом ``google_search``.

    Это обходное решение для окружений, где SDK `google.generativeai` не поддерживает
    поле tools.google_search, но прокси/API уже требуют именно его.
    """
    key = (os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY") or "").strip()
    if not key:
        return None

    parts: list[dict[str, Any]] = [{"text": user_prompt}]
    if image_path is not None and image_path.is_file():
        try:
            raw = image_path.read_bytes()
            mime = mimetypes.guess_type(str(image_path))[0] or "image/jpeg"
            parts.append(
                {
                    "inline_data": {
                        "mime_type": mime,
                        "data": base64.b64encode(raw).decode("ascii"),
                    }
                }
            )
        except Exception as e:
            logger.warning("structured hs: raw web-search запрос без фото ({}): {}", image_path, e)

    payload = {
        "systemInstruction": {"parts": [{"text": system_instruction}]},
        "contents": [{"role": "user", "parts": parts}],
        "tools": [{"google_search": {}}],
        "generationConfig": {
            "temperature": temperature,
            "maxOutputTokens": max_output_tokens,
        },
    }

    url = gemini_generate_content_rest_url(resolved_gemini_model_name())
    timeout_cfg = httpx.Timeout(connect=20.0, read=90.0, write=30.0, pool=10.0)

    max_retries = _GEMINI_RATE_LIMIT_MAX_RETRIES
    attempt = 0
    while True:
        try:
            _gemini_global_rate_limit("raw_google_search_hs")
            with httpx.Client(timeout=timeout_cfg) as client:
                resp = client.post(url, params={"key": key}, json=payload)
            if resp.status_code == 429 and attempt < max_retries:
                slept = _gemini_rate_limit_sleep(attempt)
                logger.warning(
                    "Gemini 429/rate-limit (raw google_search hs), пауза {:.1f} с (попытка {}/{}).",
                    slept,
                    attempt + 1,
                    max_retries,
                )
                attempt += 1
                continue
            resp.raise_for_status()
            data = resp.json() if resp.content else {}
            raw_text = _gemini_extract_text_from_rest_response(data)
            if not raw_text:
                logger.warning("structured hs: пустой ответ Gemini (raw google_search), fallback без tools")
                return None
            parsed = _parse_hs_classify_response_json(raw_text)
            if not isinstance(parsed, dict):
                logger.warning("structured hs: невалидный JSON от raw google_search, fallback без tools")
                return None
            return _normalize_hs_structured_payload(parsed)
        except Exception as e:
            if _is_gemini_resource_exhausted(e) and attempt < max_retries:
                slept = _gemini_rate_limit_sleep(attempt)
                logger.warning(
                    "Gemini 429/rate-limit (raw google_search hs), пауза {:.1f} с (попытка {}/{}): {}",
                    slept,
                    attempt + 1,
                    max_retries,
                    e,
                )
                attempt += 1
                continue
            logger.warning("structured hs: raw Google Search call failed: {}", e)
            return None


def format_electronics_excel_cells(*, hs_code: str, electronics_compliance: Any) -> dict[str, str]:
    """Четыре колонки отчёта по электронике; для не‑84/85/90 — прочерк «—»."""
    dash = "—"
    h2 = _chapter_two_from_hs(hs_code)
    keys = (
        "Радиочастоты (МГц/ГГц)",
        "Нотификация ФСБ",
        "Лицензия РЧЦ",
        "Обоснование по электронике",
    )
    if h2 not in HS_ELECTRONICS_CHAPTERS:
        return dict.fromkeys(keys, dash)
    if not isinstance(electronics_compliance, dict):
        return dict.fromkeys(keys, dash)
    freqs = electronics_compliance.get("frequencies")
    if isinstance(freqs, list):
        ftxt = "; ".join(str(x).strip() for x in freqs if str(x).strip()) or dash
    else:
        ftxt = dash

    def _yn(val: Any) -> str:
        return "Да" if bool(val) else "Нет"

    just = str(electronics_compliance.get("compliance_justification") or "").strip() or dash
    return {
        "Радиочастоты (МГц/ГГц)": ftxt,
        "Нотификация ФСБ": _yn(electronics_compliance.get("fss_notification_required")),
        "Лицензия РЧЦ": _yn(electronics_compliance.get("rf_license_required")),
        "Обоснование по электронике": just[:8000],
    }


def _hs_classify_payload_valid(data: dict[str, Any]) -> bool:
    rs = _normalize_reasoning_steps(data.get("opi_reasoning_steps") or data.get("reasoning_steps"))
    if len(rs) < 1:
        return False
    hs = _normalize_model_hs10(data.get("suggested_hs_code") or data.get("hs_code"))
    if len(hs) != 10:
        return False
    att = data.get("attributes")
    if not isinstance(att, dict):
        return False
    att = dict(att)
    if "purpose_and_tech" not in att and "purpose" in att:
        att["purpose_and_tech"] = att.get("purpose", "")
    for k in ("name", "manufacturer", "trademark", "model", "material", "purpose_and_tech"):
        if k not in att:
            return False
    if "confidence_score" not in data:
        return False
    sq = data.get("supplier_question_en")
    if sq is not None and not isinstance(sq, str):
        return False
    return True


def _gemini_hs_classify_structured(
    user_prompt: str,
    *,
    image_path: Path | None = None,
    max_output_tokens: int = 4096,
    temperature: float = 0.12,
    rag_system_suffix: str = "",
    use_google_search: bool = False,
    fast_mode: bool = False,
) -> dict[str, Any] | None:
    """
    Запрос к Gemini с Structured Outputs (JSON Schema).

    Для ``use_google_search=True``:
    1) сначала пробуем raw REST + tools.google_search (без response_schema),
    2) при неудаче возвращаемся к structured-запросу без tools.
    """
    key = (os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY") or "").strip()
    if not key:
        return None
    try:
        import google.generativeai as genai
    except ImportError:
        return None
    configure_google_generativeai(genai, api_key=key)
    gen_cfg_structured = genai.GenerationConfig(
        temperature=temperature,
        max_output_tokens=max_output_tokens,
        response_mime_type="application/json",
        response_schema=_HS_CLASSIFY_JSON_SCHEMA,
    )

    sys_inst_base = GEMINI_HS_CLASSIFY_SYSTEM_INSTRUCTION
    if image_path is not None and image_path.is_file():
        sys_inst_base = sys_inst_base + "\n\n" + GEMINI_VISUAL_TRADEMARK_PACKING_BLOCK
    if rag_system_suffix.strip():
        sys_inst_base = sys_inst_base + "\n\n" + rag_system_suffix.strip()
    sys_inst_no_search = sys_inst_base
    if use_google_search:
        sys_inst_no_search = (
            sys_inst_no_search
            + "\n\n[Техническое примечание] Google Search может быть недоступен. "
            "Если web-поиск не доступен, снизь confidence_score и заполни missing_info."
        )
    sys_inst_with_search = sys_inst_base + ("\n\n" + GEMINI_ELECTRONICS_RULES if use_google_search else "")

    if use_google_search and not fast_mode:
        raw_tools_data = _gemini_hs_classify_google_search_raw(
            user_prompt,
            system_instruction=sys_inst_with_search,
            image_path=image_path,
            max_output_tokens=max_output_tokens,
            temperature=temperature,
        )
        if isinstance(raw_tools_data, dict):
            return raw_tools_data

    model_no_search = genai.GenerativeModel(
        resolved_gemini_model_name(),
        system_instruction=sys_inst_no_search,
    )

    max_retries = _GEMINI_RATE_LIMIT_MAX_RETRIES
    attempt = 0
    while True:
        try:
            _gemini_global_rate_limit("structured_hs")
            pil_img = None
            if image_path is not None and image_path.is_file():
                try:
                    from PIL import Image
                except ImportError:
                    logger.warning("suggest_hs_code: нужен Pillow для мультимодального structured-запроса")
                    return None
                try:
                    pil_img = Image.open(image_path)
                except Exception as e:
                    logger.warning(
                        "_gemini_hs_classify_structured: изображение не прочитано ({}), запрос без фото: {}",
                        image_path,
                        e,
                    )
            try:
                if pil_img is not None:
                    resp = model_no_search.generate_content(
                        [user_prompt, pil_img],
                        generation_config=gen_cfg_structured,
                    )
                else:
                    resp = model_no_search.generate_content(
                        user_prompt,
                        generation_config=gen_cfg_structured,
                    )
                try:
                    raw_text = (resp.text or "").strip()
                except ValueError:
                    logger.warning("suggest_hs_code: пустой/заблокированный ответ Gemini (structured)")
                    return None
                try:
                    data = _parse_hs_classify_response_json(raw_text)
                except Exception as e:
                    logger.warning("suggest_hs_code: ошибка парсинга JSON (structured): {}", e)
                    data = None
                if not isinstance(data, dict):
                    return None
                return _normalize_hs_structured_payload(data)
            finally:
                if pil_img is not None:
                    try:
                        pil_img.close()
                    except Exception:
                        pass
        except Exception as e:
            if _is_gemini_resource_exhausted(e) and attempt < max_retries:
                slept = _gemini_rate_limit_sleep(attempt)
                logger.warning(
                    "Gemini 429/rate-limit (structured hs), пауза {:.1f} с (попытка {}/{}): {}",
                    slept,
                    attempt + 1,
                    max_retries,
                    e,
                )
                attempt += 1
                continue
            logger.warning("suggest_hs_code: structured Gemini error: {}", e)
            return None


def _compose_suggest_hs_user_prompt(
    *,
    precedents_block: str,
    vision_block: str,
    name_ru: str,
    name_cn: str,
    material: str,
    usage: str,
    brand: str,
    weight_net: str,
    origin_line: str,
) -> str:
    return (
        GEMINI_PROJECT_VAT_RULES
        + GEMINI_GEOPOLITICAL_RULES
        + GEMINI_EAEU_BASE_RULES
        + "\n"
        + (precedents_block or "")
        + "Ты классификатор по ТН ВЭД ЕАЭС. По строке спецификации определи наиболее подходящий десятизначный код. "
        "Разрешены только коды с главами (первые две цифры): "
        f"{', '.join(sorted(HS_SUGGEST_ALLOWED_CHAPTERS, key=int))}.\n\n"
        f"{vision_block}"
        f"Наименование (RU): {name_ru or '—'}\n"
        f"Наименование (CN): {name_cn or '—'}\n"
        f"Материал / состав: {material or '—'}\n"
        f"Назначение / функция: {usage or '—'}\n"
        f"Бренд: {brand or '—'}\n"
        f"Нетто (если есть): {weight_net or '—'}\n"
        f"Страна происхождения (если указана): {origin_line}\n\n"
        "Заполни поля JSON по схеме (строго application/json). Сначала opi_reasoning_steps (минимум один шаг): "
        "1) факты из инвойса/фото; 2) подходящие группы ТН ВЭД; 3) сверка с прецедентами и нормативкой из контекста; "
        "4) итоговый вывод с явным указанием применённых ОПИ. Затем suggested_hs_code — ровно 10 цифр из разрешённых глав; "
        "justification — 2–6 предложений по ОПИ ТН ВЭД, с опорой на блок прецедентов ФТС, если он приведён выше; "
        "attributes — только факты из данных и фото: name, manufacturer, trademark (при фото — по блоку "
        "«ВИЗУАЛЬНЫЙ АНАЛИЗ ТОВАРНОГО ЗНАКА»; без фото — из текста/колонки бренд или ОТСУТСТВУЕТ), model, material, "
        "purpose_and_tech (назначение и тех. характеристики; при фото — данные с шильдика сюда по возможности); "
        "confidence_score — целое число, уверенность в коде от 0 до 100 (при нехватке данных снижай и заполняй missing_info); "
        "missing_info — массив коротких пунктов, что запросить у поставщика для точной классификации, или [] если данных достаточно; "
        "supplier_question_en — ОДИН технический вопрос на английском к иностранному поставщику, если данных не хватает "
        "для выбора 10 знаков; если данных хватает — пустая строка; "
        "compliance_warnings — массив строк (сертификация, ТР ТС, ДН, шифрование, ОИС и т.д.) "
        "СТРОГО в формате: \"[Название документа] - ЕСЛИ [конкретное условие]\"; "
        "запрещены размытые фразы вроде «по необходимости»/«возможно потребуется». "
        "Если условия недостаточно определены, переноси недостающие параметры в missing_info; "
        "если существенных предупреждений нет — одна строка "
        "«Существенных предупреждений по описанию нет»; vision_insights — краткий вывод по фото "
        "(см. блок визуального анализа выше), без фото — пустая строка; box_31_description — связное описание "
        "для графы 31 в стиле блоков предв. решений и примеров декларирования из системного контекста (см. инструкцию "
        "в системном промпте про официальную терминологию и обязательные характеристики).\n"
    )


def _compose_suggest_hs_legacy_prompt(
    *,
    precedents_block: str,
    vision_block: str,
    name_ru: str,
    name_cn: str,
    material: str,
    usage: str,
    brand: str,
    weight_net: str,
    origin_line: str,
) -> str:
    """Текстовый запрос без response_schema (fallback при сбое structured)."""
    base = _compose_suggest_hs_user_prompt(
        precedents_block=precedents_block,
        vision_block=vision_block,
        name_ru=name_ru,
        name_cn=name_cn,
        material=material,
        usage=usage,
        brand=brand,
        weight_net=weight_net,
        origin_line=origin_line,
    )
    legacy_tail = (
        "Верни СТРОГО один JSON-объект без markdown (как при structured output). Обязательные ключи: "
        "opi_reasoning_steps (массив строк, шаги 1–4 как в пользовательской инструкции), suggested_hs_code, "
        "justification, attributes { name, manufacturer, trademark, model, material, purpose_and_tech }, "
        "confidence_score (целое 0–100), compliance_warnings, vision_insights, box_31_description (графа 31 — см. системный промпт), "
        "missing_info (массив строк или []), supplier_question_en (строка). "
        "Для глав 84/85/90 — опционально electronics_compliance { has_wireless_tech, has_encryption, frequencies, "
        "rf_license_required, fss_notification_required, compliance_justification }; иначе null.\n"
        "Опционально suggested_description_31 — кратко ЗАГЛАВНЫМИ по алгоритму ФТС.\n\n"
        'Пример: {"opi_reasoning_steps":["1. Факты…","2. Группы ТН ВЭД…","3. ОПИ 3б...","4. Итог по ОПИ 6…"],'
        '"suggested_hs_code":"8501510000","justification":"...","attributes":{"name":"...","manufacturer":"...",'
        '"trademark":"ОТСУТСТВУЕТ","model":"...","material":"...","purpose_and_tech":"..."},'
        '"confidence_score":80,"compliance_warnings":["Существенных предупреждений по описанию нет"],'
        '"vision_insights":"","box_31_description":"...","missing_info":["уточнить напряжение питания"],'
        '"supplier_question_en":"Please confirm input voltage range and rated power of this model.",'
        '"electronics_compliance":null,"suggested_description_31":""}'
    )
    return base + legacy_tail


def _suggested_chapter_for_rag(item_data: dict[str, Any]) -> str | None:
    sc = item_data.get("suggested_chapter")
    if sc is not None and str(sc).strip():
        return str(sc).strip()
    d = re.sub(r"\D", "", str(item_data.get("declared_hs_code") or ""))[:2]
    return d if len(d) == 2 else None


def suggest_hs_code(
    item_data: dict[str, Any],
    *,
    image_path: Any = _IMG_PATH_UNSET,
    db_session: Session | None = None,
    fast_mode: bool = False,
) -> dict[str, Any]:
    """
    Перед Gemini: RAG-контекст (включая векторный поиск 3-5 релевантных прецедентов из индекса эмбеддингов);
    дополнительно сохраняется блок «=== ПРЕЦЕДЕНТЫ ФТС ===» из keyword-поиска и поле classification_precedent.
    Классификация: Structured Outputs (response_mime_type application/json) — opi_reasoning_steps, suggested_hs_code,
    justification, attributes (в т.ч. purpose_and_tech), confidence_score, compliance_warnings, vision_insights,
    box_31_description, missing_info, supplier_question_en;
    31-я графа (сборка из атрибутов) — :func:`_build_box_31` → поля Description_31 / suggested_description_31;
    готовый ИИ-текст для графы 31 — поле box_31_description.
    Для кодов из справочника ПП РФ №908/№688 — второй запрос: vat_rate_final, vat_logic.
    При ошибке structured — повтор и legacy-промпт; битый JSON обрабатывается _parse_hs_classify_response_json,
    затем при полном провале — сырой текст в графе 31 и эвристика hs_code по 10 цифрам.

    ``image_path``: явный путь к файлу изображения; иначе используются поля ``image_path`` (колонка ``Image_Path``),
    ``_image_path`` из XLSX и файлы из ``data/raw_invoices/images/`` (по артикулу ``*.jpg`` или ``row_{ExcelRow}.jpg``,
    см. :func:`_resolve_raw_invoice_sidecar_image`).
    ``db_session``: сессия SQLAlchemy для RAG (:func:`app.services.rag_retriever.build_rag_context` через :func:`_get_rag_context`); если ``None``, открывается кратковременная
    :class:`~app.db.SessionLocal`.

    Перед сборкой RAG вызывается :func:`app.services.text_processor.normalize_product_description`: в поиск прецедентов
    передаётся ``search_keywords`` или ``clean_russian_name``, а в отчёт — поле ``normalized_product_name``.
    """
    precedents_block, precedent_csv = find_classification_precedents_for_invoice_item(item_data)
    empty: dict[str, Any] = {
        "hs_code": "",
        "suggested_hs_code": "",
        "justification": "",
        "suggested_description_31": "",
        "Description_31": "",
        "box_31_description": "",
        "Vision_Insights": "",
        "compliance_warnings": [],
        "vat_rate_final": None,
        "vat_logic": "",
        "preferential_vat_group": None,
        "classification_precedent": precedent_csv or "",
        "confidence_score": None,
        "opi_reasoning_steps": [],
        "reasoning_steps": [],
        "missing_info": None,
        "supplier_question_en": "",
        "hs_attributes": {},
        "Trois_Control": "",
        "normalized_product_name": "",
        "photo_for_analysis": "Нет",
        "electronics_compliance": None,
        "registry_check": copy.deepcopy(_REGISTRY_CHECK_EMPTY),
        "fallback_status": "",
    }

    def _resolve_existing_hs10(raw_hs: Any) -> tuple[str, str]:
        own = False
        s = db_session
        if s is None:
            s = SessionLocal()
            own = True
        try:
            return _resolve_existing_hs10_with_session(s, raw_hs)
        except Exception as e:
            logger.warning("suggest_hs_code: HS registry resolve failed for {}: {}", raw_hs, e)
            return "", "lookup_failed"
        finally:
            if own and s is not None:
                s.close()

    declared_fallback_status = ""
    declared_digits = re.sub(r"\D", "", str(item_data.get("declared_hs_code") or ""))[:10]
    if declared_digits:
        if len(declared_digits) < 10:
            logger.warning(
                "suggest_hs_code: в колонке «Код» указан неполный ТН ВЭД ({}). "
                "Автодополнение нулями отключено, будет выполнен подбор через ИИ.",
                declared_digits,
            )
            declared_digits = ""
        elif _is_aggregate_hs10(declared_digits):
            logger.warning(
                "suggest_hs_code: код из колонки «Код» выглядит как агрегат (...000000): {}. "
                "Используем ИИ-подбор конечной позиции.",
                declared_digits,
            )
            declared_digits = ""
        if len(declared_digits) == 10:
            declared_resolved, declared_status = _resolve_existing_hs10(declared_digits)
            if not declared_resolved:
                logger.warning(
                    "suggest_hs_code: код из колонки «Код» отсутствует в tnved_commodities ({}), "
                    "статус={} — выполняем ИИ-подбор.",
                    declared_digits,
                    declared_status,
                )
                declared_digits = ""
            elif declared_resolved != declared_digits:
                logger.warning(
                    "suggest_hs_code: код из колонки «Код» скорректирован по реестру {} -> {} ({})",
                    declared_digits,
                    declared_resolved,
                    declared_status,
                )
                declared_digits = declared_resolved
                declared_fallback_status = "declared_hs_registry_autocorrected"
        if len(declared_digits) == 10:
            pref_group_decl = match_preferential_vat_group(declared_digits)
            attrs_decl = _default_hs_attributes_from_item(item_data)
            desc31_decl = _build_box_31_description(attrs_decl)
            cw0, tc0 = _merge_trois_registry_risks(
                ["Существенных предупреждений по описанию нет"], attrs_decl
            )
            img_decl = _validate_product_image_path(_resolve_invoice_row_image_path(item_data))
            photo_decl = _photo_analysis_report(img_decl)
            return {
                **empty,
                "hs_code": declared_digits,
                "suggested_hs_code": declared_digits,
                "justification": "Код ТН ВЭД указан в колонке «Код» спецификации (без подбора моделью).",
                "suggested_description_31": desc31_decl[:8000],
                "Description_31": desc31_decl[:8000],
                "box_31_description": desc31_decl[:8000],
                "compliance_warnings": cw0,
                "vat_rate_final": None,
                "vat_logic": "",
                "preferential_vat_group": pref_group_decl,
                "classification_precedent": precedent_csv or "",
                "confidence_score": 100.0,
                "opi_reasoning_steps": [
                    "1. Код ТН ВЭД задан в спецификации декларантом.",
                    "2. Подбор моделью не выполнялся.",
                    "3. Прецеденты и RAG не меняют объявленный код.",
                    "4. Итог: принят код из колонки «Код».",
                ],
                "reasoning_steps": [
                    "1. Код ТН ВЭД задан в спецификации декларантом.",
                    "2. Подбор моделью не выполнялся.",
                    "3. Прецеденты и RAG не меняют объявленный код.",
                    "4. Итог: принят код из колонки «Код».",
                ],
                "missing_info": None,
                "supplier_question_en": "",
                "hs_attributes": attrs_decl,
                "Vision_Insights": "",
                "Trois_Control": tc0,
                "normalized_product_name": (
                    (item_data.get("name_ru") or item_data.get("name_cn") or "").strip()
                    or str(attrs_decl.get("name") or "").strip()
                ),
                "photo_for_analysis": photo_decl,
                "electronics_compliance": None,
                "registry_check": _compute_registry_check_payload(
                    item_data, attrs_decl, declared_digits, db_session
                ),
                "fallback_status": declared_fallback_status,
            }
    name_ru = (item_data.get("name_ru") or "").strip()
    name_cn = (item_data.get("name_cn") or "").strip()
    material = (item_data.get("material") or "").strip()
    usage = (item_data.get("usage") or "").strip()
    brand = (item_data.get("brand") or "").strip()
    weight_net = (item_data.get("weight_net") or "").strip()

    if not (os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY") or "").strip():
        blob = f"{name_ru} {name_cn} {material} {usage} {brand}".lower()
        if "шампунь" in blob:
            attrs_demo = _default_hs_attributes_from_item(item_data)
            attrs_demo["name"] = name_ru or name_cn or "ШАМПУНЬ"
            desc_demo = _build_box_31_description(attrs_demo)
            cw_d, tc_d = _merge_trois_registry_risks(
                ["Существенных предупреждений по описанию нет"], attrs_demo
            )
            img_demo = _validate_product_image_path(_resolve_invoice_row_image_path(item_data))
            photo_demo = _photo_analysis_report(img_demo)
            return {
                **empty,
                "hs_code": "3304990000",
                "suggested_hs_code": "3304990000",
                "justification": "Демо-классификация без API (маркер «шампунь» → глава 33).",
                "suggested_description_31": desc_demo[:8000],
                "Description_31": desc_demo[:8000],
                "box_31_description": desc_demo[:8000],
                "compliance_warnings": cw_d,
                "confidence_score": 100.0,
                "opi_reasoning_steps": [
                    "1. По тексту инвойса выявлен маркер «шампунь».",
                    "2. Глава 33 (косметика) соответствует номенклатуре.",
                    "3. Прецеденты не использовались (демо-режим).",
                    "4. Итог: код 3304990000 для иллюстрации.",
                ],
                "reasoning_steps": [
                    "1. По тексту инвойса выявлен маркер «шампунь».",
                    "2. Глава 33 (косметика) соответствует номенклатуре.",
                    "3. Прецеденты не использовались (демо-режим).",
                    "4. Итог: код 3304990000 для иллюстрации.",
                ],
                "missing_info": None,
                "supplier_question_en": "",
                "hs_attributes": attrs_demo,
                "Vision_Insights": "",
                "Trois_Control": tc_d,
                "normalized_product_name": str(attrs_demo.get("name") or "").strip(),
                "photo_for_analysis": photo_demo,
                "electronics_compliance": None,
                "registry_check": _compute_registry_check_payload(
                    item_data, attrs_demo, "3304990000", db_session
                ),
                "fallback_status": "demo_no_api",
            }
        logger.warning("suggest_hs_code: нет GEMINI_API_KEY / GOOGLE_API_KEY")
        return {**empty, "fallback_status": "gemini_api_key_missing"}

    normalized_name_for_report = ""

    raw_warn_cell = ""
    if image_path is _IMG_PATH_UNSET:
        raw_warn_cell = (item_data.get("image_path") or item_data.get("_image_path") or "").strip()
        resolved_img = _resolve_invoice_row_image_path(item_data)
    elif image_path is None:
        resolved_img = None
    else:
        raw_warn_cell = str(image_path).strip()
        from .vision_extractor import DocumentVisionExtractor

        resolved_img = DocumentVisionExtractor().extract_image_for_row(raw_warn_cell)

    before_validate = resolved_img
    resolved_img = _validate_product_image_path(resolved_img)
    photo_for_analysis = _photo_analysis_report(resolved_img)
    if raw_warn_cell and before_validate is None:
        logger.warning(
            "Указан путь к изображению ({}), но файл не найден или расширение не поддерживается — анализ только по тексту.",
            raw_warn_cell[:500],
        )

    vision_block = ""
    if resolved_img is not None:
        vision_block = (
            "Текст строки может быть неточным; при извлечении атрибутов и выборе кода ТН ВЭД отдавай приоритет "
            "наблюдаемым на изображении фактам.\n\n"
            + GEMINI_VISUAL_DEEP_ANALYSIS_BLOCK
            + "\n\n"
            + GEMINI_VISUAL_TRADEMARK_PACKING_BLOCK
            + "\n"
        )

    origin_line = _extract_country_origin(item_data) or "—"
    ch_hint = _suggested_chapter_for_rag(item_data)
    electronics_mode = str(ch_hint or "").strip() in HS_ELECTRONICS_CHAPTERS

    user_prompt = _compose_suggest_hs_user_prompt(
        precedents_block=precedents_block,
        vision_block=vision_block,
        name_ru=name_ru,
        name_cn=name_cn,
        material=material,
        usage=usage,
        brand=brand,
        weight_net=weight_net,
        origin_line=origin_line,
    )
    if electronics_mode:
        user_prompt = (
            user_prompt
            + "\n\nОжидается классификация в главах 84, 85 или 90. Заполни объект electronics_compliance в JSON "
            "(has_wireless_tech, has_encryption, frequencies, rf_license_required, fss_notification_required, "
            "compliance_justification), используя Google Поиск для официальных спецификаций точной модели.\n"
        )

    raw_rag_text = _product_blob_for_rag(item_data) or " ".join(
        x for x in (name_ru, name_cn, material, usage, brand) if x
    ).strip()
    product_blob_rag = raw_rag_text
    if raw_rag_text.strip():
        try:
            if fast_mode:
                normalized_name_for_report = (name_ru or name_cn or "").strip()
                product_blob_rag = (normalized_name_for_report or raw_rag_text).strip()
            else:
                from .text_processor import normalize_product_description

                nd = normalize_product_description(raw_rag_text)
                normalized_name_for_report = (nd.get("clean_russian_name") or "").strip()
                kw = (nd.get("search_keywords") or "").strip()
                product_blob_rag = (kw or normalized_name_for_report or raw_rag_text).strip()
            logger.info(
                "[Нормализация] Оригинал: {} -> Очищено: {}",
                raw_rag_text,
                (normalized_name_for_report or product_blob_rag or "—"),
            )
        except Exception as e:
            logger.warning("normalize_product_description: {}", e)
            product_blob_rag = raw_rag_text

    def _build_rag_suffix() -> str:
        from .rag_retriever import RAG_CLASSIFICATION_MANDATE, get_semantic_legal_context

        try:
            semantic_top3: list[str] = []
            pref = _nt_prefix_for_rag(product_blob_rag, ch_hint) or ""
            if not pref and ch_hint:
                c = re.sub(r"\D", "", str(ch_hint).strip())[:2].zfill(2)
                if len(c) == 2 and c.isdigit():
                    pref = c
            if db_session is not None:
                rag = _get_rag_context(db_session, product_blob_rag, ch_hint)
                semantic_top3 = get_semantic_legal_context(
                    product_blob_rag,
                    db_session,
                    hs_code_prefix=pref,
                )
            else:
                with SessionLocal() as s:
                    rag = _get_rag_context(s, product_blob_rag, ch_hint)
                    semantic_top3 = get_semantic_legal_context(
                        product_blob_rag,
                        s,
                        hs_code_prefix=pref,
                    )
            semantic_block = "=== СЕМАНТИЧЕСКИЕ ПРЕЦЕДЕНТЫ (TOP-3) ===\n"
            if semantic_top3:
                semantic_block += "\n".join(f"- {line}" for line in semantic_top3[:3])
            else:
                semantic_block += "- (релевантные векторные прецеденты не найдены)"
            suffix = f"{rag}\n\n{semantic_block}\n\n{RAG_CLASSIFICATION_MANDATE}"
            logger.info(
                "suggest_hs_code: RAG context prepared (prefix_hs={!r}, name_snippet={!r}, len={})",
                pref,
                (product_blob_rag or "")[:160],
                len(suffix),
            )
            return suffix
        except Exception as e:
            logger.warning("suggest_hs_code: RAG context build failed: {!r}", e)
            return RAG_CLASSIFICATION_MANDATE

    rag_system_suffix = _build_rag_suffix()

    structured_used = False
    data: dict[str, Any] | None = _gemini_hs_classify_structured(
        user_prompt,
        image_path=resolved_img,
        max_output_tokens=4096,
        temperature=0.12,
        rag_system_suffix=rag_system_suffix,
        use_google_search=(electronics_mode and not fast_mode),
        fast_mode=fast_mode,
    )
    if not (data and _hs_classify_payload_valid(data)):
        data = _gemini_hs_classify_structured(
            user_prompt,
            image_path=resolved_img,
            max_output_tokens=4096,
            temperature=0.05,
            rag_system_suffix=rag_system_suffix,
            use_google_search=(electronics_mode and not fast_mode),
            fast_mode=fast_mode,
        )
    if data and _hs_classify_payload_valid(data):
        structured_used = True
    else:
        legacy_prompt = _compose_suggest_hs_legacy_prompt(
            precedents_block=precedents_block,
            vision_block=vision_block,
            name_ru=name_ru,
            name_cn=name_cn,
            material=material,
            usage=usage,
            brand=brand,
            weight_net=weight_net,
            origin_line=origin_line,
        )
        legacy_prompt = (
            legacy_prompt
            + "\n\n=== СИСТЕМНЫЕ ДАННЫЕ ДЛЯ МОДЕЛИ (RAG) ===\n"
            + rag_system_suffix.strip()
            + "\n"
        )
        legacy_text = ""
        try:
            legacy_text = _gemini_generate(
                legacy_prompt,
                max_output_tokens=4096,
                temperature=0.12,
                image_path=resolved_img,
            )
        except Exception as e:
            logger.warning("suggest_hs_code: legacy fallback error: {}", e)
            return empty
        data = _parse_hs_classify_response_json(legacy_text)
        if not data:
            logger.warning("suggest_hs_code: не удалось разобрать JSON (legacy), fallback: сырой текст в графе 31")
            raw_fb = (legacy_text or "").strip()[:8000]
            attrs_fb = _default_hs_attributes_from_item(item_data)
            hs_fb = ""
            m_hs = re.search(r"\b(\d{10})\b", legacy_text or "")
            if m_hs:
                hs_fb = _normalize_model_hs10(m_hs.group(1))
            legacy_fb_status = "legacy_json_parse_failed"
            if hs_fb:
                hs_fb_resolved, hs_fb_resolve_status = _resolve_existing_hs10(hs_fb)
                if not hs_fb_resolved:
                    logger.warning(
                        "suggest_hs_code: legacy fallback HS {} отсутствует в tnved_commodities ({}), очищаем код.",
                        hs_fb,
                        hs_fb_resolve_status,
                    )
                    hs_fb = ""
                    legacy_fb_status = "legacy_json_parse_failed_hs_not_in_registry"
                elif hs_fb_resolved != hs_fb:
                    logger.warning(
                        "suggest_hs_code: legacy fallback HS autocorrect {} -> {} ({})",
                        hs_fb,
                        hs_fb_resolved,
                        hs_fb_resolve_status,
                    )
                    hs_fb = hs_fb_resolved
                    legacy_fb_status = "legacy_json_parse_failed_hs_registry_autocorrected"
            cw_fb, tc_fb = _merge_trois_registry_risks([], attrs_fb)
            ec_fb: dict[str, Any] | None = None
            if len(hs_fb) == 10 and _chapter_two_from_hs(hs_fb) in HS_ELECTRONICS_CHAPTERS:
                ec_fb = _normalize_electronics_compliance_payload(
                    data.get("electronics_compliance") if isinstance(data, dict) else None,
                    hs_two=_chapter_two_from_hs(hs_fb),
                )
            return {
                **empty,
                "hs_code": hs_fb,
                "suggested_hs_code": hs_fb,
                "justification": str(empty.get("justification") or ""),
                "suggested_description_31": raw_fb,
                "Description_31": raw_fb,
                "box_31_description": "",
                "compliance_warnings": cw_fb,
                "opi_reasoning_steps": [],
                "reasoning_steps": [],
                "missing_info": None,
                "supplier_question_en": "",
                "normalized_product_name": normalized_name_for_report,
                "photo_for_analysis": photo_for_analysis,
                "hs_attributes": attrs_fb,
                "Vision_Insights": "",
                "Trois_Control": tc_fb,
                "electronics_compliance": ec_fb,
                "registry_check": _compute_registry_check_payload(
                    item_data, attrs_fb, hs_fb, db_session
                ),
                "fallback_status": legacy_fb_status,
            }

    fallback_status = ""
    hs_value_raw = data.get("suggested_hs_code") or data.get("hs_code")
    hs = _normalize_model_hs10(hs_value_raw)
    hs_raw = re.sub(r"\D", "", str(hs_value_raw or ""))[:10]
    if hs_raw and not hs:
        logger.warning(
            "suggest_hs_code: модель вернула недопустимый hs_code={} (агрегат/глава вне allow-list), помечаем как пустой.",
            hs_raw,
        )
        fallback_status = "invalid_model_hs"

    hs_registry_source = hs if len(hs) == 10 else hs_raw
    if len(hs_registry_source) == 10:
        hs_resolved, hs_resolve_status = _resolve_existing_hs10(hs_registry_source)
        if not hs_resolved:
            logger.warning(
                "suggest_hs_code: hs={} отсутствует в tnved_commodities ({}), отклоняем код.",
                hs_registry_source,
                hs_resolve_status,
            )
            hs = ""
            if not fallback_status:
                fallback_status = "hs_not_in_registry"
        else:
            if hs_resolved != hs_registry_source:
                logger.warning(
                    "suggest_hs_code: hs autocorrect {} -> {} ({})",
                    hs_registry_source,
                    hs_resolved,
                    hs_resolve_status,
                )
                if not fallback_status:
                    fallback_status = "hs_registry_autocorrected"
            hs = hs_resolved
    just = str(data.get("justification") or "").strip()
    box31_ai = str(data.get("box_31_description") or "").strip()[:8000]
    cw = _normalize_compliance_warnings(data.get("compliance_warnings"))
    vision_insights = ""
    if isinstance(data, dict):
        vision_insights = str(data.get("vision_insights") or "").strip()[:2000]

    confidence: float | None = None
    if isinstance(data, dict) and "confidence_score" in data:
        confidence = _parse_confidence_score(data.get("confidence_score"))
    opi_reasoning_steps_out: list[str] = []
    missing_info_out: list[str] | None = None
    supplier_question_en_out = ""
    if isinstance(data, dict):
        opi_reasoning_steps_out = _normalize_reasoning_steps(
            data.get("opi_reasoning_steps") or data.get("reasoning_steps")
        )
        mi_list = _normalize_missing_info_list(data.get("missing_info"))
        missing_info_out = mi_list if mi_list else None
        supplier_question_en_out = str(data.get("supplier_question_en") or "").strip()[:2000]
    if not structured_used and not opi_reasoning_steps_out:
        opi_reasoning_steps_out = [
            "1. Ответ модели без structured-схемы (legacy).",
            "2. Проверьте полноту JSON при необходимости.",
        ]
    if structured_used and missing_info_out and not supplier_question_en_out:
        supplier_question_en_out = (
            "Please provide missing technical parameters (material, operating principle, power/voltage and exact model) "
            "to finalize the 10-digit HS classification."
        )
    attrs_norm = _normalize_hs_attributes_from_llm(data.get("attributes")) if structured_used else {}
    desc31_struct = _build_box_31(attrs_norm) if structured_used else ""

    if not structured_used:
        desc31_legacy = str(data.get("suggested_description_31") or "").strip()
        if isinstance(data.get("attributes"), dict):
            attrs_norm = _normalize_hs_attributes_from_llm(data.get("attributes"))
        else:
            attrs_norm = _default_hs_attributes_from_item(item_data)
        if desc31_legacy:
            attrs_norm["name"] = desc31_legacy[:512]
        desc31_struct = desc31_legacy or _build_box_31(attrs_norm)
    elif not (desc31_struct or "").strip():
        desc31_struct = _build_box_31(_default_hs_attributes_from_item(item_data))

    cw, trois_cell = _merge_trois_registry_risks(cw, attrs_norm)
    cw, trois_cell = _append_photo_intellectual_property_trademark_risks(
        cw,
        trois_cell,
        attrs_norm,
        had_image_in_request=resolved_img is not None,
    )
    cw = _sanitize_compliance_warnings_for_synthetic_items(
        cw, item_data=item_data, attrs=attrs_norm
    )
    hard_cw = _hard_regulatory_risk_warnings(item_data, hs or hs_raw)
    if hard_cw:
        existing = {str(x).strip().casefold() for x in cw if str(x).strip()}
        for line in hard_cw:
            t = str(line or "").strip()
            if not t:
                continue
            if t.casefold() in existing:
                continue
            cw.append(t)
            existing.add(t.casefold())

    if len(hs) < 10:
        try:
            src = json.dumps(data, ensure_ascii=False) if data else ""
        except TypeError:
            src = str(data)
        m = re.search(r"\b(\d{10})\b", src)
        if m:
            hs = _normalize_model_hs10(m.group(1))
    electronics_out: dict[str, Any] | None = None
    if len(hs) == 10 and _chapter_two_from_hs(hs) in HS_ELECTRONICS_CHAPTERS:
        electronics_out = _normalize_electronics_compliance_payload(
            data.get("electronics_compliance") if isinstance(data, dict) else None,
            hs_two=_chapter_two_from_hs(hs),
        )
    if len(hs) != 10:
        if not fallback_status:
            fallback_status = "hs_not_determined"
        return {
            "hs_code": "",
            "suggested_hs_code": "",
            "justification": just[:4000],
            "suggested_description_31": desc31_struct[:8000],
            "Description_31": desc31_struct[:8000],
            "box_31_description": box31_ai,
            "compliance_warnings": cw,
            "vat_rate_final": None,
            "vat_logic": "",
            "preferential_vat_group": None,
            "classification_precedent": precedent_csv or "",
            "confidence_score": confidence,
            "opi_reasoning_steps": opi_reasoning_steps_out,
            "reasoning_steps": opi_reasoning_steps_out,
            "missing_info": missing_info_out,
            "supplier_question_en": supplier_question_en_out,
            "hs_attributes": attrs_norm,
            "Vision_Insights": vision_insights if structured_used else "",
            "Trois_Control": trois_cell,
            "normalized_product_name": normalized_name_for_report,
            "photo_for_analysis": photo_for_analysis,
            "electronics_compliance": electronics_out,
            "registry_check": _compute_registry_check_payload(
                item_data, attrs_norm, hs, db_session
            ),
            "fallback_status": fallback_status,
        }

    if structured_used and confidence is not None and confidence < 70.0:
        if _LOW_CONFIDENCE_HS_SUFFIX.strip() not in just:
            just = (just + _LOW_CONFIDENCE_HS_SUFFIX).strip()

    pref_group = match_preferential_vat_group(hs)
    vat_final: int | float | None = None
    vat_logic_out = ""
    if pref_group is not None:
        ve = gemini_vat_expertise_preferential(hs, item_data, group=pref_group, image_path=resolved_img)
        vat_final = ve.get("vat_rate_final")
        vat_logic_out = str(ve.get("vat_logic") or "").strip()[:8000]

    return {
        "hs_code": hs,
        "suggested_hs_code": hs,
        "justification": just[:4000],
        "suggested_description_31": desc31_struct[:8000],
        "Description_31": desc31_struct[:8000],
        "box_31_description": box31_ai,
        "compliance_warnings": cw,
        "vat_rate_final": vat_final,
        "vat_logic": vat_logic_out,
        "preferential_vat_group": pref_group,
        "classification_precedent": precedent_csv or "",
        "confidence_score": confidence,
        "opi_reasoning_steps": opi_reasoning_steps_out,
        "reasoning_steps": opi_reasoning_steps_out,
        "missing_info": missing_info_out,
        "supplier_question_en": supplier_question_en_out,
        "hs_attributes": attrs_norm,
        "Vision_Insights": vision_insights if structured_used else "",
        "Trois_Control": trois_cell,
        "normalized_product_name": normalized_name_for_report,
        "photo_for_analysis": photo_for_analysis,
        "electronics_compliance": electronics_out,
        "registry_check": _compute_registry_check_payload(
            item_data, attrs_norm, hs, db_session
        ),
        "fallback_status": fallback_status,
    }


def _hard_regulatory_risk_warnings(item_data: dict[str, Any] | None, hs_code: str) -> list[str]:
    """
    Детерминированные (не-LLM) предупреждения по:
    - geo_special_duties (антидемпинг/повышенная ставка) с учётом страны происхождения;
    - sanction_import_risks;
    - акцизу из hs_rates.
    """
    hs = re.sub(r"\D", "", str(hs_code or ""))[:10]
    if len(hs) < 4:
        return []

    origin_raw = _extract_country_origin(item_data or {})
    origin_iso = _normalize_country_iso(origin_raw)
    geo = InvoiceAnalyzer().check_geopolitical_risks(hs, origin_raw)
    manufacturer = ""
    if item_data:
        manufacturer = (
            str(
                item_data.get("manufacturer")
                or item_data.get("brand")
                or item_data.get("name_ru")
                or item_data.get("name_cn")
                or ""
            )
            .strip()
        )
    sanctions = check_sanction_risks(hs, manufacturer, origin_raw, None)

    lines: list[str] = []
    if origin_iso and (geo.get("duty_rate_override") is not None) and not geo.get("embargo_blocked"):
        rate_txt = _format_geopolitical_duty_for_parse(geo.get("duty_rate_override"))
        rate_txt = rate_txt.replace("%", "").strip() if isinstance(rate_txt, str) else str(rate_txt)
        note = (geo.get("geopolitical_duty_note") or "").strip()
        tail = f" Основание: {note[:220]}." if note else ""
        lines.append(
            f"!!! [АНТИДЕМПИНГ] Ставка: {rate_txt}% (HS {hs}, Origin {origin_iso}).{tail}"
        )

    sanc_status = (geo.get("Sanction_Status") or "").strip()
    sanc_risk = (geo.get("sanction_risk") or "").strip()
    if sanc_status in {"Риск", "Запрещено"} or sanc_risk:
        if sanc_status == "Запрещено":
            msg = "[САНКЦИОННЫЙ РИСК] Товар запрещен к ввозу/вывозу по санкционным ограничениям."
        else:
            msg = "[САНКЦИОННЫЙ РИСК] Товар ограничен к ввозу/вывозу; требуется усиленная проверка."
        if sanc_risk:
            msg = f"{msg} Детали: {sanc_risk[:260]}"
        lines.append(f"!!! {msg}")
    for a in sanctions.get("alerts") or []:
        t = str(a or "").strip()
        if t:
            lines.append(f"!!! {t}")
    for cr in sanctions.get("country_rules") or []:
        t = str(cr or "").strip()
        if t:
            lines.append(f"[СТРАНОВОЕ ПРАВИЛО] {t}")

    try:
        with SessionLocal() as db:
            rate = _best_hs_rate(db, hs)
            if rate is not None:
                ex_type = (rate.excise_type or "none").strip().lower()
                ex_val = float(rate.excise_value or 0.0)
                if ex_type in {"percent", "fixed"} and ex_val > 0.0:
                    basis = (rate.excise_basis or "").strip()
                    if ex_type == "percent":
                        msg = f"[АКЦИЗ] Применяется ставка {ex_val:g}% от таможенной стоимости."
                    else:
                        msg = f"[АКЦИЗ] Применяется фиксированная ставка {ex_val:g} за единицу."
                    if basis:
                        msg = f"{msg} Основание: {basis[:260]}"
                    lines.append(msg)
    except Exception as e:
        logger.warning("_hard_regulatory_risk_warnings: {}", e)

    out: list[str] = []
    seen: set[str] = set()
    for line in lines:
        t = str(line or "").strip()
        if not t:
            continue
        k = t.casefold()
        if k in seen:
            continue
        seen.add(k)
        out.append(t)
    return out


def analyze_item_risks(item_data: dict[str, Any], hs_code: str) -> str:
    """
    Оценка рисков через Gemini: ОИС/бренд, военное или двойного назначения, шифрование и иные разрешения.
    Возвращает связный текст для колонки ai_risk_notes.
    При наличии фото (``Image_Path`` / ``_image_path``) — мультимодальный запрос (логотипы на фото → риск РОИС/ОИС).
    """
    if not (os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY") or "").strip():
        hard_only = _hard_regulatory_risk_warnings(item_data, hs_code)
        return "\n".join(hard_only).strip()

    name_ru = (item_data.get("name_ru") or "").strip()
    name_cn = (item_data.get("name_cn") or "").strip()
    brand = (item_data.get("brand") or "").strip()
    material = (item_data.get("material") or "").strip()
    usage = (item_data.get("usage") or "").strip()
    hs = re.sub(r"\D", "", hs_code or "")[:10]

    resolved_img = _validate_product_image_path(_resolve_invoice_row_image_path(item_data))
    hard_risks = _hard_regulatory_risk_warnings(item_data, hs)

    vision_risk = ""
    if resolved_img is not None:
        vision_risk = (
            "К запросу приложено фото товара. Оцени видимые логотипы, словесные и фигуративные знаки на изделии/упаковке. "
            "Если по фото виден известный бренд (часы, электроника и т.п.), а в тексте бренд не указан или указан слабо, "
            "в risk_notes добавь явное предупреждение в духе: "
            "«ВНИМАНИЕ: Возможен объект интеллектуальной собственности (РОИС). Проверьте право на использование товарного знака». "
            "Сопоставляй фото с пунктом про ОИС/контрафакт.\n\n"
            + GEMINI_VISUAL_TRADEMARK_PACKING_BLOCK
            + "\n\n"
        )

    prompt = (
        GEMINI_PROJECT_VAT_RULES
        + GEMINI_EAEU_BASE_RULES
        + "\n"
        + "Ты таможенный аналитик по рискам. Оцени строку товара для ввоза в ЕАЭС.\n\n"
        f"{vision_risk}"
        f"ТН ВЭД (подобранный): {hs or 'не указан'}\n"
        f"Наименование RU: {name_ru or '—'}\n"
        f"Наименование CN: {name_cn or '—'}\n"
        f"Бренд (если есть): {brand or '—'}\n"
        f"Материал: {material or '—'}\n"
        f"Назначение: {usage or '—'}\n\n"
        "Проверь и кратко опиши на русском в одном поле risk_notes (5–12 предложений или короткий список абзацев):\n"
        "1) Может ли товар затрагивать объекты интеллектуальной собственности (контрафакт бренда, если бренд указан).\n"
        "2) Нужны ли особые разрешения: военная продукция, двойного назначения, криптография/шифрование, "
        "радиоэлектроника, санитарные/вет/фито меры — только если уместно по описанию.\n"
        "Если рисков не видно, явно напиши «Существенных рисков по описанию не выявлено» и укажи оговорку "
        "о необходимости сверки с фактом и документами.\n\n"
        'Верни СТРОГО один JSON без markdown: {"risk_notes":"..."}'
    )
    try:
        text = _gemini_generate(
            prompt,
            max_output_tokens=1536,
            temperature=0.2,
            image_path=resolved_img,
        )
    except Exception as e:
        logger.warning("analyze_item_risks: {}", e)
        tail = "SOFT-FALLBACK: AI-анализ рисков временно недоступен из-за сбоя LLM; требуется ручная проверка."
        if hard_risks:
            return "\n".join([*hard_risks, tail]).strip()
        return tail
    data = _extract_json_object(text)
    notes = str(data.get("risk_notes") or data.get("ai_risk_notes") or "").strip()
    notes = _sanitize_risk_notes_for_synthetic_items(notes, item_data)
    if hard_risks:
        notes = "\n".join([*hard_risks, notes] if notes else [*hard_risks]).strip()
    if len(notes) > 12000:
        return notes[:11900] + "…"
    return notes


def _best_hs_rate(db: Session, code10: str) -> HsRate | None:
    """Строка hs_rates с самым длинным hs_prefix, который является префиксом code10."""
    p = re.sub(r"\D", "", code10)[:10]
    if len(p) < 4:
        return None
    exact = db.query(HsRate).filter(HsRate.hs_code == p).first()
    if exact:
        return exact
    best: HsRate | None = None
    best_len = -1
    for r in db.query(HsRate).filter(HsRate.hs_prefix.isnot(None)).all():
        pref = (r.hs_prefix or "").strip()
        if not pref or not p.startswith(pref):
            continue
        lp = len(pref)
        if lp > best_len:
            best = r
            best_len = lp
    return best


_SWEET_DRINK_EXCISE_2026_RUB_PER_L = 11.0


def _item_text_blob_for_excise(item_data: dict[str, Any] | None) -> str:
    if not item_data:
        return ""
    return " ".join(
        str(item_data.get(k) or "")
        for k in ("name_ru", "name_cn", "material", "usage", "description", "composition")
    ).casefold()


def _has_sugar_or_sweetener(item_data: dict[str, Any] | None) -> bool:
    b = _item_text_blob_for_excise(item_data)
    if not b:
        return False
    markers = (
        "сахар",
        "сахароз",
        "подсласт",
        "sweetener",
        "sugar",
        "sucralose",
        "аспартам",
        "фруктоз",
        "глюкоз",
        "сироп",
    )
    return any(m in b for m in markers)


def _apply_sweet_drink_excise_2026(out: dict[str, Any], hs_code: str, item_data: dict[str, Any] | None) -> None:
    hs = re.sub(r"\D", "", str(hs_code or ""))[:10]
    if not hs.startswith("2202"):
        return
    if not _has_sugar_or_sweetener(item_data):
        return
    out["excise_type"] = "fixed"
    out["excise_value"] = float(_SWEET_DRINK_EXCISE_2026_RUB_PER_L)
    out["excise_basis"] = "2026: сахаросодержащие напитки — 11 руб/л (актуализировать по НК РФ)."
    nt = out.setdefault("non_tariff", [])
    warn = {
        "measure_type": "excise_alert",
        "document_required": "Оформление только на акцизных таможенных постах (Приказ Минфина №27н).",
        "description": "Для сахаросодержащих напитков требуется оформление на акцизном таможенном посту.",
        "regulatory_act": "Приказ Минфина РФ №27н",
    }
    key = (warn["measure_type"], warn["document_required"])
    exists = any(
        isinstance(x, dict)
        and (str(x.get("measure_type") or ""), str(x.get("document_required") or "")) == key
        for x in nt
    )
    if not exists:
        nt.insert(0, warn)


def _apply_geo_embargo_to_enrichment(out: dict[str, Any], geo: dict[str, Any]) -> None:
    """Помечает строку как заблокированную эмбарго; не дублирует запись в non_tariff."""
    if not geo.get("embargo_blocked"):
        return
    out["duty_rate"] = "ЗАПРЕТ ВВОЗА"
    out["import_embargo"] = True
    out["base_duty_amount"] = None
    out["vat_amount"] = None
    out["total_tax_pay"] = None
    nt = out.setdefault("non_tariff", [])
    if nt and isinstance(nt[0], dict) and nt[0].get("measure_type") == "embargo":
        return
    basis = (geo.get("embargo_document_basis") or geo.get("geopolitical_duty_note") or "").strip()
    link = (geo.get("embargo_document_link") or "").strip()
    nt.insert(
        0,
        {
            "measure_type": "embargo",
            "document_required": "Запрет ввоза (geo_special_duties)",
            "description": (basis[:500] if basis else "Ввоз запрещён нормативными мерами"),
            "regulatory_act": (link or (geo.get("Applied_Special_Duty") or ""))[:255],
        },
    )


def enrich_with_customs_data(
    hs_code: str,
    item_data: dict[str, Any] | None = None,
    *,
    vat_import_override: float | None = None,
) -> dict[str, Any]:
    """
    Ставки из hs_rates и нетарифные меры из non_tariff_measures для 10-значного кода.

    Поле ``duty_rate`` для отчёта декларанту — **текст исходной нормы** (как в законе/ЕТТ) или ``"{число}%"`` из БД.
    ``base_duty_amount``: адвалор от таможенной стоимости, специфика €/кг × нетто × курс EUR (ЦБ), правила
    MAX («не менее»), ADD («+»), STANDARD — см. ``_parse_duty_rate`` / ``_compute_base_duty_rub``.
    ``vat_amount`` = (таможенная стоимость + ``base_duty_amount``) × ставка НДС (10% / 22% по множителям проекта).
    ``total_tax_pay`` = пошлина + НДС.
    """
    p = re.sub(r"\D", "", hs_code or "")[:10]
    out: dict[str, Any] = {
        "hs_code": p,
        "duty_rate": None,
        "vat_import_rate": None,
        "vat_rule_basis": "",
        "excise_type": None,
        "excise_value": None,
        "excise_basis": "",
        "valid_from": "",
        "non_tariff": [],
        "customs_value": None,
        "customs_value_base": None,
        "base_duty_amount": None,
        "vat_amount": None,
        "total_tax_pay": None,
        "Country_Risk_Status": "Нейтральная",
        "Applied_Special_Duty": "",
        "Required_Certificates": "Сертификат происхождения — по общему порядку при необходимости.",
        "Sanction_Status": "Безопасно",
        "sanction_risk": "",
        "geopolitical_duty_note": "",
        "import_embargo": False,
    }
    if len(p) == 10:
        with SessionLocal() as _db_hs:
            p_resolved, p_resolve_status = _resolve_existing_hs10_with_session(_db_hs, p)
        if not p_resolved:
            out["hs_code"] = ""
            out["duty_rate"] = "КОД ТН ВЭД НЕ НАЙДЕН В СПРАВОЧНИКЕ"
            out["non_tariff"].append(
                {
                    "measure_type": "invalid_hs",
                    "document_required": "Код ТН ВЭД отсутствует в локальном справочнике",
                    "description": (
                        f"Код {p} не найден в tnved_commodities (status={p_resolve_status}). "
                        "Проверьте классификацию и синхронизацию справочника."
                    )[:500],
                    "regulatory_act": "",
                }
            )
            return out
        if p_resolved != p:
            out["non_tariff"].append(
                {
                    "measure_type": "hs_registry_autocorrect",
                    "document_required": "Код ТН ВЭД скорректирован по локальному справочнику",
                    "description": f"{p} -> {p_resolved} ({p_resolve_status})",
                    "regulatory_act": "",
                }
            )
            p = p_resolved
            out["hs_code"] = p
    geo = InvoiceAnalyzer().check_geopolitical_risks(p, _extract_country_origin(item_data))
    out.update(
        {
            "Country_Risk_Status": geo.get("Country_Risk_Status") or out["Country_Risk_Status"],
            "Applied_Special_Duty": geo.get("Applied_Special_Duty") or "",
            "Required_Certificates": geo.get("Required_Certificates") or out["Required_Certificates"],
            "Sanction_Status": geo.get("Sanction_Status") or out["Sanction_Status"],
            "sanction_risk": geo.get("sanction_risk") or "",
            "geopolitical_duty_note": geo.get("geopolitical_duty_note") or "",
            "import_embargo": bool(geo.get("embargo_blocked")),
        }
    )
    _apply_geo_embargo_to_enrichment(out, geo)
    if len(p) < 10:
        return out

    with SessionLocal() as db:
        exact_row = db.query(HsRate).filter(HsRate.hs_code == p).first()
        rate = _best_hs_rate(db, p)
        if rate is not None and not out.get("import_embargo"):
            dr_db = str(rate.duty_rate).strip() if rate.duty_rate is not None else ""
            out["duty_rate"] = dr_db if dr_db else "0"
            out["vat_import_rate"] = float(rate.vat_import_rate)
            out["excise_type"] = rate.excise_type or "none"
            out["excise_value"] = float(rate.excise_value or 0.0)
            out["excise_basis"] = (rate.excise_basis or "")[:2000]
            out["valid_from"] = (rate.valid_from or "")[:20]
            if exact_row is None:
                out["non_tariff"].append(
                    {
                        "measure_type": "rate_fallback",
                        "document_required": "Ставка взята по ближайшему префиксу hs_rates",
                        "description": f"Для кода {p} использована ставка по префиксу {rate.hs_prefix or rate.hs_code or ''}.",
                        "regulatory_act": (rate.vat_rule_basis or "")[:255],
                    }
                )
        elif len(p) == 10 and not out.get("import_embargo"):
            # Нет ни точной ставки, ни префиксной ставки.
            out["duty_rate"] = "ТРЕБУЕТСЯ СИНХРОНИЗАЦИЯ"
            out["vat_import_rate"] = None

        measures = (
            db.query(NonTariffMeasure)
            .filter(NonTariffMeasure.commodity_code == p)
            .order_by(NonTariffMeasure.id)
            .limit(50)
            .all()
        )
        for m in measures:
            out["non_tariff"].append(
                {
                    "measure_type": m.measure_type,
                    "document_required": (m.document_required or "")[:255],
                    "description": (m.description or "")[:500],
                    "regulatory_act": (m.regulatory_act or "")[:255],
                }
            )

        try:
            vat_rate_resolved, vat_basis = resolve_vat_rate_for_hs(p, db)
            out["vat_rule_basis"] = (vat_basis or "")[:255]
            if vat_import_override is None:
                out["vat_import_rate"] = float(vat_rate_resolved)
        except Exception as e:
            logger.warning("resolve_vat_rate_for_hs: {}", e)

        try:
            apply_compliance_resolution_to_enrichment(out, p, item_data, db)
        except Exception as e:
            logger.warning("apply_compliance_resolution_to_enrichment: {}", e)

    if vat_import_override is not None and vat_import_override in (10.0, float(DEFAULT_VAT_RATE)):
        out["vat_import_rate"] = float(vat_import_override)

    _apply_sweet_drink_excise_2026(out, p, item_data)

    _apply_geo_embargo_to_enrichment(out, geo)

    dr0 = out.get("duty_rate")
    duty_overridable = False
    if not geo.get("embargo_blocked"):
        if isinstance(dr0, (int, float)) and not isinstance(dr0, bool):
            duty_overridable = True
        elif isinstance(dr0, str):
            s0 = dr0.strip()
            if s0 and "ЗАПРЕТ" not in s0 and "ТРЕБУЕТСЯ СИНХРОНИЗАЦИЯ" not in s0:
                duty_overridable = True
            elif "СИНХРОНИЗАЦИЯ" in s0:
                duty_overridable = True
    if (
        not geo.get("embargo_blocked")
        and geo.get("duty_rate_override") is not None
        and duty_overridable
    ):
        out["duty_rate"] = _format_geopolitical_duty_for_parse(geo.get("duty_rate_override"))
        note = (geo.get("geopolitical_duty_note") or "").strip()
        if note:
            out["geopolitical_duty_note"] = note
            out["non_tariff"].insert(
                0,
                {
                    "measure_type": "geopolitical_duty",
                    "document_required": "Повышенная пошлина (geo_special_duties)",
                    "description": note[:500],
                    "regulatory_act": (geo.get("Applied_Special_Duty") or "ПП РФ №2140")[:255],
                },
            )

    vr_raw = out.get("vat_import_rate")
    if vr_raw is None:
        out["vat_import_rate"] = float(DEFAULT_VAT_RATE)
    else:
        try:
            vr_num = float(vr_raw)
            if abs(vr_num - 20.0) < 1e-9:
                out["vat_import_rate"] = float(DEFAULT_VAT_RATE)
        except (TypeError, ValueError):
            out["vat_import_rate"] = float(DEFAULT_VAT_RATE)

    from .currency_sync import CurrencyService

    eur_rate = float(CurrencyService.get_eur_rate())

    customs_val = _customs_value_from_line(item_data)
    dr = out.get("duty_rate")
    vr = out.get("vat_import_rate")

    def _push_duty_warnings(lines: list[str]) -> None:
        for w in lines:
            if not (w or "").strip():
                continue
            out["non_tariff"].insert(
                0,
                {
                    "measure_type": "duty_calculation",
                    "document_required": w.strip()[:255],
                    "description": w.strip()[:500],
                    "regulatory_act": "",
                },
            )

    if customs_val is not None and vr is not None and isinstance(vr, (int, float)):
        vr_f = float(vr)
        if abs(vr_f - 10.0) < 1e-9:
            vat_pct = VAT_IMPORT_MULTIPLIER_PREFERENTIAL
        elif abs(vr_f - float(DEFAULT_VAT_RATE)) < 1e-9:
            vat_pct = VAT_IMPORT_MULTIPLIER_STANDARD
        else:
            vat_pct = vr_f / 100.0

        calc_ok = False
        if dr is not None:
            if isinstance(dr, str):
                ds = dr.strip()
                if ds and "ТРЕБУЕТСЯ" not in ds and "ЗАПРЕТ" not in ds:
                    calc_ok = True
            elif isinstance(dr, (int, float)) and not isinstance(dr, bool):
                calc_ok = True

        if calc_ok:
            parsed = _parse_duty_rate(dr)
            wn = _parse_number((item_data or {}).get("weight_net"))
            wg = _parse_number((item_data or {}).get("weight_gross"))
            base_duty, warns = _compute_base_duty_rub(parsed, float(customs_val), wn, wg, eur_rate)
            _push_duty_warnings(warns)
            vat_amt = round((float(customs_val) + base_duty) * vat_pct, 2)
            out["customs_value"] = customs_val
            out["customs_value_base"] = customs_val
            out["base_duty_amount"] = base_duty
            out["vat_amount"] = vat_amt
            out["total_tax_pay"] = round(base_duty + vat_amt, 2)

    # Safe bridge: прикладываем новый структурированный профиль мер (payment_profile),
    # но не позволяем ему уронить legacy enrichment-пайплайн.
    try:
        if len(p) == 10 and customs_val is not None and float(customs_val) > 0:
            # Локальный импорт, чтобы избежать циклического импорта на уровне модуля.
            from .payment_profile_builder import build_full_payment_profile

            country_origin = _extract_country_origin(item_data)
            profile_payload: dict[str, Any] = {
                "hs_code": p,
                "customs_value": float(customs_val),
                "freight": 0.0,
                "country": country_origin,
            }
            profile = build_full_payment_profile(
                payload=profile_payload,
                hs_code=p,
                country=country_origin,
                item_data=item_data,
            )
            out["payment_profile"] = profile.model_dump()
    except Exception as e:
        logger.warning("payment_profile bridge skipped: {}", e)

    _sanitize_enrichment_certs_and_non_tariff(out, item_data)
    return out


def format_duty(enrichment: dict[str, Any]) -> str:
    dr = enrichment.get("duty_rate")
    if dr is None:
        return "—"
    if isinstance(dr, str):
        return dr
    try:
        return f"{float(dr):g}%"
    except (TypeError, ValueError):
        return str(dr)


def format_non_tariff(enrichment: dict[str, Any], *, max_items: int | None = 5) -> str:
    """max_items=None — все меры (для Excel-отчёта); иначе сокращённый вывод для консоли."""
    items = enrichment.get("non_tariff") or []
    if not items:
        return "—"
    slice_ = items if max_items is None else items[: max(0, int(max_items))]
    parts: list[str] = []
    for it in slice_:
        t = it.get("measure_type") or ""
        doc = it.get("document_required") or it.get("description") or ""
        parts.append(f"{t}: {doc[:500]}".strip())
    if max_items is not None and len(items) > len(slice_):
        parts.append(f"… ещё {len(items) - len(slice_)}")
    out = "; ".join(parts)
    if len(out) > 32000:
        return out[:31900] + "… [обрезано для Excel]"
    return out


# --- Чекпоинты и оформление итогового Excel (скрипт test_invoice_parsing) ---

_INVOICE_WRAP_COLUMNS = (
    "Нормализованное наименование",
    "Фото",
    "Радиочастоты (МГц/ГГц)",
    "Нотификация ФСБ",
    "Лицензия РЧЦ",
    "Обоснование по электронике",
    "ГРАФА_31",
    "Description_31",
    "Графа 31 (Готовое описание)",
    "Ставка экосбора",
    "Vision_Insights",
    "Calculation_Notes",
    "Trois_Control",
    "Проверка реестров: статус",
    "Проверка реестров: номер",
    "Проверка реестров: срок/статус",
    "Проверка реестров: рекомендация",
    "Проверка реестров (СГР): статус",
    "Проверка реестров (СГР): номер",
    "Проверка реестров (СГР): дата/статус",
    "Проверка реестров (СГР): рекомендация",
    "non_tariff_measures",
    "ai_risk_notes",
    "ОБОСНОВАНИЕ_НДС",
    "classification_precedent",
    "ОПИ шаги (LLM)",
    "Вопрос поставщику (EN)",
    "Required_Certificates",
    "sanction_risk",
    "geopolitical_duty_note",
    "Запросить у поставщика",
)
_VAT_10_FILL = "C6EFCE"  # светло-зелёный для ставки НДС 10%
_AI_CONFIDENCE_LOW_FILL = "FFF2CC"  # светло-жёлтый: уверенность ИИ < 80%


def _invoice_excel_vat_baseline_note_text() -> str:
    return (
        f"Примечание: расчёт суммы НДС (поле vat_amount) и итоговых платежей выполнен по формуле "
        f"(таможенная стоимость customs_value + пошлина) × {VAT_IMPORT_MULTIPLIER_STANDARD:.2f} при базовой ставке {DEFAULT_VAT_RATE}% "
        "(льготная 10% — только при явном обосновании по ПП РФ в соответствующих строках). "
        "Пошлина и НДС считаются от customs_value (не от цены товара без фрахта/Incoterms). "
        "В ячейках vat_amount заданы формулы Excel с множителем 0.22 (или 0.1 при ставке 10%)."
    )


_INVOICE_SUMMARY_NUMERIC = (
    "weight_net",
    "quantity",
    "total_cost_estimate",
    "item_price_rub",
    "allocated_freight_rub",
    "customs_value",
    "customs_value_base",
    "landed_cost_freight_addon",
    "eur_rate_cb",
    "base_duty_amount",
    "vat_amount",
    "total_tax_pay",
    "unit_landed_cost",
)
_WARNING_FILL = "FFC7CE"


def _apply_invoice_excel_base_duty_formulas(ws, header_map: dict[str, int], header_row: int) -> None:
    """
    Если в ячейке «Ставка пошлины» число — подставляет формулу customs_value * duty_rate / 100.
    Если ставка задана текстом нормы (не число), формула не пишется: остаётся рассчитанная сумма из пайплайна.
    """
    from openpyxl.utils.cell import get_column_letter

    cv_c = header_map.get("customs_value") or header_map.get("customs_value_base")
    dr_c = header_map.get("duty_rate")
    bd_c = header_map.get("base_duty_amount")
    if not (cv_c and dr_c and bd_c):
        return
    cv_l = get_column_letter(cv_c)
    dr_l = get_column_letter(dr_c)
    first_data_row = header_row + 1
    for r in range(first_data_row, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").strip().upper() == "ИТОГО":
            continue
        dr_v = ws.cell(row=r, column=dr_c).value
        if isinstance(dr_v, str) and ("СИНХРОН" in str(dr_v).upper() or "ТРЕБУЕТСЯ" in str(dr_v).upper()):
            continue
        try:
            float(dr_v)
        except (TypeError, ValueError):
            continue
        ws.cell(row=r, column=bd_c).value = f"={cv_l}{r}*{dr_l}{r}/100"


def _apply_invoice_excel_vat_amount_formulas(ws, header_map: dict[str, int], header_row: int) -> None:
    """
    Подставляет в столбец vat_amount формулу Excel: (customs_value + duty) * 0.22
    или * 0.1 при vat_rate = 10 (льгота).
    """
    from openpyxl.utils.cell import get_column_letter

    cv_c = header_map.get("customs_value") or header_map.get("customs_value_base")
    bd_c = header_map.get("base_duty_amount")
    vr_c = header_map.get("vat_rate")
    va_c = header_map.get("vat_amount")
    if not (cv_c and bd_c and vr_c and va_c):
        return
    cv_l = get_column_letter(cv_c)
    bd_l = get_column_letter(bd_c)
    vr_l = get_column_letter(vr_c)
    first_data_row = header_row + 1
    std = VAT_IMPORT_MULTIPLIER_STANDARD
    pref = VAT_IMPORT_MULTIPLIER_PREFERENTIAL
    for r in range(first_data_row, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").strip().upper() == "ИТОГО":
            continue
        duty_v = ws.cell(row=r, column=bd_c).value
        cust_v = ws.cell(row=r, column=cv_c).value
        if isinstance(duty_v, str) and "СИНХРОН" in str(duty_v).upper():
            continue
        try:
            float(cust_v)
        except (TypeError, ValueError):
            continue
        # duty может быть формулой — не требуем float(duty_v) до записи vat
        ws.cell(row=r, column=va_c).value = (
            f"=({cv_l}{r}+{bd_l}{r})*IF({vr_l}{r}=10,{pref},{std})"
        )


def _apply_invoice_excel_unit_landed_formulas(ws, header_map: dict[str, int], header_row: int) -> None:
    """
    Себестоимость 1 ед. на складе:
    (customs_value + base_duty_amount + vat_amount + landed_cost_freight_addon) / quantity
    """
    from openpyxl.utils.cell import get_column_letter

    cv_c = header_map.get("customs_value") or header_map.get("customs_value_base")
    bd_c = header_map.get("base_duty_amount")
    va_c = header_map.get("vat_amount")
    fa_c = header_map.get("landed_cost_freight_addon")
    q_c = header_map.get("quantity")
    ul_c = header_map.get("unit_landed_cost")
    if not (cv_c and bd_c and va_c and q_c and ul_c):
        return
    cv_l = get_column_letter(cv_c)
    bd_l = get_column_letter(bd_c)
    va_l = get_column_letter(va_c)
    q_l = get_column_letter(q_c)
    fa_l = get_column_letter(fa_c) if fa_c else ""
    first_data_row = header_row + 1
    for r in range(first_data_row, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").strip().upper() == "ИТОГО":
            continue
        duty_v = ws.cell(row=r, column=bd_c).value
        if isinstance(duty_v, str) and "СИНХРОН" in str(duty_v).upper():
            continue
        if fa_l:
            expr = f"=IF({q_l}{r}>0,({cv_l}{r}+{bd_l}{r}+{va_l}{r}+{fa_l}{r})/{q_l}{r},\"\")"
        else:
            expr = f"=IF({q_l}{r}>0,({cv_l}{r}+{bd_l}{r}+{va_l}{r})/{q_l}{r},\"\")"
        ws.cell(row=r, column=ul_c).value = expr


def invoice_checkpoint_paths(source_xlsx: Path, temp_results_dir: Path) -> tuple[Path, Path]:
    """Пары файлов: CSV со строками и JSON с метаданными для возобновления."""
    temp_results_dir.mkdir(parents=True, exist_ok=True)
    key = hashlib.md5(str(source_xlsx.resolve()).encode("utf-8")).hexdigest()[:16]
    base = temp_results_dir / f"checkpoint_{key}"
    return base.with_suffix(".csv"), base.with_suffix(".meta.json")


def save_invoice_checkpoint(
    rows: list[dict[str, Any]],
    *,
    csv_path: Path,
    meta_path: Path,
    meta: dict[str, Any],
) -> None:
    """Полная перезапись CSV и meta (атомарно по смыслу: сначала CSV)."""
    pd.DataFrame(rows).to_csv(csv_path, index=False, encoding="utf-8-sig")
    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")


def load_invoice_checkpoint_rows(csv_path: Path) -> list[dict[str, Any]]:
    if not csv_path.is_file() or csv_path.stat().st_size == 0:
        return []
    df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str)
    df = df.fillna("")
    return df.to_dict("records")


def load_invoice_checkpoint_meta(meta_path: Path) -> dict[str, Any] | None:
    if not meta_path.is_file():
        return None
    try:
        obj = json.loads(meta_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None
    return obj if isinstance(obj, dict) else None


def invoice_checkpoint_matches_source(
    meta: dict[str, Any],
    *,
    source: Path,
    max_rows: int,
    planned_total: int,
) -> bool:
    try:
        saved = Path(str(meta.get("source_path") or "")).resolve()
    except OSError:
        return False
    if saved != source.resolve():
        return False
    if int(meta.get("max_rows", -1)) != int(max_rows):
        return False
    if int(meta.get("planned_total", -1)) != int(planned_total):
        return False
    return True


def _reorder_invoice_columns_raw_first(df: pd.DataFrame) -> pd.DataFrame:
    """Все колонки ``[RAW] …`` в начале таблицы, затем расчётные/системные."""
    if df is None or df.empty:
        return df
    cols = list(df.columns)
    raw_cols = [c for c in cols if _is_raw_mirror_column(c)]
    rest = [c for c in cols if not _is_raw_mirror_column(c)]
    return df[raw_cols + rest]


def build_invoice_summary_row(out_df: pd.DataFrame) -> dict[str, Any]:
    """Одна строка «ИТОГО»: суммы по весу, стоимости и налогам (и вспомогательные числовые поля)."""
    cols = list(out_df.columns)
    row: dict[str, Any] = {c: "" for c in cols}
    if not cols:
        return row
    row[cols[0]] = "ИТОГО"
    for k in _INVOICE_SUMMARY_NUMERIC:
        if k not in out_df.columns:
            continue
        s = 0.0
        for v in out_df[k]:
            n = _parse_number(v)
            if n is not None:
                s += float(n)
        row[k] = round(s, 2) if s else ""
    return row


def write_invoice_report_excel(
    out_df: pd.DataFrame,
    out_path: Path,
    *,
    include_summary_row: bool = True,
) -> None:
    """
    Сохраняет отчёт в .xlsx: опциональная строка итогов, перенос текста, подсветка ячеек с «ВНИМАНИЕ»/«РОИС».
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df_out = _reorder_invoice_columns_raw_first(out_df.copy())
    if include_summary_row and not df_out.empty:
        summary = build_invoice_summary_row(df_out)
        df_out = pd.concat([df_out, pd.DataFrame([summary])], ignore_index=True)
    df_out.to_excel(out_path, index=False, engine="openpyxl")
    _apply_invoice_excel_workbook_styling(out_path, include_vat_baseline_note=True)


def _apply_invoice_excel_workbook_styling(
    xlsx_path: Path,
    *,
    include_vat_baseline_note: bool = True,
) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, PatternFill

    wb = load_workbook(xlsx_path)
    try:
        ws = wb.active
        header_row = 1
        if include_vat_baseline_note and ws.max_row >= 1:
            ws.insert_rows(1)
            last_col = max(1, int(ws.max_column))
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
            note_cell = ws.cell(row=1, column=1)
            note_cell.value = _invoice_excel_vat_baseline_note_text()
            note_cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws.row_dimensions[1].height = 48
            header_row = 2

        header_map: dict[str, int] = {}
        for cell in ws[header_row]:
            if cell.value is not None and str(cell.value).strip():
                header_map[str(cell.value).strip()] = int(cell.column)
        col_name_by_idx = {idx: name for name, idx in header_map.items()}

        warn_fill = PatternFill(start_color=_WARNING_FILL, end_color=_WARNING_FILL, fill_type="solid")
        vat10_fill = PatternFill(start_color=_VAT_10_FILL, end_color=_VAT_10_FILL, fill_type="solid")
        ai_conf_low_fill = PatternFill(
            start_color=_AI_CONFIDENCE_LOW_FILL,
            end_color=_AI_CONFIDENCE_LOW_FILL,
            fill_type="solid",
        )
        top_wrap = Alignment(wrap_text=True, vertical="top")

        vat_col_idx = header_map.get("vat_rate")
        ai_conf_col_idx = header_map.get("Уверенность ИИ (%)")

        first_data_row = header_row + 1
        for r in range(first_data_row, ws.max_row + 1):
            if str(ws.cell(row=r, column=1).value or "").strip().upper() == "ИТОГО":
                continue
            for col_idx in range(1, ws.max_column + 1):
                c = ws.cell(row=r, column=col_idx)
                hname = col_name_by_idx.get(col_idx, "")
                if hname in _INVOICE_WRAP_COLUMNS:
                    c.alignment = top_wrap
                v = c.value
                if v is None:
                    continue
                s = str(v)
                if "ВНИМАНИЕ" in s or "РОИС" in s:
                    c.fill = warn_fill
                if vat_col_idx == col_idx:
                    num = _parse_number(v)
                    if num is not None and abs(float(num) - 10.0) < 1e-6:
                        c.fill = vat10_fill
                if ai_conf_col_idx == col_idx:
                    cn = _parse_number(v)
                    if cn is not None and float(cn) < 80.0:
                        c.fill = ai_conf_low_fill

        try:
            _apply_invoice_excel_base_duty_formulas(ws, header_map, header_row)
            _apply_invoice_excel_vat_amount_formulas(ws, header_map, header_row)
            _apply_invoice_excel_unit_landed_formulas(ws, header_map, header_row)
        except Exception as e:
            logger.warning("Excel-формулы пошлины/НДС/landed cost не применены: {}", e)

        wb.save(xlsx_path)
    finally:
        wb.close()
