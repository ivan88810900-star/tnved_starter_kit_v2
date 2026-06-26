"""Экспорт классифицированного пакинг-листа в Excel."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

EXPORT_HEADERS = ("Перевод (RU)", "Код ТН ВЭД", "Уверенность %", "Обоснование")


def _header_row(data_start_row: int) -> int:
    return max(1, int(data_start_row) - 1)


def _confidence_pct(value: Any) -> str | None:
    if value is None:
        return None
    try:
        v = float(value)
    except (TypeError, ValueError):
        return str(value)
    if 0 <= v <= 1:
        return round(v * 100, 1)
    return round(v, 1)


def export_classified_packing_list(
    original_path: str | Path,
    *,
    results_by_row: dict[int, dict[str, Any]],
    col_map: dict[str, int],
    data_start_row: int,
    output_path: str | Path,
) -> Path:
    """
    Открывает оригинальный xlsx, вставляет 4 колонки после колонки «название»
    (или после последней заполненной, если name не найден).
    """
    original_path = Path(original_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(original_path)
    try:
        ws = wb.active
        name_col = col_map.get("name") or max(col_map.values(), default=ws.max_column or 1)
        insert_at = int(name_col) + 1
        ws.insert_cols(insert_at, amount=len(EXPORT_HEADERS))

        hdr_row = _header_row(data_start_row)
        for offset, title in enumerate(EXPORT_HEADERS):
            ws.cell(row=hdr_row, column=insert_at + offset, value=title)

        for row_num, data in results_by_row.items():
            ws.cell(row=row_num, column=insert_at, value=data.get("translation_used") or "")
            ws.cell(row=row_num, column=insert_at + 1, value=data.get("hs_code") or "")
            conf = _confidence_pct(data.get("hs_confidence"))
            ws.cell(row=row_num, column=insert_at + 2, value=conf if conf is not None else "")
            ws.cell(row=row_num, column=insert_at + 3, value=data.get("hs_rationale") or "")

        wb.save(output_path)
    finally:
        wb.close()
    return output_path
