from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill


RED_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _alerts_text(documents: list[dict[str, Any]]) -> str:
    if not documents:
        return "—"
    bits: list[str] = []
    for d in documents:
        status = str(d.get("compliance_status") or "REQUIRED")
        title = str(d.get("title") or d.get("doc_type") or "").strip()
        if not title:
            continue
        bits.append(f"[{status}] {title}")
    return "; ".join(bits)[:4000] if bits else "—"


def _is_critical_row(profile: dict[str, Any]) -> bool:
    if bool(profile.get("blocking_issue")):
        return True
    docs = list(profile.get("documents") or [])
    for d in docs:
        st = str(d.get("compliance_status") or "").upper()
        dt = str(d.get("doc_type") or "").upper()
        if st == "CRITICAL_RISK" or dt == "SANCTION_CONTROL":
            return True
    return False


def _is_matched_row(profile: dict[str, Any]) -> bool:
    docs = list(profile.get("documents") or [])
    for d in docs:
        st = str(d.get("compliance_status") or "").upper()
        if st == "MATCHED":
            return True
        if d.get("registry_match"):
            return True
    return False


def _is_warning_row(profile: dict[str, Any]) -> bool:
    docs = list(profile.get("documents") or [])
    for d in docs:
        st = str(d.get("compliance_status") or "").upper()
        if st == "WARNING":
            return True
    return False


def _duty_rate(profile: dict[str, Any], item: dict[str, Any]) -> Any:
    if item.get("duty_rate") is not None:
        return item.get("duty_rate")
    br = dict(profile.get("breakdown") or {})
    if br.get("duty_rate") is not None:
        return br.get("duty_rate")
    return "N/A"


def _vat_rate(profile: dict[str, Any], item: dict[str, Any]) -> Any:
    if item.get("vat_rate") is not None:
        return item.get("vat_rate")
    br = dict(profile.get("breakdown") or {})
    if br.get("vat_rate") is not None:
        return br.get("vat_rate")
    return "N/A"


def _compliance_status_text(profile: dict[str, Any]) -> str:
    docs = list(profile.get("documents") or [])
    if not docs:
        return "—"
    statuses = sorted({str(d.get("compliance_status") or "REQUIRED").upper() for d in docs})
    return ", ".join(statuses)


def generate_final_customs_excel(items: list[dict[str, Any]]) -> bytes:
    """
    Формирует XLSX для финального таможенного файла.

    expected item shape:
    {
      "item_description": str,
      "ai_technical_description": str | None,
      "payment_profile": PaymentProfileResponse.model_dump(),
      "duty_rate": float | None,
      "vat_rate": float | None
    }
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Customs Export"

    headers = [
        "Описание из инвойса",
        "Тех. описание ИИ (Vision)",
        "Код ТН ВЭД",
        "Ставка пошлины",
        "НДС",
        "Итого платежей",
        "Комплаенс-статус",
    ]
    ws.append(headers)

    for row_idx, it in enumerate(items, start=2):
        profile = dict(it.get("payment_profile") or {})
        breakdown = dict(profile.get("breakdown") or {})
        docs = list(profile.get("documents") or [])

        ws.append(
            [
                str(it.get("item_description") or ""),
                str(it.get("ai_technical_description") or ""),
                str(profile.get("hs_code") or ""),
                _duty_rate(profile, it),
                _vat_rate(profile, it),
                float(breakdown.get("total_payable") or 0.0),
                (_alerts_text(docs) + " | statuses: " + _compliance_status_text(profile))[:8000],
            ]
        )

        if _is_critical_row(profile):
            fill = RED_FILL
        elif _is_matched_row(profile):
            fill = GREEN_FILL
        elif _is_warning_row(profile):
            fill = YELLOW_FILL
        else:
            fill = None
        if fill is not None:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    for col in ("A", "B", "G"):
        ws.column_dimensions[col].width = 48
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 20

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def generate_final_customs_excel_from_ved_result(result: dict[str, Any]) -> bytes:
    """
    Адаптер: строит экспортный XLSX из payload результата ved-intelligent-analyze (job.result).
    """
    src_items = list((result or {}).get("items") or [])
    rows: list[dict[str, Any]] = []
    for it in src_items:
        if not isinstance(it, dict):
            continue
        profile = dict(it.get("payment_profile") or {})
        rows.append(
            {
                "item_description": str(it.get("name") or it.get("description") or ""),
                "ai_technical_description": str(it.get("ai_visual_description") or it.get("technical_description") or ""),
                "payment_profile": profile,
                "duty_rate": it.get("duty_rate"),
                "vat_rate": it.get("vat_rate") or it.get("vat_import_rate"),
            }
        )
    return generate_final_customs_excel(rows)

